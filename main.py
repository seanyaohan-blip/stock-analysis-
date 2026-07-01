#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
V2.8.5 ETF / A股统一投资决策仪表盘

功能说明（给非编程用户）：
1. 读取 watchlist.csv（观察池）和 positions.csv（持仓）
2. 用 AKShare 抓取指数、ETF、A股 的准实时行情
3. 生成 Excel 到 output/V2.8.5_每日行情输出_日期时间.xlsx
4. 只做纪律提醒，不连接券商、不下单、不自动交易

使用方法：
    第一步：pip install -r requirements.txt
    第二步：python main.py
    第三步：打开 output 文件夹里的 Excel 查看结果
"""

from __future__ import annotations

import json
import re
import subprocess
import time
import urllib.request
from datetime import datetime, time as dt_time
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import SeriesLabel
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook

try:
    import akshare as ak
except ImportError:
    ak = None

# ==================== 路径配置 ====================
BASE_DIR = Path(__file__).resolve().parent
WATCHLIST_CSV = BASE_DIR / "watchlist.csv"
POSITIONS_CSV = BASE_DIR / "positions.csv"
DECISION_INPUTS_CSV = BASE_DIR / "decision_inputs.csv"
FIRST_YEAR_ALLOCATION_CSV = BASE_DIR / "first_year_allocation.csv"
INVESTMENT_PROFILE_CSV = BASE_DIR / "investment_profile.csv"
DATA_SOURCES_CSV = BASE_DIR / "data_sources.csv"
BUY_POINT_PLAN_CSV = BASE_DIR / "buy_point_plan.csv"
OUTPUT_DIR = BASE_DIR / "output"
FRAMEWORK_VERSION = "V2.8.5"
FRAMEWORK_DISPLAY_VERSION = "V2.8.5-QM"
PRODUCT_NAME = "ETF / A股统一投资决策仪表盘"
LIVE_QUOTE_SOURCES = {"行情接口", "东方财富补齐", "腾讯行情补齐"}
MISSING_TEXT_VALUES = {"", "-", "--", "—", "nan", "none", "null", "n/a", "na"}
POSITIVE_MARKET_FIELDS = {"Latest", "Open", "High", "Low", "PrevClose", "Avg20Amount", "PrevDayLow"}
NUMERIC_MARKET_FIELDS = {"PctChg", "Amount"}
BUY_RANGE_PRICE_FIELDS = {
    "latest": "最新价",
    "high": "最高",
    "low": "最低",
    "prev_close": "昨收",
    "prev_day_low": "昨日低点",
}
TRUSTED_PREV_DAY_LOW_SOURCES = {"东方财富日K统计", "腾讯日K补齐"}

FRONT_SHEET_ORDER = [
    "01_今日决策",
    "02_今日动作",
    "03_买入候选",
    "04_阶段推进",
    "买点计划",
    "05_持仓风险",
    "06_年度配置",
    "07_组合总览",
]

DETAIL_SHEET_ORDER = [
    "Double_Anchor",
    "Emotion",
    "Quality_Score",
    "Exposure",
    "Market_Data",
    "Buy_Filter",
    "Positions",
    "Broker_Snapshot",
    "Watchlist",
    "长期跟踪个股",
    "Investment_Profile",
    "Data_Sources",
    "Buy_Point_Plan_Source",
    "Framework_Rules",
    "Checks",
    "使用说明",
]

# V2.8.5 的硬门槛集中放在这里，避免规则散落在代码里。
QUALITY_CORE_MIN = 8.0
QUALITY_OBSERVE_MIN = 6.0
EMOTION_PAUSE_LEVEL = 4
BUY_STANDARD_MIN = 5
BUY_HALF_MIN = 4
THEME_LIMITS = {
    "半导体/芯片设备": 0.18,
    "AI算力/云计算": 0.15,
    "机器人/自动化": 0.18,
    "新能源/绿电": 0.25,
    "创新药/生物医药": 0.15,
    "科技成长合计": 0.50,
}

MAINLINE_CYCLE_STAGES = "触底期→突破期→主线聚焦期→加速期→背驰上涨期→退潮期"
TREND_BUY_PRIORITY = "回踩买点 > 突破买点 > 止跌买点"
RETREAT_THREE_FACTORS = "主线证伪 × 资金退潮 × 估值透支"
VOLUME_ABSORPTION_SEQUENCE = "爆量突破→缩量回踩→不破关键支撑→再次放量转强"
V285_SUPPLEMENT_SOURCE = "V2.8.5 第十九部分：爆量与量质吸买点确认规则"
MENTAL_MODEL_SOURCE = "V2.8.5-QM：多元思维模型认知防错层"
QUANT_MODULE_SOURCE = "V2.8.5-Q：量化验证与执行纪律模块"
DIVIDEND_STOCK_SOURCE = "上传文本：长期稳定分红个股配置与高息陷阱规则"

DIVIDEND_CORE_STOCK_CODES = {"600900", "600941", "601398", "601939", "601088", "000333"}
DIVIDEND_ENHANCED_STOCK_CODES = {
    "601288", "601988", "600377", "600642", "600023",
    "601225", "600028", "000651", "600690",
}


def make_output_xlsx_path() -> Path:
    """每次运行都生成带日期时间后缀的新 Excel，方便保留历史记录。"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return OUTPUT_DIR / f"{FRAMEWORK_VERSION}_每日行情输出_{timestamp}.xlsx"

# ==================== 字段映射（AKShare 字段可能变化，这里做兼容） ====================
# 标准字段名 -> 可能出现的 AKShare 中文列名
FIELD_ALIASES: dict[str, list[str]] = {
    "symbol": ["代码", "code", "symbol"],
    "name": ["名称", "name"],
    "latest": ["最新价", "最新", "close", "现价"],
    "pct_chg": ["涨跌幅", "pct_chg"],
    "open": ["今开", "开盘价", "open"],
    "high": ["最高", "最高价", "high"],
    "low": ["最低", "最低价", "low"],
    "prev_close": ["昨收", "pre_close", "prev_close"],
    "amount": ["成交额", "amount"],
    "premium": ["基金折价率", "折溢价", "premium"],
    "volume_ratio": ["量比", "volume_ratio"],
}

TOTAL_ANCHORS = [
    ("000001.SH", "上证综指"),
    ("000300.SH", "沪深300"),
    ("000510.CSI", "中证A500"),
]

GROWTH_ANCHORS = [
    ("000688.SH", "科创50"),
    ("399006.SZ", "创业板指"),
]

BUY_FILTER_COLUMNS = [
    "买点灯号", "Code", "Name", "建议买入区间", "Role", "质量评分", "质量状态", "情绪温度", "反馈状态",
    "趋势结构", "关键支撑纪律", "单次仓位上限", "爆量/量质吸纪律",
    "机会成本检查", "认知防错", "决策树情景", "赔率/最大损失", "反向失败路径", "能力圈边界",
    "资产角色", "量化验证分", "量化验证状态", "量化买点规则", "正期望检查", "量化风控规则",
    "红利配置角色", "分红质量检查", "红利买点规则", "高息陷阱否决",
    "当前仓位", "目标仓位", "剩余额度", "仓位状态",
    "阶段状态", "阶段目标下限", "阶段目标上限", "阶段占用率", "阶段验证", "阶段推进结论", "下一档单次上限", "阶段阻断原因", "阶段动作说明",
    "仓位决策口径", "全资产当前仓位", "全资产目标仓位", "全资产剩余额度", "全资产仓位状态",
    "证券账户仓位", "证券账户目标仓位", "证券账户仓位状态",
    "Latest", "PctChg", "Open", "High", "Low", "PrevClose", "PrevDayLow", "PrevDayLowSource",
    "QuoteTime", "DataSource", "接口状态", "失败原因", "重试结果", "评分输入状态", "评分缺失项",
    "日内位置", "量能倍数", "分时结构", "量价关系", "份额变动", "折溢价",
    "龙头同步", "次日验证", "通过项", "通过明细", "未通过项", "待确认项", "一票否决", "否决原因", "建议",
]


def format_code(code: str) -> str:
    """统一代码格式，避免 Excel 把 002594 变成 2594。"""
    code = str(code).strip()
    upper = code.upper()
    if upper.endswith((".SH", ".SZ", ".CSI")):
        return upper
    pure = code.split(".")[0]
    if pure.isdigit():
        return pure.zfill(6)
    return code


def to_float(value) -> float | None:
    """安全转成浮点数，空值或异常值返回 None。"""
    if value is None or pd.isna(value):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def is_recent_complete_quote(quote_time, now: datetime | None = None) -> bool:
    """行情必须有有效时间，且不得早于当前日期三个自然日。"""
    if quote_time is None or str(quote_time).strip() in {"", "nan", "None"}:
        return False
    try:
        parsed = pd.to_datetime(quote_time, errors="raise")
        quote_dt = parsed.to_pydatetime() if hasattr(parsed, "to_pydatetime") else parsed
    except (TypeError, ValueError):
        return False
    current = now or datetime.now()
    age_days = (current.date() - quote_dt.date()).days
    return 0 <= age_days <= 3


def price_tick(code: str, name: str) -> float:
    """ETF 使用千分位报价，普通 A 股使用分位报价。"""
    code_text = format_code(code).split(".")[0]
    if "ETF" in str(name).upper() or code_text.startswith(("15", "16", "51", "56", "58")):
        return 0.001
    return 0.01


def round_to_tick(value: float, tick: float) -> float:
    """按证券最小价格变动单位四舍五入。"""
    decimal_tick = Decimal(str(tick))
    rounded_units = (Decimal(str(value)) / decimal_tick).quantize(Decimal("1"), rounding=ROUND_HALF_UP)
    return float(rounded_units * decimal_tick)


def price_decimals(code: str, name: str) -> int:
    """根据最小报价单位决定表格展示精度。"""
    return 3 if price_tick(code, name) == 0.001 else 2


def format_price_value(value, code: str, name: str) -> str:
    """把价格格式化为与 ETF/A股报价单位一致的文本。"""
    numeric = to_float(value)
    if numeric is None:
        return ""
    decimals = price_decimals(code, name)
    return f"{numeric:.{decimals}f}"


def format_price_range(price_range: tuple[float, float], code: str, name: str) -> str:
    """格式化买点/观察区间。"""
    decimals = price_decimals(code, name)
    lower, upper = price_range
    return f"{lower:.{decimals}f}–{upper:.{decimals}f}"


def calculate_buy_range(
    code: str,
    name: str,
    signal: str,
    latest,
    high,
    low,
    prev_close,
    prev_day_low,
) -> tuple[float, float] | None:
    """按 V2.8.5 的回踩、不追高和昨日低点纪律计算复核区间。"""
    values = [to_float(item) for item in (latest, high, low, prev_close, prev_day_low)]
    if any(item is None for item in values) or signal not in {"绿", "黄"}:
        return None
    latest_f, high_f, low_f, prev_close_f, prev_day_low_f = values
    intraday_range = high_f - low_f
    if intraday_range <= 0 or min(values) <= 0:
        return None

    if signal == "绿":
        lower = max(prev_day_low_f, low_f + intraday_range * 0.25)
        upper = min(latest_f, prev_close_f * 1.005, low_f + intraday_range * 0.50)
    else:
        lower = max(prev_day_low_f, low_f + intraday_range * 0.15)
        upper = min(latest_f * 0.995, prev_close_f, low_f + intraday_range * 0.35)

    tick = price_tick(code, name)
    lower_rounded = round_to_tick(lower, tick)
    upper_rounded = round_to_tick(upper, tick)
    if upper_rounded < lower_rounded:
        return None
    return lower_rounded, upper_rounded


def calculate_reclaim_prev_low_range(
    code: str,
    name: str,
    prev_day_low,
) -> tuple[float, float] | None:
    """跌破昨日低点后，给出重新站稳后的窄观察区间。"""
    prev_low_f = to_float(prev_day_low)
    if prev_low_f is None or prev_low_f <= 0:
        return None
    tick = price_tick(code, name)
    lower = round_to_tick(prev_low_f, tick)
    upper = round_to_tick(max(prev_low_f * 1.005, prev_low_f + tick), tick)
    if upper <= lower:
        upper = round_to_tick(lower + tick, tick)
    return lower, upper


def calculate_reference_buy_range(
    code: str,
    name: str,
    latest,
    high,
    low,
    prev_close,
    prev_day_low,
) -> tuple[float, float] | None:
    """在暂不建议买入时，给出不可执行的观察区间，避免只输出纯文字。"""
    latest_f = to_float(latest)
    prev_low_f = to_float(prev_day_low)
    if latest_f is not None and prev_low_f is not None and latest_f < prev_low_f:
        return calculate_reclaim_prev_low_range(code, name, prev_day_low)

    price_range = calculate_buy_range(
        code=code,
        name=name,
        signal="黄",
        latest=latest,
        high=high,
        low=low,
        prev_close=prev_close,
        prev_day_low=prev_day_low,
    )
    if price_range is not None:
        return price_range

    if latest_f is None or latest_f <= 0:
        return None
    tick = price_tick(code, name)
    lower = round_to_tick(max(latest_f * 0.985, tick), tick)
    upper = round_to_tick(max(latest_f * 0.995, lower + tick), tick)
    if upper < lower:
        upper = round_to_tick(lower + tick, tick)
    return lower, upper


def first_valid(*values):
    """返回第一个非空值，用于行情缺失时使用券商截图数据补充。"""
    for value in values:
        if value is None:
            continue
        if isinstance(value, str):
            text = value.strip()
            if text.lower() in MISSING_TEXT_VALUES:
                continue
        try:
            if pd.isna(value):
                continue
        except TypeError:
            pass
        return value
    return None


def has_market_field_value(column: str, value) -> bool:
    """Return whether a market-data field is usable for freshness diagnostics."""
    if first_valid(value) is None:
        return False
    if column in POSITIVE_MARKET_FIELDS:
        numeric = to_float(value)
        return numeric is not None and numeric > 0
    if column in NUMERIC_MARKET_FIELDS:
        return to_float(value) is not None
    return True


def missing_latest_mask(series: pd.Series) -> pd.Series:
    """Latest must be a positive numeric price; '-' and zero are incomplete quotes."""
    numeric = pd.to_numeric(series, errors="coerce")
    return numeric.isna() | (numeric <= 0)


def append_note(existing, note: str) -> str:
    text = str(existing or "").strip()
    if not text or text.lower() == "nan":
        return note
    if note in text:
        return text
    return f"{text}；{note}"


def describe_weight_status(current_weight: float | None, target_weight: float | None, label: str = "目标") -> str:
    """统一描述仓位是否低于/达到/超过目标，避免不同表格口径不一致。"""
    if target_weight is None:
        return f"未设置{label}仓位"
    if current_weight is None:
        return "当前仓位缺失，需人工确认"
    gap = target_weight - current_weight
    if gap < 0:
        return f"超{label} {abs(gap):.2%}"
    if gap == 0:
        return f"已达{label}仓位"
    return f"低于{label}，剩余 {gap:.2%}"


def remove_note_phrases(existing, phrases: list[str]) -> str:
    text = str(existing or "").strip()
    if not text or text.lower() == "nan":
        return ""
    for phrase in phrases:
        text = text.replace(f"；{phrase}", "").replace(phrase, "")
    return text.strip("；")


def missing_market_field_labels(row, fields: dict[str, str]) -> list[str]:
    """Return human-readable labels for missing fields in a market-data row."""
    missing: list[str] = []
    for column, label in fields.items():
        try:
            value = row.get(column)
        except AttributeError:
            value = None
        if not has_market_field_value(column, value):
            missing.append(label)
    return missing


def missing_buy_range_price_labels(
    latest,
    high,
    low,
    prev_close,
    prev_day_low,
) -> list[str]:
    """买入区间必须使用完整且为正的价格字段。"""
    values = {
        "latest": latest,
        "high": high,
        "low": low,
        "prev_close": prev_close,
        "prev_day_low": prev_day_low,
    }
    missing: list[str] = []
    for key, value in values.items():
        numeric = to_float(value)
        if numeric is None or numeric <= 0:
            missing.append(BUY_RANGE_PRICE_FIELDS[key])
    return missing


def buy_range_data_issue(
    *,
    latest,
    high,
    low,
    prev_close,
    prev_day_low,
    quote_time,
    data_source: str,
    prev_day_low_source: str = "",
    now: datetime | None = None,
) -> str:
    """Return the explicit data issue that blocks a numeric buy range."""
    issues: list[str] = []
    source = str(data_source or "").strip()
    if source not in LIVE_QUOTE_SOURCES:
        issues.append("行情来源")
    if not is_recent_complete_quote(quote_time, now=now):
        issues.append("行情时间")
    issues.extend(missing_buy_range_price_labels(latest, high, low, prev_close, prev_day_low))

    prev_source = str(prev_day_low_source or "").strip()
    if prev_source and prev_source not in TRUSTED_PREV_DAY_LOW_SOURCES:
        issues.append("昨日低点来源")

    return "、".join(dict.fromkeys(issues))


def clean_blocker_reason(value) -> str:
    text = str(value or "").strip()
    if text == "无" or text.lower() in MISSING_TEXT_VALUES:
        return ""
    return text


def quote_data_blocker_reason(row, range_issue: str = "") -> str:
    issue = str(range_issue or "").strip()
    if issue:
        return f"行情未刷新或数据不完整：{issue}"

    status = clean_blocker_reason(row.get("接口状态", ""))
    failure_reason = clean_blocker_reason(row.get("失败原因", ""))
    if status and re.search(r"未刷新|缺失|失败", status):
        return f"{status}：{failure_reason}" if failure_reason else status

    latest_missing = to_float(row.get("Latest")) is None
    pct_missing = to_float(row.get("PctChg")) is None
    if latest_missing or pct_missing:
        missing = []
        if latest_missing:
            missing.append("最新价")
        if pct_missing:
            missing.append("涨跌幅")
        return f"行情未刷新或数据不完整：{'、'.join(missing)}"

    return ""


def build_market_interface_diagnostics(row, now: datetime | None = None) -> dict[str, str]:
    """Summarize live quote freshness and retry outcome for one Market_Data row."""
    source = str(row.get("DataSource", "") or "").strip()
    quote_time = row.get("QuoteTime")
    latest = first_valid(row.get("Latest"))
    quote_missing = missing_market_field_labels(
        row,
        {
            "Latest": "最新价",
            "PctChg": "涨跌幅",
            "Open": "今开",
            "High": "最高",
            "Low": "最低",
            "PrevClose": "昨收",
            "Amount": "成交额",
            "QuoteTime": "行情时间",
        },
    )
    stats_missing = missing_market_field_labels(
        row,
        {
            "Avg20Amount": "20日均额",
            "PrevDayLow": "昨日低点",
        },
    )
    avg20_source = str(row.get("Avg20AmountSource", "") or "").strip()
    prev_day_low_source = str(row.get("PrevDayLowSource", "") or "").strip()
    if first_valid(row.get("Avg20Amount")) is not None and "替代" in avg20_source:
        stats_missing.append("20日均额来源")
    if first_valid(row.get("PrevDayLow")) is not None and prev_day_low_source and prev_day_low_source not in TRUSTED_PREV_DAY_LOW_SOURCES:
        stats_missing.append("昨日低点来源")
    recent_live_quote = source in LIVE_QUOTE_SOURCES and is_recent_complete_quote(quote_time, now=now)

    if recent_live_quote and not quote_missing and not stats_missing:
        status = "当日行情已刷新"
        failure_reason = "无"
    elif recent_live_quote and not quote_missing:
        status = "行情字段缺失"
        failure_reason = "实时行情已返回但买入区间或量能字段不完整"
    elif source in LIVE_QUOTE_SOURCES:
        status = "行情字段缺失"
        failure_reason = "实时行情已返回但字段不完整"
    elif source == "券商截图" and latest is not None:
        status = "行情未刷新，券商截图兜底"
        failure_reason = "实时行情接口未返回当日完整数据；仅用券商截图价格作持仓市值参考"
    else:
        status = "行情未刷新"
        failure_reason = "AKShare实时行情未命中；东方财富/腾讯当日行情补齐未返回"

    if quote_missing and failure_reason != "无":
        failure_reason = append_note(failure_reason, "缺失字段：" + "、".join(quote_missing))
    elif quote_missing:
        failure_reason = "缺失字段：" + "、".join(quote_missing)

    if stats_missing:
        retry_result = "日K首轮失败后已重试，仍缺：" + "、".join(stats_missing)
    elif recent_live_quote and not quote_missing:
        retry_result = "实时行情与日K补齐可用"
    else:
        retry_result = "实时行情补齐已尝试，未取得当日完整行情"

    return {
        "接口状态": status,
        "失败原因": failure_reason,
        "重试结果": retry_result,
    }


def attach_market_interface_diagnostics(market_data: pd.DataFrame) -> pd.DataFrame:
    """Add structured interface diagnostics used by dashboard and audit sheets."""
    if market_data.empty:
        return market_data
    out = market_data.copy()
    for column in ("接口状态", "失败原因", "重试结果"):
        if column not in out.columns:
            out[column] = ""
    for idx, row in out.iterrows():
        diagnostics = build_market_interface_diagnostics(row)
        for column, value in diagnostics.items():
            out.at[idx, column] = value
    return out


def print_actual_columns(source_name: str, df: pd.DataFrame) -> None:
    """当字段对不上时，打印实际字段名，方便排错。"""
    print(f"\n【字段排错】接口：{source_name}")
    print(f"【实际字段】{list(df.columns)}")


def normalize_akshare_df(df: pd.DataFrame, source_name: str) -> pd.DataFrame:
    """
    把 AKShare 返回的数据列名，统一成英文标准名。
    如果缺少关键字段，会打印实际字段名，不会静默失败。
    """
    rename_map: dict[str, str] = {}
    missing: list[str] = []

    for standard_name, aliases in FIELD_ALIASES.items():
        hit = next((alias for alias in aliases if alias in df.columns), None)
        if hit:
            rename_map[hit] = standard_name
        else:
            missing.append(standard_name)

    if missing:
        print(f"\n【字段提示】{source_name} 缺少这些标准字段：{missing}")
        print_actual_columns(source_name, df)

    out = df.rename(columns=rename_map).copy()

    if "symbol" in out.columns:
        out["symbol"] = out["symbol"].astype(str).str.strip()

    return out


# ==================== 读取 CSV ====================
def load_watchlist() -> pd.DataFrame:
    """读取观察池 CSV。"""
    if not WATCHLIST_CSV.exists():
        raise FileNotFoundError(f"找不到观察池文件：{WATCHLIST_CSV}")

    df = pd.read_csv(WATCHLIST_CSV, dtype={"Code": str}, encoding="utf-8-sig")
    df["Code"] = df["Code"].astype(str).str.strip().map(format_code)
    return df


def load_decision_inputs() -> dict[str, dict[str, str]]:
    """读取少量人工决策输入；空白值不会被当作有效证据。"""
    if not DECISION_INPUTS_CSV.exists():
        return {}
    frame = pd.read_csv(DECISION_INPUTS_CSV, dtype=str, encoding="utf-8-sig").fillna("")
    rows: dict[str, dict[str, str]] = {}
    for _, row in frame.iterrows():
        key = str(row.get("Key", "")).strip()
        if key:
            rows[key] = {column: str(row.get(column, "")).strip() for column in frame.columns}
    return rows


def load_optional_csv(path: Path, expected_columns: list[str]) -> pd.DataFrame:
    """读取可选配置表；缺失时返回带表头的空表，避免主流程中断。"""
    if not path.exists():
        return pd.DataFrame(columns=expected_columns)
    frame = pd.read_csv(path, dtype=str, encoding="utf-8-sig").fillna("")
    for column in expected_columns:
        if column not in frame.columns:
            frame[column] = ""
    return frame[expected_columns]


def build_investment_profile(decision_inputs: dict[str, dict[str, str]]) -> pd.DataFrame:
    """汇总投资偏好与纪律设置，写入每日输出方便复核。"""
    profile = load_optional_csv(INVESTMENT_PROFILE_CSV, ["Section", "Key", "Value", "Notes"])
    if profile.empty:
        return profile
    profile = profile.copy()
    overrides = {
        "cash_buffer_min": decision_input_number(decision_inputs, "cash_buffer_min"),
        "cash_buffer_target": decision_input_number(decision_inputs, "cash_buffer_target"),
        "emotion_temperature": decision_input_number(decision_inputs, "emotion_temperature"),
        "mainline_stage": decision_input_text(decision_inputs, "mainline_stage"),
        "trend_buy_type": decision_input_text(decision_inputs, "trend_buy_type"),
        "quant_module_status": decision_input_text(decision_inputs, "quant_module_status"),
        "retreat_factor_status": decision_input_text(decision_inputs, "retreat_factor_status"),
    }
    profile["当前生效值"] = [
        overrides.get(str(row.get("Key", "")).strip()) if overrides.get(str(row.get("Key", "")).strip()) is not None else row.get("Value", "")
        for _, row in profile.iterrows()
    ]
    profile["维护文件"] = "investment_profile.csv"
    profile.loc[
        profile["Key"].isin([
            "cash_buffer_min",
            "cash_buffer_target",
            "emotion_temperature",
            "mainline_stage",
            "trend_buy_type",
            "quant_module_status",
            "retreat_factor_status",
        ]),
        "维护文件",
    ] = "decision_inputs.csv"
    return profile


def build_data_sources() -> pd.DataFrame:
    """汇总行情、持仓、配置和盘前报告的数据来源。"""
    return load_optional_csv(
        DATA_SOURCES_CSV,
        ["Module", "DataType", "PrimarySource", "FallbackSource", "FreshnessRule", "FailureBehavior", "ConfigFile"],
    )


def load_buy_point_plan() -> pd.DataFrame:
    """读取人工维护的未来买点计划，用于输出可执行前的价格区间和条件。"""
    columns = [
        "Code", "Name", "资产角色", "当前价参考",
        "轻仓观察区间", "标准买点区间", "强买点区间", "不买/暂停区",
        "下一笔合理金额", "成立条件", "风险暂停条件", "来源日期",
    ]
    frame = load_optional_csv(BUY_POINT_PLAN_CSV, columns)
    if frame.empty:
        return frame
    frame = frame.copy()
    frame["Code"] = frame["Code"].astype(str).str.strip().map(format_code)
    return frame


def decision_input_number(inputs: dict[str, dict[str, str]], key: str) -> float | None:
    item = inputs.get(key, {})
    return to_float(item.get("Value"))


def decision_input_text(inputs: dict[str, dict[str, str]], key: str) -> str | None:
    text = str(inputs.get(key, {}).get("Value", "") or "").strip()
    return text or None


def load_first_year_allocation() -> pd.DataFrame:
    """读取第一年全资产配置目标；比例和金额必须保持数值格式。"""
    if not FIRST_YEAR_ALLOCATION_CSV.exists():
        return pd.DataFrame()
    frame = pd.read_csv(FIRST_YEAR_ALLOCATION_CSV, dtype={"Codes": str}, encoding="utf-8-sig").fillna("")
    for column in [
        "TargetWeight", "SourceTargetAmount", "SourceCurrentAmount",
        "TargetReturnLow", "TargetReturnHigh",
    ]:
        frame[column] = pd.to_numeric(frame[column], errors="coerce")
    frame["Codes"] = frame["Codes"].astype(str).str.strip()
    frame["SourceAsOf"] = frame.get("SourceAsOf", "").replace("", "原表未标注")
    return frame


def load_positions() -> tuple[dict[str, float], pd.DataFrame]:
    """
    读取持仓 CSV。
    第一行 # 开头的是账户汇总信息，例如总资产。
    """
    if not POSITIONS_CSV.exists():
        raise FileNotFoundError(f"找不到持仓文件：{POSITIONS_CSV}")

    meta: dict[str, float] = {}
    with POSITIONS_CSV.open(encoding="utf-8-sig") as f:
        first_line = f.readline().strip()
        if first_line.startswith("#"):
            parts = [p.strip() for p in first_line.lstrip("#").split(",")]
            for i in range(0, len(parts) - 1, 2):
                key = parts[i]
                try:
                    meta[key] = float(parts[i + 1])
                except ValueError:
                    pass

    positions = pd.read_csv(POSITIONS_CSV, comment="#", dtype={"Code": str}, encoding="utf-8-sig")
    positions["Code"] = positions["Code"].astype(str).str.strip().map(format_code)
    return meta, positions


# ==================== 代码分类 ====================
def is_index_code(code: str) -> bool:
    """判断是否为指数代码，例如 000001.SH、399006.SZ。"""
    code = code.upper()
    return code.endswith((".SH", ".SZ", ".CSI"))


def normalize_index_symbol(code: str) -> str:
    """把 000001.SH 转成 akshare 指数表里常见的 sh000001 形式。"""
    code = code.upper()
    if code.endswith(".SH"):
        return f"sh{code[:-3]}"
    if code.endswith(".SZ"):
        return f"sz{code[:-3]}"
    if code.endswith(".CSI"):
        return code[:-4].lower()
    return code.lower()


def normalize_index_symbols(code: str) -> set[str]:
    """返回指数代码可能出现的多种接口格式。"""
    code = code.upper()
    symbols = {normalize_index_symbol(code)}
    if code.endswith(".SH"):
        pure = code[:-3]
        symbols.add(pure)
        symbols.add(f"sh{pure}")
    elif code.endswith(".SZ"):
        pure = code[:-3]
        symbols.add(pure)
        symbols.add(f"sz{pure}")
    elif code.endswith(".CSI"):
        pure = code[:-4]
        symbols.add(pure)
        symbols.add(f"sh{pure}")
    return symbols


def classify_code(code: str) -> str:
    """
    判断代码类型：
    - index：指数
    - etf：ETF（常见 1/5 开头 6 位代码）
    - stock：A股股票
    """
    if is_index_code(code):
        return "index"

    pure = code.split(".")[0]
    if pure.isdigit() and len(pure) == 6 and pure.startswith(("1", "5")):
        return "etf"
    if pure.isdigit() and len(pure) == 6:
        return "stock"
    return "unknown"


# ==================== 抓取行情 ====================
def safe_fetch(source_name: str, fetch_func) -> pd.DataFrame:
    """统一抓取入口：失败时给出清晰中文提示。"""
    print(f"正在抓取：{source_name} ...")
    try:
        df = fetch_func()
        df = normalize_akshare_df(df, source_name)
        print(f"  完成，共 {len(df)} 条。")
        return df
    except Exception as exc:
        print(f"【抓取失败】{source_name}：{exc}")
        print("  你可以稍后重试，或检查网络/代理设置。")
        return pd.DataFrame()


def fetch_all_spot() -> dict[str, pd.DataFrame]:
    if ak is None:
        print("未安装 AKShare，将使用东方财富兜底行情补齐。")
        return {"index": pd.DataFrame(), "etf": pd.DataFrame(), "stock": pd.DataFrame()}

    """分别抓取指数、ETF、A股 三组行情。"""
    index_df = safe_fetch("指数行情 stock_zh_index_spot_em", ak.stock_zh_index_spot_em)
    if index_df.empty:
        index_df = safe_fetch("指数行情 stock_zh_index_spot_sina", ak.stock_zh_index_spot_sina)

    return {
        "index": index_df,
        "etf": safe_fetch("ETF行情 fund_etf_spot_em", ak.fund_etf_spot_em),
        "stock": safe_fetch("A股行情 stock_zh_a_spot_em", ak.stock_zh_a_spot_em),
    }


def eastmoney_secid(code: str) -> str:
    code = format_code(code)
    if code == "000001.SH":
        return "1.000001"
    if code == "000300.SH":
        return "1.000300"
    if code == "000510.CSI":
        return "1.000510"
    if code == "000688.SH":
        return "1.000688"
    if code == "399006.SZ":
        return "0.399006"
    pure = code.split(".")[0]
    market = "0" if pure.startswith(("0", "1", "2", "3")) else "1"
    return f"{market}.{pure}"


def fetch_json(url: str, retries: int = 3) -> dict:
    headers = {
        "User-Agent": "Mozilla/5.0",
        "Referer": "https://quote.eastmoney.com/",
    }
    last_exc: Exception | None = None
    for attempt in range(retries):
        try:
            req = urllib.request.Request(url, headers=headers)
            with urllib.request.urlopen(req, timeout=25) as resp:
                return json.loads(resp.read().decode("utf-8"))
        except Exception as exc:
            last_exc = exc
            if attempt < retries - 1:
                time.sleep(0.8 * (attempt + 1))
    for curl_binary in ("curl.exe", "curl"):
        try:
            result = subprocess.run(
                [
                    curl_binary,
                    "-sS",
                    "-L",
                    "--max-time",
                    "25",
                    "-A",
                    headers["User-Agent"],
                    "-e",
                    headers["Referer"],
                    url,
                ],
                check=True,
                capture_output=True,
                text=True,
                encoding="utf-8",
            )
            return json.loads(result.stdout)
        except Exception:
            continue
    raise last_exc if last_exc else RuntimeError("empty response")


def fetch_text(url: str, encoding: str = "utf-8", retries: int = 3) -> str:
    """Fetch a text endpoint with retries and an explicit response encoding."""
    headers = {
        "User-Agent": "Mozilla/5.0",
        "Referer": "https://gu.qq.com/",
    }
    last_exc: Exception | None = None
    for attempt in range(retries):
        try:
            req = urllib.request.Request(url, headers=headers)
            with urllib.request.urlopen(req, timeout=25) as resp:
                return resp.read().decode(encoding, errors="replace")
        except Exception as exc:
            last_exc = exc
            if attempt < retries - 1:
                time.sleep(0.8 * (attempt + 1))
    raise last_exc if last_exc else RuntimeError("empty response")


def tencent_symbol(code: str) -> str:
    """Convert a project code to Tencent's sh/sz quote symbol."""
    code = format_code(code)
    pure = code.split(".")[0]
    if code.endswith((".SH", ".CSI")):
        return f"sh{pure}"
    if code.endswith(".SZ"):
        return f"sz{pure}"
    if pure.startswith(("0", "1", "2", "3")):
        return f"sz{pure}"
    return f"sh{pure}"


def fetch_tencent_quotes(codes: list[str]) -> dict[str, dict]:
    """Fetch today's quotes from Tencent when Eastmoney's spot endpoint fails."""
    if not codes:
        return {}

    symbols = [tencent_symbol(code) for code in codes]
    symbol_to_code = {tencent_symbol(code): format_code(code) for code in codes}
    url = "https://qt.gtimg.cn/q=" + ",".join(symbols)
    try:
        text = fetch_text(url, encoding="gb18030")
    except Exception as exc:
        print(f"【腾讯行情补齐失败】{exc}")
        return {}

    today = datetime.now().strftime("%Y%m%d")
    out: dict[str, dict] = {}
    stale_symbols: list[str] = []
    for line in text.splitlines():
        if "=" not in line:
            continue
        key, raw = line.split("=", 1)
        symbol = key.strip().removeprefix("v_").lower()
        fields = raw.strip().rstrip(";").strip('"').split("~")
        if len(fields) <= 37:
            continue
        quote_time_raw = fields[30].strip()
        if not quote_time_raw.startswith(today):
            stale_symbols.append(symbol)
            continue
        try:
            quote_time = datetime.strptime(quote_time_raw, "%Y%m%d%H%M%S").strftime("%Y-%m-%d %H:%M:%S")
        except ValueError:
            quote_time = quote_time_raw

        amount = to_float(fields[37])
        code = symbol_to_code.get(symbol, format_code(fields[2]))
        out[code] = {
            "name": fields[1].strip(),
            "latest": to_float(fields[3]),
            "pct_chg": to_float(fields[32]),
            "open": to_float(fields[5]),
            "high": to_float(fields[33]),
            "low": to_float(fields[34]),
            "prev_close": to_float(fields[4]),
            "amount": amount * 10000 if amount is not None else None,
            "quote_time": quote_time,
        }

    if stale_symbols:
        print(f"【腾讯行情陈旧，已拒绝】{stale_symbols}")
    print(f"腾讯行情补齐：取得 {len(out)}/{len(codes)} 条，数据日期 {today}。")
    return out


def fetch_tencent_kline_stats(code: str, current_amount) -> dict:
    """Use Tencent daily volume to estimate an equivalent 20-day amount baseline."""
    symbol = tencent_symbol(code)
    url = (
        "https://web.ifzq.gtimg.cn/appstock/app/fqkline/get"
        f"?param={symbol},day,,,21,qfq"
    )
    try:
        data = fetch_json(url).get("data", {}).get(symbol, {})
        rows = data.get("day") or data.get("qfqday") or []
        rows = rows[-20:]
        volumes = [float(row[5]) for row in rows if len(row) > 5 and to_float(row[5]) is not None]
        current_volume = to_float(rows[-1][5]) if rows and len(rows[-1]) > 5 else None
        amount_value = to_float(current_amount)
        avg20_amount = None
        if volumes and current_volume and current_volume > 0 and amount_value is not None:
            avg20_volume = sum(volumes) / len(volumes)
            avg20_amount = round(amount_value * avg20_volume / current_volume, 2)
        prev_day_low = to_float(rows[-2][4]) if len(rows) >= 2 and len(rows[-2]) > 4 else None
        return {
            "Avg20Amount": avg20_amount,
            "Avg20AmountSource": "腾讯日K量能等价估算",
            "PrevDayLow": prev_day_low,
            "PrevDayLowSource": "腾讯日K补齐",
        }
    except Exception as exc:
        print(f"【腾讯日K补齐失败】{code}：{exc}")
        return {
            "Avg20Amount": None,
            "Avg20AmountSource": None,
            "PrevDayLow": None,
            "PrevDayLowSource": None,
        }


def fetch_eastmoney_quotes(codes: list[str]) -> dict[str, dict]:
    if not codes:
        return {}
    fields = "f12,f14,f2,f3,f6,f15,f16,f17,f18,f124"
    url = (
        "https://push2.eastmoney.com/api/qt/ulist.np/get"
        f"?fltt=2&invt=2&fields={fields}&secids="
        + ",".join(eastmoney_secid(code) for code in codes)
    )
    try:
        data = fetch_json(url).get("data", {}).get("diff", [])
    except Exception as exc:
        print(f"【东方财富行情补齐失败】{exc}")
        return {}

    today = datetime.now().strftime("%Y%m%d")
    out: dict[str, dict] = {}
    for item in data:
        code = format_code(item.get("f12"))
        if code == "000001":
            code = "000001.SH"
        elif code == "000300":
            code = "000300.SH"
        elif code == "000510":
            code = "000510.CSI"
        elif code == "000688":
            code = "000688.SH"
        elif code == "399006":
            code = "399006.SZ"
        quote_time = None
        quote_timestamp = to_float(item.get("f124"))
        if quote_timestamp is not None:
            quote_time = datetime.fromtimestamp(quote_timestamp).strftime("%Y-%m-%d %H:%M:%S")
            if not quote_time.replace("-", "").startswith(today):
                print(f"【东方财富行情陈旧，已拒绝】{code} {quote_time}")
                continue
        out[code] = {
            "name": item.get("f14"),
            "latest": item.get("f2"),
            "pct_chg": item.get("f3"),
            "open": item.get("f17"),
            "high": item.get("f15"),
            "low": item.get("f16"),
            "prev_close": item.get("f18"),
            "amount": item.get("f6"),
            "quote_time": quote_time,
        }
    return out


def fetch_eastmoney_kline_stats(code: str) -> dict:
    url = (
        "https://push2his.eastmoney.com/api/qt/stock/kline/get"
        f"?secid={eastmoney_secid(code)}&fields1=f1,f2,f3,f4,f5,f6"
        "&fields2=f51,f52,f53,f54,f55,f56,f57&klt=101&fqt=1&end=20500101&lmt=21"
    )
    try:
        klines = fetch_json(url).get("data", {}).get("klines") or []
        rows = [line.split(",") for line in klines]
        amounts = [float(row[6]) for row in rows[-20:] if len(row) > 6]
        prev_day_low = float(rows[-2][4]) if len(rows) >= 2 and len(rows[-2]) > 4 else None
        return {
            "Avg20Amount": round(sum(amounts) / len(amounts), 2) if amounts else None,
            "Avg20AmountSource": "东方财富日K统计",
            "PrevDayLow": prev_day_low,
            "PrevDayLowSource": "东方财富日K统计",
        }
    except Exception:
        return {
            "Avg20Amount": None,
            "Avg20AmountSource": None,
            "PrevDayLow": None,
            "PrevDayLowSource": None,
        }


def enrich_market_data_with_eastmoney(market_data: pd.DataFrame) -> pd.DataFrame:
    """Fill missing quote fields, 20-day average amount, and previous-day low."""
    if market_data.empty or "Code" not in market_data.columns:
        return market_data

    out = market_data.copy()
    for mixed_column in ("ETFShareChg", "Premium", "LeaderStatus"):
        if mixed_column in out.columns:
            out[mixed_column] = out[mixed_column].astype("object")
    codes = [format_code(code) for code in out["Code"].tolist()]
    quotes = fetch_eastmoney_quotes(codes)
    missing_codes = [code for code in codes if code not in quotes]
    tencent_quotes = fetch_tencent_quotes(missing_codes)

    stats_by_code: dict[str, dict] = {}
    failed_stats: list[str] = []
    for code in codes:
        stats = fetch_eastmoney_kline_stats(code)
        if first_valid(stats.get("Avg20Amount")) is None or first_valid(stats.get("PrevDayLow")) is None:
            live_quote = quotes.get(code) or tencent_quotes.get(code) or {}
            fallback_stats = fetch_tencent_kline_stats(code, live_quote.get("amount"))
            for key in ("Avg20Amount", "Avg20AmountSource", "PrevDayLow", "PrevDayLowSource"):
                if first_valid(stats.get(key)) is None and first_valid(fallback_stats.get(key)) is not None:
                    stats[key] = fallback_stats.get(key)
        stats_by_code[code] = stats
        if first_valid(stats.get("Avg20Amount")) is None or first_valid(stats.get("PrevDayLow")) is None:
            failed_stats.append(code)
        time.sleep(0.4)

    if failed_stats:
        print(f"日K统计首轮失败 {len(failed_stats)} 条，等待后重试：{failed_stats}")
        time.sleep(3)
        for code in failed_stats:
            retry_stats = fetch_eastmoney_kline_stats(code)
            merged_stats = stats_by_code[code]
            for key in ("Avg20Amount", "Avg20AmountSource", "PrevDayLow", "PrevDayLowSource"):
                if first_valid(retry_stats.get(key)) is not None:
                    merged_stats[key] = retry_stats.get(key)
            stats_by_code[code] = merged_stats
            time.sleep(0.8)

    for idx, row in out.iterrows():
        code = format_code(row["Code"])
        quote = quotes.get(code) or tencent_quotes.get(code)
        quote_source = "东方财富补齐" if code in quotes else "腾讯行情补齐"
        if quote:
            mapping = {
                "Name": "name",
                "Latest": "latest",
                "PctChg": "pct_chg",
                "Open": "open",
                "High": "high",
                "Low": "low",
                "PrevClose": "prev_close",
                "Amount": "amount",
            }
            for col, key in mapping.items():
                if col not in out.columns:
                    continue
                value = quote.get(key)
                if first_valid(value) is None:
                    continue
                if col == "Name" and first_valid(out.at[idx, col]) is not None:
                    continue
                out.at[idx, col] = value
            if "QuoteTime" in out.columns and first_valid(quote.get("quote_time")) is not None:
                out.at[idx, "QuoteTime"] = quote.get("quote_time")
            if str(out.at[idx, "DataSource"]).strip() in ("券商截图", "", "nan"):
                out.at[idx, "DataSource"] = quote_source
            if "Notes" in out.columns:
                out.at[idx, "Notes"] = remove_note_phrases(
                    out.at[idx, "Notes"],
                    ["未取到行情，使用券商截图价格", "未取到行情"],
                )

        stats = stats_by_code.get(code, {"Avg20Amount": None, "PrevDayLow": None})
        if "Avg20Amount" in out.columns and first_valid(stats.get("Avg20Amount")) is not None:
            out.at[idx, "Avg20Amount"] = stats.get("Avg20Amount")
            if "Avg20AmountSource" in out.columns:
                out.at[idx, "Avg20AmountSource"] = stats.get("Avg20AmountSource") or "东方财富日K统计"
            if stats.get("Avg20AmountSource") == "腾讯日K量能等价估算" and "Notes" in out.columns:
                out.at[idx, "Notes"] = append_note(
                    out.at[idx, "Notes"],
                    "东方财富日K重试失败；20日均额按腾讯日K成交量等价估算",
                )
        elif "Avg20Amount" in out.columns and first_valid(out.at[idx, "Avg20Amount"]) is None:
            amount_value = out.at[idx, "Amount"] if "Amount" in out.columns else None
            if first_valid(amount_value) is not None:
                out.at[idx, "Avg20Amount"] = amount_value
                if "Avg20AmountSource" in out.columns:
                    out.at[idx, "Avg20AmountSource"] = "当日成交额替代"
                if "Notes" in out.columns:
                    out.at[idx, "Notes"] = append_note(out.at[idx, "Notes"], "20日均额接口失败，临时用当日成交额替代")
        if "PrevDayLow" in out.columns and first_valid(stats.get("PrevDayLow")) is not None:
            out.at[idx, "PrevDayLow"] = stats.get("PrevDayLow")
            if "PrevDayLowSource" in out.columns:
                out.at[idx, "PrevDayLowSource"] = stats.get("PrevDayLowSource") or "东方财富日K统计"
            if stats.get("PrevDayLowSource") == "腾讯日K补齐" and "Notes" in out.columns:
                out.at[idx, "Notes"] = append_note(
                    out.at[idx, "Notes"],
                    "东方财富昨日低点重试失败；已由腾讯日K补齐",
                )
        elif "PrevDayLow" in out.columns and first_valid(out.at[idx, "PrevDayLow"]) is None:
            low_value = out.at[idx, "Low"] if "Low" in out.columns else None
            if first_valid(low_value) is not None:
                out.at[idx, "PrevDayLow"] = low_value
                if "PrevDayLowSource" in out.columns:
                    out.at[idx, "PrevDayLowSource"] = "当日低点替代"
                if "Notes" in out.columns:
                    out.at[idx, "Notes"] = append_note(out.at[idx, "Notes"], "昨日低点接口失败，临时用当日低点替代")
        if "ETFShareChg" in out.columns and first_valid(out.at[idx, "ETFShareChg"]) is None:
            out.at[idx, "ETFShareChg"] = describe_share_proxy(
                out.at[idx, "Amount"] if "Amount" in out.columns else None,
                out.at[idx, "Avg20Amount"] if "Avg20Amount" in out.columns else None,
                out.at[idx, "Avg20AmountSource"] if "Avg20AmountSource" in out.columns else "",
            )
        if "LeaderStatus" in out.columns and str(out.at[idx, "LeaderStatus"]).strip() in ("待填", "", "nan"):
            out.at[idx, "LeaderStatus"] = describe_leader_proxy(
                code,
                out.at[idx, "Name"] if "Name" in out.columns else "",
                out.at[idx, "Role"] if "Role" in out.columns else "",
                out.at[idx, "PctChg"] if "PctChg" in out.columns else None,
            )
        if "Premium" in out.columns and first_valid(out.at[idx, "Premium"]) is None:
            intraday_pos = calc_intraday_position(
                out.at[idx, "Latest"] if "Latest" in out.columns else None,
                out.at[idx, "Low"] if "Low" in out.columns else None,
                out.at[idx, "High"] if "High" in out.columns else None,
            )
            out.at[idx, "Premium"] = describe_premium_proxy(
                out.at[idx, "PctChg"] if "PctChg" in out.columns else None,
                intraday_pos,
            )

    return attach_market_interface_diagnostics(out)


def lookup_quote(code: str, spot_map: dict[str, pd.DataFrame]) -> dict:
    """
    根据代码，从已抓取的行情表里查找一条记录。
    找不到时返回空字典，后续会标记“未取到行情”。
    """
    kind = classify_code(code)

    if kind == "index":
        pool = spot_map.get("index", pd.DataFrame())
        if pool.empty:
            return {}
        symbols = normalize_index_symbols(code)
        pool = pool.copy()
        pool["symbol"] = pool["symbol"].astype(str).str.lower()
        hit = pool[pool["symbol"].isin(symbols)]
        return hit.iloc[0].to_dict() if not hit.empty else {}

    if kind == "etf":
        pool = spot_map.get("etf", pd.DataFrame())
        if pool.empty:
            return {}
        symbol = code.zfill(6)
        pool = pool.copy()
        pool["symbol"] = pool["symbol"].astype(str).str.zfill(6)
        hit = pool[pool["symbol"] == symbol]
        return hit.iloc[0].to_dict() if not hit.empty else {}

    if kind == "stock":
        pool = spot_map.get("stock", pd.DataFrame())
        if pool.empty:
            return {}
        symbol = code.zfill(6)
        pool = pool.copy()
        pool["symbol"] = pool["symbol"].astype(str).str.zfill(6)
        hit = pool[pool["symbol"] == symbol]
        return hit.iloc[0].to_dict() if not hit.empty else {}

    return {}


def build_market_data(
    watchlist: pd.DataFrame,
    spot_map: dict[str, pd.DataFrame],
    positions: pd.DataFrame | None = None,
) -> pd.DataFrame:
    """生成 Market_Data 工作表。"""
    rows: list[dict] = []
    position_by_code: dict[str, dict] = {}
    if positions is not None and not positions.empty:
        position_by_code = {
            str(item["Code"]).strip(): item.to_dict()
            for _, item in positions.iterrows()
        }

    items: list[dict] = []
    seen_codes: set[str] = set()

    for _, item in watchlist.iterrows():
        code = str(item["Code"]).strip()
        seen_codes.add(code)
        items.append(item.to_dict())

    if positions is not None:
        for _, item in positions.iterrows():
            code = str(item["Code"]).strip()
            if code in seen_codes:
                continue
            seen_codes.add(code)
            item_dict = item.to_dict()
            item_dict["Notes"] = "持仓补充行情"
            items.append(item_dict)

    for item in items:
        code = str(item["Code"]).strip()
        quote = lookup_quote(code, spot_map)
        position_item = position_by_code.get(code, {})
        has_quote = bool(quote)
        data_source = "行情接口" if has_quote else "券商截图"

        row = {
            "Code": code,
            "Name": quote.get("name") or item.get("Name", ""),
            "Role": item.get("Role", ""),
            "Latest": first_valid(quote.get("latest"), position_item.get("Latest")),
            "PctChg": quote.get("pct_chg"),
            "Open": quote.get("open"),
            "High": quote.get("high"),
            "Low": quote.get("low"),
            "PrevClose": quote.get("prev_close"),
            "Amount": quote.get("amount"),
            "Avg20Amount": None,
            "Avg20AmountSource": None,
            "ETFShareChg": None,
            "Premium": quote.get("premium"),
            "LeaderStatus": "待填",
            "PrevDayLow": None,
            "PrevDayLowSource": None,
            "QuoteTime": None,
            "DataSource": data_source,
            "接口状态": "待补齐",
            "失败原因": "待运行行情补齐接口",
            "重试结果": "待运行日K补齐接口",
            "Notes": item.get("Notes", ""),
        }

        if not has_quote:
            if row["Latest"] is not None:
                row["Notes"] = (str(row["Notes"]) + "；未取到行情，使用券商截图价格").strip("；")
            else:
                row["Notes"] = (str(row["Notes"]) + "；未取到行情").strip("；")

        rows.append(row)

    columns = [
        "Code", "Name", "Role", "Latest", "PctChg", "Open", "High", "Low", "PrevClose",
        "Amount", "Avg20Amount", "Avg20AmountSource", "ETFShareChg", "Premium", "LeaderStatus", "PrevDayLow",
        "PrevDayLowSource", "QuoteTime", "DataSource", "接口状态", "失败原因", "重试结果", "Notes",
    ]
    return pd.DataFrame(rows, columns=columns)


# ==================== 买点过滤器（最小版，仅做提醒） ====================
def calc_intraday_position(latest, low, high) -> float | None:
    """日内位置：0=接近最低，1=接近最高。"""
    try:
        latest = float(latest)
        low = float(low)
        high = float(high)
        if high <= low:
            return None
        return round((latest - low) / (high - low), 4)
    except (TypeError, ValueError):
        return None


def describe_share_proxy(amount, avg20_amount, avg20_source: str = "") -> str:
    """Use turnover activity as a proxy when fund-share change is unavailable."""
    if "替代" in str(avg20_source or ""):
        return "量能替代：20日均额接口失败，按当日成交额临时基准"
    vr = volume_pace_ratio(amount, avg20_amount)
    if vr is None:
        return "量能替代：缺少成交额/20日均额"
    if vr >= 1.8:
        return f"量能替代：时间校正后显著放量 {vr:.2f}x"
    if vr >= 1.2:
        return f"量能替代：时间校正后温和放量 {vr:.2f}x"
    if vr >= 0.7:
        return f"量能替代：时间校正后正常 {vr:.2f}x"
    return f"量能替代：时间校正后缩量 {vr:.2f}x"


def describe_premium_proxy(pct_chg, intraday_pos) -> str:
    """Use price movement and intraday position as a proxy when real-time IOPV premium is unavailable."""
    pct = to_float(pct_chg)
    if pct is None:
        return "折溢价替代：缺少涨跌幅"
    pos_text = "日内位置缺失"
    if intraday_pos is not None:
        if intraday_pos >= 0.85:
            pos_text = "接近日内高位"
        elif intraday_pos <= 0.25:
            pos_text = "接近日内低位"
        else:
            pos_text = "日内中位"
    return f"折溢价替代：涨跌幅{pct:.2f}% / {pos_text}"


def describe_leader_proxy(code: str, name: str, role: str, pct_chg) -> str:
    """Use watchlist role and instrument name as a proxy for leader/core status."""
    role_text = str(role or "").strip()
    name_text = str(name or "").strip()
    pct = to_float(pct_chg)
    if role_text:
        if "核心" in role_text:
            return f"角色替代：核心持有（{role_text}）"
        if "宽基" in role_text:
            return f"角色替代：宽基锚点（{role_text}）"
        if "防守" in role_text or "债" in name_text or "红利" in name_text:
            return f"角色替代：防守资产（{role_text}）"
        return f"角色替代：{role_text}"
    if name_text.endswith("ETF") or "ETF" in name_text:
        return "名称替代：主题/行业ETF"
    if pct is not None and pct >= 1.0:
        return "强弱替代：跑出相对强势"
    return "替代：普通观察标的"


def first_valid_text(*values, fallback: str = "") -> str:
    value = first_valid(*values)
    if value is None:
        return fallback
    text = str(value).strip()
    if text in {"暂无接口", "待填", "待人工确认", "nan", ""}:
        return fallback
    return text


def is_dividend_stock_candidate(code: str, name: str, role: str, theme: str, asset_type: str, notes: str) -> bool:
    """识别红利现金流个股；ETF红利低波仍按低波权益ETF处理。"""
    normalized = format_code(code)
    text = " ".join(str(value or "") for value in [name, role, theme, asset_type, notes])
    is_stock = str(asset_type).strip().lower() == "stock"
    return is_stock and (
        normalized in DIVIDEND_CORE_STOCK_CODES
        or normalized in DIVIDEND_ENHANCED_STOCK_CODES
        or "红利个股" in text
        or "红利现金流" in text
        or "稳定分红" in text
    )


def dividend_stock_profile(code: str, name: str, role: str, theme: str, asset_type: str, notes: str) -> dict[str, str]:
    """红利个股专用模板，防止把低波现金流层误判成成长突破仓。"""
    if not is_dividend_stock_candidate(code, name, role, theme, asset_type, notes):
        return {
            "红利配置角色": "",
            "分红质量检查": "",
            "红利买点规则": "",
            "高息陷阱否决": "",
        }
    normalized = format_code(code)
    if normalized in DIVIDEND_CORE_STOCK_CODES:
        role_text = "A档红利底仓候选"
        single_limit = "初始0.5%–1%；成熟1.5%–3%"
    else:
        role_text = "B档红利增强候选"
        single_limit = "初始0.3%–0.8%；成熟1%–2%"
    return {
        "红利配置角色": role_text,
        "分红质量检查": "至少5年稳定分红；分红率30%–75%优先；经营现金流覆盖分红；盈利不连续明显下滑",
        "红利买点规则": f"缩量回踩或横盘止跌；股息率安全垫达标；PE/PB处于历史中低位；{single_limit}",
        "高息陷阱否决": "分红率>100%、高息来自股价暴跌、盈利/现金流恶化、举债分红、资本开支压力或只因快分红而买，均一票否决",
    }


def volume_structure_profile(code: str, name: str, role: str, theme: str, asset_type: str, notes: str) -> dict[str, str]:
    """把V2.8.5补充模块转成每日表格可读的趋势结构纪律。"""
    code = format_code(code)
    text = " ".join(str(value or "") for value in [name, role, theme, asset_type, notes])
    if is_index_code(code):
        return {
            "趋势结构": "指数锚点",
            "关键支撑纪律": "只判断市场权限，不执行买卖",
            "单次仓位上限": "不可交易",
            "爆量/量质吸纪律": "仅用于双锚和风格强弱判断",
        }
    if code == "002594" or "2030战略观察" in text:
        return {
            "趋势结构": "2030战略观察：基本面+估值+技术三重确认",
            "关键支撑纪律": "周线企稳，日线缩量回踩，不破前低，重新站回20/60日线",
            "单次仓位上限": "现有底仓；合格加仓后全资产上限3%",
            "爆量/量质吸纪律": "不做T不追涨；股价涨但利润不涨时暂停或减仓复审",
        }
    if "遗留仓" in text:
        return {
            "趋势结构": "遗留仓反弹复审",
            "关键支撑纪律": "不因回踩补仓，反弹按退出纪律处理",
            "单次仓位上限": "0（禁止新增）",
            "爆量/量质吸纪律": "爆量只作减仓/退出复审，不作为买点",
        }
    if code == "159516":
        return {
            "趋势结构": "第一核心：爆量突破→缩量回踩→量质吸确认",
            "关键支撑纪律": "不破爆量阳线低点/5日线/10日线；龙头不破位",
            "单次仓位上限": "突破试探0.5%–1%；回踩≤1%；强确认1%–2%",
            "爆量/量质吸纪律": "连续上涨不追；回踩确认后分批；放量滞涨暂停新增",
        }
    if code == "159381":
        return {
            "趋势结构": "AI硬件增强：龙头缩量企稳后的回踩确认",
            "关键支撑纪律": "不破5日/10日；中际旭创/新易盛/天孚通信等龙头不破位",
            "单次仓位上限": "观察ETF半额；回踩确认≤1%",
            "爆量/量质吸纪律": "不追研报催化高开；优先级低于159516",
        }
    if code == "562500":
        return {
            "趋势结构": "核心持有：止跌三部曲/量质吸再启动",
            "关键支撑纪律": "放量下跌结束→缩量企稳→不再创新低→次日红K",
            "单次仓位上限": "当前持有不加；重新确认前为0",
            "爆量/量质吸纪律": "不因回调直接补；等待龙头成分止跌和份额稳定",
        }
    if is_dividend_stock_candidate(code, name, role, theme, asset_type, notes):
        return {
            "趋势结构": "红利现金流：缩量回踩+股息率安全垫",
            "关键支撑纪律": "不追除权/分红前冲高；优先缩量回调、横盘止跌、未放量破位",
            "单次仓位上限": "A档0.5%–1%；B档0.3%–0.8%；红利个股组合先建3%–5%",
            "爆量/量质吸纪律": "爆量上涨不是买点；红利股以现金流、估值安全垫和回踩确认优先",
        }
    if str(asset_type).strip().lower() == "stock" or "个股" in text:
        return {
            "趋势结构": "个股小仓验证：爆量后缩量回踩",
            "关键支撑纪律": "不破爆量K线低点/平台下沿/5日线/10日线",
            "单次仓位上限": "首仓0.3%–0.5%；单股≤1%，特殊龙头≤2%",
            "爆量/量质吸纪律": "资产质量≥8、主线未退潮、无放量滞涨、情绪≤4",
        }
    if "ETF" in str(name) or str(asset_type).strip().upper() == "ETF":
        return {
            "趋势结构": "ETF回踩确认优先；突破只小仓试探",
            "关键支撑纪律": "优先看爆量阳线低点、5日线、10日线、平台上沿",
            "单次仓位上限": "突破试探0.5%–1%；回踩确认≤1%",
            "爆量/量质吸纪律": "爆量不是买点；缩量回踩不破后再复核",
        }
    return {
        "趋势结构": "等待回踩确认",
        "关键支撑纪律": "关键位不破且次日转强后再复核",
        "单次仓位上限": "需人工确认",
        "爆量/量质吸纪律": "不因单日放量直接加仓",
    }


def cognitive_guardrail_profile(code: str, name: str, role: str, theme: str, asset_type: str, notes: str) -> dict[str, str]:
    """把多元思维模型转成投资决策前的认知防错提示。"""
    code = format_code(code)
    text = " ".join(str(value or "") for value in [name, role, theme, asset_type, notes])
    default_tree = "强势延续=持有不追；缩量回踩=小步复核；失败转弱=暂停/复审"
    default_loss = "先估关键位跌破后的损失；逻辑好但赔率差不买"
    default_bias = "若机会成本、反证、最大损失、能力圈中两项答不清，不买"

    if is_index_code(code):
        return {
            "机会成本检查": "指数只作市场锚点，不占买入预算",
            "认知防错": "避免把指数短线波动误读为单一标的买点",
            "决策树情景": "双锚转强=开放复核；成长红=暂停新增；系统破位=只做风控",
            "赔率/最大损失": "用于判断市场权限，不直接计算单标的赔率",
            "反向失败路径": "宽基或成长锚持续弱于对照指数，说明市场权限下降",
            "能力圈边界": "市场锚点，非交易标的",
        }
    if code == "002594" or "2030战略观察" in text:
        return {
            "机会成本检查": "不得挤占159516/159381等核心ETF预算；只用战略观察仓额度复核",
            "认知防错": "优秀公司不等于立即加仓；避免信仰加仓、沉没成本和做T冲动",
            "决策树情景": "基本面改善+估值安全垫+技术确认=小额复核；任一缺失=持有观察；风险信号出现=暂停/降级",
            "赔率/最大损失": "上限按全资产3%控制；先估海外、利润率、现金流失败时的回撤空间",
            "反向失败路径": "海外受阻、价格战恶化、净利率长期低于3%-4%、现金流恶化或股价涨但利润不涨",
            "能力圈边界": "新能源战略观察，不是当前第一主线进攻仓",
        }
    if "遗留仓" in text:
        return {
            "机会成本检查": "占用核心主线和现金等待买点的资金效率，优先反弹压缩",
            "认知防错": "亏损不是补仓理由；避免沉没成本、损失规避和做T冲动",
            "决策树情景": "反弹转强=减仓复核；继续走弱=保持冻结；重新过主线/质量/买点才可升级",
            "赔率/最大损失": "先看继续持有的机会成本和再次下跌空间",
            "反向失败路径": "主线外、质量弱、放量滞涨或关键位跌破，继续降级",
            "能力圈边界": "非当前核心能力圈，只允许退出复审",
        }
    if code == "159516":
        return {
            "机会成本检查": "新增资金第一优先；替代选择是159381/A500/可转债/现金等待回踩",
            "认知防错": "不要因半导体利好刷屏确认偏误；必须找净赎回、龙头不涨、放量滞涨等反证",
            "决策树情景": default_tree,
            "赔率/最大损失": "只在回踩不破关键位时提高赔率；跌破爆量K低点暂停新增",
            "反向失败路径": "订单不及预期、ETF连续净赎回、科创50弱于沪深300、龙头放量下跌",
            "能力圈边界": "能力圈核心：AI基础设施/半导体设备ETF",
        }
    if code == "159381":
        return {
            "机会成本检查": "优先级低于159516；只有AI硬件赔率明显更优时才占用新增预算",
            "认知防错": "AI利好和研报容易造成易得性偏差；必须核对龙头、份额、情绪和价格是否已反映",
            "决策树情景": default_tree,
            "赔率/最大损失": "高开低走、净赎回或接近日内低位时赔率差，不买",
            "反向失败路径": "光模块/CPO/AI芯片龙头破位、资金退潮、估值透支",
            "能力圈边界": "能力圈内增强仓，但不得超过159516优先级",
        }
    if code == "562500":
        return {
            "机会成本检查": "证券账户占比已高；新增资金机会成本通常高于159516/159381回踩",
            "认知防错": "盈利仓看趋势，不因怕利润回吐乱卖；也不因回调直接补",
            "决策树情景": "强势延续=持有不追；缩量企稳且次日红K=复审；放量下跌=暂停新增",
            "赔率/最大损失": "等待止跌三部曲提高赔率；跌破关键位进入复审",
            "反向失败路径": "龙头分化、放量滞涨、情绪过热、ETF份额退潮",
            "能力圈边界": "能力圈核心持有，但当前纪律为持有不加",
        }
    if is_dividend_stock_candidate(code, name, role, theme, asset_type, notes):
        return {
            "机会成本检查": "红利个股使用低波现金流预算；先比较红利低波ETF、A500、现金等待和核心成长回踩",
            "认知防错": "不要被高股息率锚定；先确认高息不是股价暴跌、盈利下滑或举债分红造成",
            "决策树情景": "缩量回踩且股息率安全垫达标=小仓；冲高除权前=不追；削减分红/现金流恶化=降级",
            "赔率/最大损失": "收益来自现金流稳定和波动降低；单只先小仓，跌破长期平台或分红逻辑恶化立即复审",
            "反向失败路径": "利润连续下滑、现金流覆盖不足、分红率>100%、资本开支抬升或政策/周期冲击",
            "能力圈边界": "红利现金流层，不能替代AI/半导体核心成长收益发动机",
        }
    if str(asset_type).strip().lower() == "stock" or "个股" in text:
        return {
            "机会成本检查": "个股资金机会成本高于核心ETF；首仓只用于验证，不替代159516/159381",
            "认知防错": "先找反证：利好不涨、放量滞涨、龙头分化、财务/估值证据不足",
            "决策树情景": "缩量回踩不破=小仓验证；强势上冲=不追；跌破爆量K低点=复审/止损",
            "赔率/最大损失": "首仓0.3%–0.5%，先定义跌破关键位的最大亏损",
            "反向失败路径": "订单/Capex下修、利润率恶化、估值透支、板块退潮",
            "能力圈边界": "AI基础设施/半导体硬件能力圈内观察，未补完整评分前不买",
        }
    if "可转债" in text or "国债" in text or "红利" in text or "低波" in text or "防御" in text:
        return {
            "机会成本检查": "用于反脆弱结构和现金替代，不与核心成长仓直接抢进攻预算",
            "认知防错": "不要因短期跑输科技而放弃冗余备份和安全边际",
            "决策树情景": "市场过热=提高防御价值；回踩买点=提供弹药；系统走弱=缓冲波动",
            "赔率/最大损失": "看组合波动降低和现金安全垫贡献，不追求单日弹性",
            "反向失败路径": "估值拥挤、利率冲击或与宽基重复暴露过高",
            "能力圈边界": "组合反脆弱与安全垫工具",
        }
    return {
        "机会成本检查": "与159516、159381、A500/可转债、现金等待买点比较后再决定",
        "认知防错": default_bias,
        "决策树情景": default_tree,
        "赔率/最大损失": default_loss,
        "反向失败路径": "先研究它怎么失败，再决定是否复核买点",
        "能力圈边界": "能力圈边界待确认，不能重仓",
    }


def quant_role_profile(code: str, name: str, role: str, theme: str, asset_type: str, notes: str) -> dict[str, str]:
    """把量化交易文档中的资产角色差异，落成每日表格可读的执行模板。"""
    code = format_code(code)
    text = " ".join(str(value or "") for value in [name, role, theme, asset_type, notes])
    name_text = str(name or "")
    if is_index_code(code):
        return {
            "资产角色": "指数环境锚点",
            "量化买点规则": "只记录涨跌、成交额、强弱和市场广度；不产生交易指令",
            "正期望检查": "用于判断市场权限，不计算单标的期望收益",
            "量化风控规则": "指数破位、成交萎缩或成长锚转红时降低权益新增权限",
        }
    if code == "002594" or "2030战略观察" in text:
        return {
            "资产角色": "2030战略观察个股",
            "量化买点规则": "基本面至少3项改善，估值有安全垫，技术买点通过后才小额复核",
            "正期望检查": "不能因亏损摊低成本；只有长期情景赔率优于核心ETF机会成本时才成立",
            "量化风控规则": "海外/利润率/智能化/现金流恶化暂停加仓；估值先行透支则减仓或锁定利润",
        }
    if "遗留仓" in text:
        return {
            "资产角色": "降级/遗留仓",
            "量化买点规则": "不使用买点规则新增；只看反弹减仓、跌破复审和机会成本",
            "正期望检查": "沉没成本不计入期望收益；若继续持有弱于核心ETF/现金等待，则压缩",
            "量化风控规则": "反弹无量、利好不涨、放量下跌或继续弱于主线时减仓/退出复审",
        }
    if code == "159516":
        return {
            "资产角色": "核心成长ETF",
            "量化买点规则": "缩量回踩不破爆量K低点/5日/10日线，龙头同步，份额不连续赎回，次日转强",
            "正期望检查": "新增资金第一优先，但只有回踩确认提高胜率且跌破关键位损失可控时才成立",
            "量化风控规则": "连续上涨不追；净赎回、龙头不涨、放量滞涨或跌破爆量低点则暂停新增",
        }
    if code == "159381":
        return {
            "资产角色": "AI硬件增强ETF",
            "量化买点规则": "光模块/CPO/AI芯片龙头缩量企稳，ETF不破5日/10日，次日重新放量转强",
            "正期望检查": "必须优于159516/A500/可转债的机会成本，且不能被研报催化后的高开透支",
            "量化风控规则": "高开低走、净赎回、硬件链龙头破位或估值透支时禁止新增",
        }
    if code == "562500":
        return {
            "资产角色": "核心持有ETF",
            "量化买点规则": "持有不追；重新买点需放量下跌结束、缩量企稳、不再创新低、龙头止跌、次日红K",
            "正期望检查": "当前新增机会成本通常高于159516/159381回踩，未完成止跌三部曲前不成立",
            "量化风控规则": "放量滞涨、龙头分化、份额退潮或情绪≥4时暂停新增并复审",
        }
    if code in {"159338", "512890", "510210"} or any(key in text for key in ("A500", "红利低波", "上证指数")):
        return {
            "资产角色": "低波权益ETF",
            "量化买点规则": "按配置节奏分批；不在指数大涨日追买，优先缩量回踩、权重股止跌、份额稳定",
            "正期望检查": "以降低组合波动和改善全资产结构为收益来源，不和核心成长ETF比单日弹性",
            "量化风控规则": "系统性破位、权重股集体破位或与宽基重复暴露过高时暂停新增/换仓",
        }
    if is_dividend_stock_candidate(code, name, role, theme, asset_type, notes):
        return {
            "资产角色": "红利现金流个股",
            "量化买点规则": "分红连续性、现金流覆盖、股息率安全垫、历史中低估值和缩量回踩同时满足后分批",
            "正期望检查": "正期望来自稳定现金分红+估值均值修复+组合波动降低，不来自短线突破追涨",
            "量化风控规则": "削减分红、现金流低于分红、分红率过高且业绩下滑、政策/周期冲击时降级或退出",
        }
    if "可转债" in text:
        return {
            "资产角色": "弹性低波ETF",
            "量化买点规则": "每次小额分批；股市不单边暴跌、转债估值不过热、ETF缩量回调或横盘企稳",
            "正期望检查": "收益来自下跌缓冲与反弹弹性，不用于短线追涨或替代核心成长买点",
            "量化风控规则": "转债溢价明显过热、利率快速上行或权益系统性下跌扩散时暂停新增",
        }
    if "国债" in text or "纯防御" in text:
        return {
            "资产角色": "纯防御ETF",
            "量化买点规则": "作为现金安全垫和心理稳定器；不因短期跑输科技而切走",
            "正期望检查": "期望收益来自保留弹药和降低组合断裂风险，不追求高弹性",
            "量化风控规则": "权益仓提高时必须保留；利率急升或安全垫显著超目标时再复核",
        }
    if str(asset_type).strip().lower() == "stock" or "个股" in text:
        return {
            "资产角色": "卫星验证个股",
            "量化买点规则": "质量≥8、主线未退潮、爆量后缩量回踩、不破爆量K低点、前排龙头、情绪≤4",
            "正期望检查": "个股机会成本高于核心ETF；首仓只验证产业链弹性，先定义跌破关键位最大损失",
            "量化风控规则": "订单/Capex下修、利润率恶化、放量滞涨、后排补涨或跌破关键位时快速复审",
        }
    if "ETF" in name_text.upper() or str(asset_type).strip().upper() == "ETF":
        return {
            "资产角色": "观察ETF",
            "量化买点规则": "按资产角色重评分；不能套用核心成长ETF买点，不追单日放量",
            "正期望检查": "先比较159516、159381、A500、可转债和现金等待买点",
            "量化风控规则": "角色不清、证据不足、份额恶化或不在能力圈时只观察",
        }
    return {
        "资产角色": "待分类资产",
        "量化买点规则": "先补资产角色、质量评分、资金/价格结构与证伪条件",
        "正期望检查": "未能估计胜率、赔率、失败损失和替代选择前不买",
        "量化风控规则": "能力圈外或证据链不完整时只观察",
    }


def quant_validation_result(
    *,
    pass_items: list[str],
    fail_items: list[str],
    pending_items: list[str],
    quality_value: float | None,
    emotion_temperature: int,
    veto: str,
    hard_block_kind: str,
) -> dict[str, object]:
    """将买点过滤器六项检查转成 V2.8.5-Q 的10分制验证结果。"""
    score = 0
    score += min(2, sum(1 for item in pass_items if item in {"分时结构", "量价关系"}))
    score += min(2, sum(1 for item in pass_items if item in {"份额/筹码", "折溢价/估值"}))
    score += min(2, sum(1 for item in pass_items if item in {"龙头同步", "次日验证"}))
    if quality_value is not None and quality_value >= QUALITY_CORE_MIN:
        score += 1
    elif quality_value is not None and quality_value >= QUALITY_OBSERVE_MIN:
        score += 0.5
    if emotion_temperature < EMOTION_PAUSE_LEVEL:
        score += 1
    if not fail_items and len(pending_items) <= 1:
        score += 2
    elif len(fail_items) <= 1:
        score += 1
    score = round(float(score), 1)

    if hard_block_kind or veto == "是":
        status = "阻断/不买"
    elif score >= 8:
        status = "标准量化验证"
    elif score >= 6:
        status = "黄灯观察"
    elif score >= 4:
        status = "不买"
    else:
        status = "风险复审"

    if score >= 8 and veto != "是" and not hard_block_kind:
        expected = f"通过{len(pass_items)}项，正期望可人工复核；先定义跌破关键位损失"
    elif score >= 6 and veto != "是":
        expected = f"通过{len(pass_items)}项，仍需补：{'、'.join(pending_items) if pending_items else '反证/赔率'}"
    else:
        blockers = "、".join(fail_items + pending_items[:2]) or "硬性阻断/证据不足"
        expected = f"正期望不足或待补；主要约束：{blockers}"
    return {
        "量化验证分": score,
        "量化验证状态": status,
        "正期望检查": expected,
    }


# ==================== V2.8.5 质量、情绪与穿透风控 ====================
def trading_session_progress(now: datetime | None = None) -> float:
    """Return the completed fraction of a normal A-share trading day."""
    current = now or datetime.now()
    t = current.time()
    morning_start = dt_time(9, 30)
    morning_end = dt_time(11, 30)
    afternoon_start = dt_time(13, 0)
    afternoon_end = dt_time(15, 0)
    if t <= morning_start:
        return 0.05
    if t < morning_end:
        elapsed = (current - current.replace(hour=9, minute=30, second=0, microsecond=0)).total_seconds()
        return max(elapsed / (4 * 3600), 0.05)
    if t < afternoon_start:
        return 0.5
    if t < afternoon_end:
        elapsed = (current - current.replace(hour=13, minute=0, second=0, microsecond=0)).total_seconds()
        return min(0.5 + elapsed / (4 * 3600), 1.0)
    return 1.0


def volume_pace_ratio(amount, avg20_amount) -> float | None:
    """Time-adjusted turnover pace; avoids labeling morning data as false contraction."""
    raw = volume_ratio(amount, avg20_amount)
    if raw is None:
        return None
    return round(raw / trading_session_progress(), 2)


def classify_theme(name: str, role: str = "") -> str:
    text = f"{name} {role}"
    if any(key in text for key in ("红利现金流", "红利个股", "长江电力", "中国移动", "工商银行", "建设银行", "中国神华", "美的集团")):
        return "红利现金流"
    if any(key in text for key in ("半导体", "芯片")):
        return "半导体/芯片设备"
    if any(key in text for key in ("人工智能", "云计算", "算力", "中望")):
        return "AI算力/云计算"
    if any(key in text for key in ("机器人", "工控", "自动化")):
        return "机器人/自动化"
    if any(key in text for key in ("创新药", "生物医药", "医药")):
        return "创新药/生物医药"
    if any(key in text for key in ("比亚迪", "绿电", "储能", "电池", "新能源")):
        return "新能源/绿电"
    if any(key in text for key in ("国债", "可转债", "红利", "A500", "上证")):
        return "防御/宽基"
    return "其他"


def feedback_state(pct_chg, intraday_pos, pace_ratio) -> str:
    pct = to_float(pct_chg)
    pace = to_float(pace_ratio)
    if pct is None:
        return "待确认"
    if pct < -1.5 or (intraday_pos is not None and intraday_pos <= 0.2 and pct < 0):
        return "退潮期"
    if pct >= 3.0 and (pace or 0) >= 1.5:
        return "加速期"
    if pct >= 1.0 and (pace or 0) >= 1.0:
        return "成长期"
    if pct >= 0 and intraday_pos is not None and intraday_pos >= 0.45:
        return "启动期"
    if pct >= 2.0 and (pace or 0) < 0.8:
        return "一致期"
    return "观察期"


def build_quality_score(watchlist: pd.DataFrame, market_data: pd.DataFrame) -> pd.DataFrame:
    """V2.8.5 质量准入。

    代理指标只展示已知部分，不再把 7 分制机械折算成 10 分制。
    只有观察池中的人工完整评分，或未来接入的完整 10 分数据，才可以通过质量门槛。
    """
    market_by_code = market_data.set_index("Code")
    rows: list[dict] = []
    for _, item in watchlist.iterrows():
        code = format_code(item.get("Code"))
        name = str(item.get("Name", ""))
        role = str(item.get("Role", ""))
        quote = market_by_code.loc[code] if code in market_by_code.index else None
        amount = quote.get("Amount") if quote is not None else None
        avg20 = quote.get("Avg20Amount") if quote is not None else None
        pct = to_float(quote.get("PctChg") if quote is not None else None)
        pace = volume_pace_ratio(amount, avg20)
        asset_type = str(item.get("Asset Type", "")).strip().lower()
        is_index = is_index_code(code) or asset_type == "index"
        is_etf = ("ETF" in name.upper() or asset_type == "etf") and not is_index
        manual_score = to_float(item.get("Manual Quality Score"))
        manual_evidence = str(item.get("Quality Evidence", "") or "").strip()
        missing_quality_inputs: list[str] = []
        if manual_score is None:
            missing_quality_inputs.append("Manual Quality Score")
        elif not 0 <= manual_score <= 10:
            missing_quality_inputs.append("Manual Quality Score需0-10")
        if not manual_evidence:
            missing_quality_inputs.append("Quality Evidence")

        if is_index:
            proxy_score = None
            available = 0
            score_type = "指数锚点"
            detail = "指数不参与质量准入评分"
        elif is_etf:
            purity = 3 if any(key in role for key in ("核心", "战术")) else 2
            leader = 2 if "核心" in role else 1
            liquidity = 2 if (avg20 or 0) >= 5e8 else 1 if (avg20 or 0) >= 1e8 else 0
            proxy_score = purity + leader + liquidity
            available = 7
            score_type = "ETF质量评分"
            detail = f"代理：纯度{purity}/3；龙头覆盖{leader}/2；流动性{liquidity}/2；资金认可与规模稳定性待补"
        else:
            leadership = 1 if "遗留仓" in role else 2 if "核心" in role else 1
            cycle = 2 if pct is not None and pct > 0 and (pace or 0) >= 1 else 1 if pct is not None and pct >= 0 else 0
            proxy_score = leadership + cycle
            available = 5
            score_type = "个股质量评分"
            detail = f"代理：龙头性{leadership}/3；景气{cycle}/2；盈利、估值、机构数据待补"

        valid_manual = manual_score is not None and 0 <= manual_score <= 10 and bool(manual_evidence)
        effective_score = round(manual_score, 1) if valid_manual else None
        coverage = 1.0 if valid_manual else available / 10 if available else 0
        if is_index:
            score_source = "不适用"
            input_status = "指数锚点不参与质量评分"
            missing_text = "无"
        elif valid_manual:
            score_source = "watchlist.csv人工完整评分"
            input_status = "完整"
            missing_text = "无"
        else:
            score_source = "代理展示，不放行"
            input_status = "待补"
            missing_text = "、".join(missing_quality_inputs) if missing_quality_inputs else "完整评分未确认"
        if is_index:
            status = "不适用"
        elif effective_score is None:
            status = "数据不足，需人工评分"
        elif effective_score >= QUALITY_CORE_MIN:
            status = "核心候选"
        elif effective_score >= QUALITY_OBSERVE_MIN:
            status = "观察候选"
        else:
            status = "低于6分，禁止新增"
        if valid_manual:
            detail = f"人工完整评分：{effective_score:.1f}/10；依据：{manual_evidence}"
        rows.append(
            {
                "Code": code,
                "Name": name,
                "类型": score_type,
                "主题": classify_theme(name, role),
                "可用原始分": proxy_score,
                "可用满分": available,
                "折算质量分": effective_score,
                "数据完整度": coverage,
                "质量状态": status,
                "评分来源": score_source,
                "评分输入状态": input_status,
                "评分缺失项": missing_text,
                "评分明细": detail,
                "数据边界": "代理分不得折算放行；完整人工评分必须同时填写分数与证据",
            }
        )
    return pd.DataFrame(rows)


def build_emotion_thermometer(
    market_data: pd.DataFrame,
    decision_inputs: dict[str, dict[str, str]] | None = None,
) -> pd.DataFrame:
    """五档情绪温度计。

    人工温度优先；没有人工输入时，只输出“观察池行情代理”，不冒充全市场情绪。
    """
    inputs = decision_inputs or {}
    tradable = market_data[~market_data["Code"].astype(str).str.contains(r"\.(?:SH|SZ|CSI)$", regex=True)].copy()
    pcts = pd.to_numeric(tradable.get("PctChg"), errors="coerce").dropna()
    pace_values = [
        volume_pace_ratio(row.get("Amount"), row.get("Avg20Amount"))
        for _, row in tradable.iterrows()
    ]
    pace_values = [value for value in pace_values if value is not None]
    breadth = float((pcts > 0).mean()) if not pcts.empty else None
    hot_share = float((pcts >= 3).mean()) if not pcts.empty else None
    median_pct = float(pcts.median()) if not pcts.empty else None
    median_pace = float(pd.Series(pace_values).median()) if pace_values else None

    points = 1
    if median_pct is not None and median_pct >= 0.5:
        points += 1
    if breadth is not None and breadth >= 0.6:
        points += 1
    if hot_share is not None and hot_share >= 0.2:
        points += 1
    if median_pace is not None and median_pace >= 1.5:
        points += 1
    proxy_temperature = min(points, 5)
    manual_temperature = decision_input_number(inputs, "emotion_temperature")
    manual_valid = manual_temperature is not None and 1 <= manual_temperature <= 5
    temperature = int(manual_temperature) if manual_valid else proxy_temperature
    reliability = "高（人工确认）" if manual_valid else "低（观察池代理）"
    source = inputs.get("emotion_temperature", {}).get("Source", "") if manual_valid else "观察池涨跌与成交节奏"
    labels = {1: "冰点", 2: "回暖", 3: "温热", 4: "火热", 5: "狂热"}
    action = {
        1: "允许重视，可按买点纪律建仓",
        2: "允许建仓",
        3: "以持有为主，新增需确认非追高",
        4: "暂停新增，等待降温",
        5: "启动减仓复审",
    }[temperature]
    return pd.DataFrame(
        [
            {"指标": "情绪温度", "数值": temperature, "状态": labels[temperature], "解释": action, "可靠性": reliability, "来源": source},
            {"指标": "观察池代理温度", "数值": proxy_temperature, "状态": labels[proxy_temperature], "解释": "仅用于提示，不代表全市场", "可靠性": "低", "来源": "观察池"},
            {"指标": "上涨广度", "数值": breadth, "状态": "观察池上涨占比", "解释": "代理指标", "可靠性": "低", "来源": "观察池"},
            {"指标": "涨幅≥3%占比", "数值": hot_share, "状态": "短线热度", "解释": "代理指标", "可靠性": "低", "来源": "观察池"},
            {"指标": "涨跌幅中位数", "数值": median_pct, "状态": "观察池中枢", "解释": "百分比点", "可靠性": "低", "来源": "观察池"},
            {"指标": "时间校正成交节奏", "数值": median_pace, "状态": "相对20日均额", "解释": "已按交易时段折算", "可靠性": "中", "来源": "行情接口"},
            {"指标": "数据边界", "数值": None, "状态": "媒体/社区/ETF申赎/两融/涨停热度待补", "解释": "缺失项不参与自动判分", "可靠性": "提示", "来源": "decision_inputs.csv"},
        ]
    )


def build_exposure_summary(positions_sheet: pd.DataFrame, account_meta: dict[str, float]) -> pd.DataFrame:
    account_total = to_float(account_meta.get("account_total")) or 0
    total_assets = to_float(account_meta.get("total_assets")) or 0
    rows: list[dict] = []
    for _, item in positions_sheet.iterrows():
        market_value = to_float(item.get("Market Value")) or 0
        account_weight = market_value / account_total if account_total else None
        full_weight = market_value / total_assets if total_assets else None
        rows.append(
            {
                "维度": "标的",
                "分类": classify_theme(str(item.get("Name", "")), str(item.get("Role", ""))),
                "Code": item.get("Code"),
                "Name": item.get("Name"),
                "市值": market_value,
                "证券账户仓位": account_weight,
                "全资产穿透仓位": full_weight,
                "目标仓位": to_float(item.get("Target Weight")),
                "主题上限": None,
                "上限剩余额度": None,
                "上限状态": "明细见主题汇总",
                "风控口径": "证券账户 + 全资产穿透",
                "风险提示": item.get(f"{FRAMEWORK_VERSION} Action", item.get("V2.8.4 Action", "")),
            }
        )
    detail = pd.DataFrame(rows)
    groups: list[dict] = []
    if not detail.empty:
        for category, group in detail.groupby("分类", dropna=False):
            full_weight = group["全资产穿透仓位"].sum()
            limit = THEME_LIMITS.get(str(category))
            groups.append(
                {
                    "维度": "主题汇总",
                    "分类": category,
                    "Code": "",
                    "Name": f"{category}合计",
                    "市值": group["市值"].sum(),
                    "证券账户仓位": group["证券账户仓位"].sum(),
                    "全资产穿透仓位": full_weight,
                    "目标仓位": group["目标仓位"].sum(min_count=1),
                    "主题上限": limit,
                    "上限剩余额度": max(limit - full_weight, 0.0) if limit is not None else None,
                    "上限状态": "超限" if limit is not None and full_weight > limit else "范围内" if limit is not None else "未设上限",
                    "风控口径": "全资产穿透",
                    "风险提示": "检查产业链簇上限与重复暴露",
                }
            )
        tech_categories = ["半导体/芯片设备", "AI算力/云计算", "机器人/自动化"]
        tech = detail[detail["分类"].isin(tech_categories)]
        tech_weight = tech["全资产穿透仓位"].sum()
        tech_limit = THEME_LIMITS["科技成长合计"]
        groups.insert(
            0,
            {
                "维度": "组合汇总",
                "分类": "科技成长合计",
                "Code": "",
                "Name": "半导体 + AI + 机器人",
                "市值": tech["市值"].sum(),
                "证券账户仓位": tech["证券账户仓位"].sum(),
                "全资产穿透仓位": tech_weight,
                "目标仓位": tech["目标仓位"].sum(min_count=1),
                "主题上限": tech_limit,
                "上限剩余额度": max(tech_limit - tech_weight, 0.0),
                "上限状态": "超限" if tech_weight > tech_limit else "范围内",
                "风控口径": "全资产穿透",
                "风险提示": "科技成长全资产穿透目标/上限按45%-50%复核",
            },
        )
    return pd.DataFrame(groups + detail.to_dict("records"), columns=detail.columns)


def build_first_year_allocation(
    allocation: pd.DataFrame,
    positions_sheet: pd.DataFrame,
    account_meta: dict[str, float],
) -> pd.DataFrame:
    """把1st年配置表转换为动态全资产计划，并用最新持仓重新计算进度。"""
    if allocation.empty:
        return pd.DataFrame()
    total_assets = to_float(account_meta.get("total_assets")) or 0
    market_value_by_code = {
        format_code(row.get("Code")): to_float(row.get("Market Value")) or 0
        for _, row in positions_sheet.iterrows()
    }
    cash_like_by_code = {
        "CASH": to_float(account_meta.get("available_cash")) or 0,
    }
    rows: list[dict[str, object]] = []
    for _, item in allocation.iterrows():
        codes = [format_code(code) for code in str(item.get("Codes", "")).split("|") if str(code).strip()]
        current_amount = sum(market_value_by_code.get(code, cash_like_by_code.get(code, 0)) for code in codes)
        target_weight = to_float(item.get("TargetWeight")) or 0
        source_target = to_float(item.get("SourceTargetAmount"))
        source_current = to_float(item.get("SourceCurrentAmount"))
        dynamic_target = total_assets * target_weight if total_assets else source_target
        gap = (dynamic_target - current_amount) if dynamic_target is not None else None
        completion = current_amount / dynamic_target if dynamic_target else 0.0
        current_weight = current_amount / total_assets if total_assets else None
        if not codes:
            progress_status = "待选择合格标的"
        elif gap is not None and gap <= 0:
            progress_status = "达到或超过年度目标"
        elif completion is not None and completion >= 0.8:
            progress_status = "接近年度目标"
        else:
            progress_status = "配置中"
        rows.append(
            {
                "配置键": item.get("AllocationKey"),
                "配置项": item.get("Label"),
                "资产层级": item.get("AssetLayer"),
                "映射代码": "、".join(codes) if codes else "待选",
                "年度目标占比": target_weight,
                "配置表目标金额": source_target,
                "按当前全资产目标金额": dynamic_target,
                "最新持仓金额": current_amount,
                "配置表当前金额约": source_current,
                "年度资金缺口": gap,
                "年度完成率": completion,
                "当前全资产占比": current_weight,
                "目标收益下限": to_float(item.get("TargetReturnLow")),
                "目标收益上限": to_float(item.get("TargetReturnHigh")),
                "配置状态": item.get("Status"),
                "进度状态": progress_status,
                "执行约束": item.get("ExecutionConstraint"),
                "来源": item.get("Source"),
                "来源日期": item.get("SourceAsOf"),
            }
        )
    return pd.DataFrame(rows)


def summarize_first_year_allocation(first_year: pd.DataFrame, account_meta: dict[str, float]) -> dict[str, float]:
    if first_year.empty:
        return {}
    total_assets = to_float(account_meta.get("total_assets")) or 0
    target_weight = pd.to_numeric(first_year["年度目标占比"], errors="coerce").fillna(0).sum()
    target_amount = pd.to_numeric(first_year["按当前全资产目标金额"], errors="coerce").fillna(0).sum()
    current_amount = pd.to_numeric(first_year["最新持仓金额"], errors="coerce").fillna(0).sum()
    gap = target_amount - current_amount
    return {
        "年度目标占比": float(target_weight),
        "年度目标金额": float(target_amount),
        "年度当前金额": float(current_amount),
        "年度资金缺口": float(gap),
        "年度完成率": float(current_amount / target_amount) if target_amount else 0.0,
        "年度当前全资产占比": float(current_amount / total_assets) if total_assets else 0.0,
        "未配置目标占比": float(max(1 - target_weight, 0)),
    }


def first_year_lookup(first_year: pd.DataFrame) -> dict[str, dict[str, object]]:
    lookup: dict[str, dict[str, object]] = {}
    if first_year.empty:
        return lookup
    for _, row in first_year.iterrows():
        for code in str(row.get("映射代码", "")).split("、"):
            code = format_code(code)
            if code and code != "待选":
                lookup[code] = row.to_dict()
    return lookup


def attach_first_year_fields(frame: pd.DataFrame, first_year: pd.DataFrame) -> pd.DataFrame:
    """给含Code的相关表格追加年度配置字段，不改变原有规则列。"""
    if frame.empty or "Code" not in frame.columns:
        return frame
    lookup = first_year_lookup(first_year)
    output = frame.copy()
    fields = {
        "年度配置项": "配置项",
        "年度全资产目标": "年度目标占比",
        "年度目标金额": "按当前全资产目标金额",
        "年度组内当前金额": "最新持仓金额",
        "年度资金缺口": "年度资金缺口",
        "年度完成率": "年度完成率",
        "年度配置状态": "进度状态",
    }
    for output_column, source_column in fields.items():
        output[output_column] = [lookup.get(format_code(code), {}).get(source_column) for code in output["Code"]]
    return output


STAGE_BANDS = [
    ("观察池", 0.00, 0.00),
    ("观察仓", 0.00, 0.20),
    ("验证仓", 0.20, 0.40),
    ("核心仓", 0.40, 0.70),
    ("目标仓", 0.70, 1.00),
]


def stage_band_for_ratio(ratio: float | None) -> tuple[str, float | None, float | None]:
    """按第一年目标完成度划分阶段，不让年度缺口直接变成买入信号。"""
    if ratio is None:
        return "不适用", None, None
    if ratio <= 0:
        return "观察池", 0.0, 0.0
    if ratio < 0.20:
        return "观察仓", 0.0, 0.20
    if ratio < 0.40:
        return "验证仓", 0.20, 0.40
    if ratio < 0.70:
        return "核心仓", 0.40, 0.70
    if ratio < 1.00:
        return "目标仓", 0.70, 1.00
    return "已达第一年目标", 1.00, 1.00


def is_core_low_config_candidate(row: pd.Series) -> bool:
    text = " ".join(
        str(row.get(column, "") or "")
        for column in ["Name", "Role", "资产角色", "年度配置项", "趋势结构"]
    )
    core_keywords = ("核心", "半导体", "人工智能", "AI", "机器人", "创新药")
    downgrade_keywords = ("遗留", "降级", "禁止新增", "指数锚点")
    return any(keyword in text for keyword in core_keywords) and not any(keyword in text for keyword in downgrade_keywords)


def evaluate_stage_progression(row: pd.Series) -> dict[str, object]:
    target_weight = to_float(first_valid(row.get("目标仓位"), row.get("全资产目标仓位"), row.get("年度全资产目标")))
    current_weight = to_float(first_valid(row.get("当前仓位"), row.get("全资产当前仓位")))
    remaining_weight = to_float(first_valid(row.get("剩余额度"), row.get("全资产剩余额度"), row.get("年度资金缺口")))
    quality_value = to_float(row.get("质量评分"))
    temperature = to_float(row.get("情绪温度"))
    signal = str(row.get("买点灯号", "灰") or "灰")
    veto = str(row.get("一票否决", "否") or "否")
    decision_basis = str(row.get("仓位决策口径", "") or "")
    reason = clean_blocker_reason(row.get("否决原因", "")) or clean_blocker_reason(row.get("阻断原因", ""))

    ratio = None
    if target_weight is not None and target_weight > 0 and current_weight is not None:
        ratio = current_weight / target_weight
    stage, lower, upper = stage_band_for_ratio(ratio)
    stage_verified = "是" if signal in {"绿", "黄"} and veto == "否" else "否"

    if "指数锚点" in decision_basis:
        conclusion = "不适用"
        blocker = "指数只作市场锚点，不参与加仓阶段推进"
    elif target_weight is None:
        conclusion = "暂停/降级"
        blocker = "缺少第一年配置目标，不能用证券账户仓位替代"
    elif current_weight is None:
        conclusion = "暂停/降级"
        blocker = "当前全资产仓位缺失，先核对持仓"
    elif remaining_weight is not None and remaining_weight <= 0:
        conclusion = "已达目标"
        blocker = "已达或超过第一年目标，不因强势继续加仓"
    elif quality_value is not None and quality_value < QUALITY_CORE_MIN and is_core_low_config_candidate(row):
        conclusion = "等待"
        blocker = f"核心候选质量分{quality_value:.1f}低于{QUALITY_CORE_MIN:.0f}分"
    elif temperature is not None and temperature >= EMOTION_PAUSE_LEVEL:
        conclusion = "暂停/降级"
        blocker = f"情绪温度{temperature:.0f}级，暂停新增"
    elif veto == "是" or signal == "红":
        conclusion = "暂停/降级"
        blocker = reason or "买点过滤器存在一票否决或红灯"
    elif ratio is not None and upper is not None and ratio >= upper and stage_verified == "是":
        conclusion = "允许下一档"
        blocker = "当前阶段已达标且买点验证通过"
    elif stage_verified == "是":
        conclusion = "补足当前阶段"
        blocker = "买点验证通过，但当前阶段尚未达上沿"
    elif is_core_low_config_candidate(row) and quality_value is not None and quality_value >= QUALITY_CORE_MIN and signal != "红":
        conclusion = "小额防踏空"
        blocker = "核心低配但买点未完整确认，只允许小额试探"
    else:
        conclusion = "等待"
        blocker = reason or "阶段验证未完成，年度缺口不是买入信号"

    if conclusion == "允许下一档":
        single_limit = min(remaining_weight or 0.0, 0.015)
        action_note = "打开下一阶段复核；仍优先回踩确认，不突破追满。"
    elif conclusion == "补足当前阶段":
        stage_room = max((target_weight or 0.0) * (upper or 0.0) - (current_weight or 0.0), 0.0)
        single_limit = min(remaining_weight or stage_room or 0.0, stage_room or 0.0, 0.010)
        action_note = "只补当前阶段上沿以内，不自动跨入下一档。"
    elif conclusion == "小额防踏空":
        single_limit = min(remaining_weight or 0.0, 0.005)
        action_note = "单次小额验证，买后必须设置确认位和撤销条件。"
    else:
        single_limit = 0.0
        action_note = "不新增；记录下一次触发条件。"

    return {
        "阶段状态": stage,
        "阶段目标下限": lower,
        "阶段目标上限": upper,
        "阶段占用率": ratio,
        "阶段验证": stage_verified,
        "阶段推进结论": conclusion,
        "下一档单次上限": single_limit,
        "阶段阻断原因": blocker,
        "阶段动作说明": action_note,
    }


def apply_stage_progression_fields(frame: pd.DataFrame) -> pd.DataFrame:
    if frame.empty:
        return frame.copy()
    result = frame.copy()
    stage_rows = [evaluate_stage_progression(row) for _, row in result.iterrows()]
    stage_df = pd.DataFrame(stage_rows, index=result.index)
    for column in stage_df.columns:
        result[column] = stage_df[column]
    return result


def build_stage_progression_view(buy_filter: pd.DataFrame, market_permission: str) -> pd.DataFrame:
    if buy_filter.empty:
        return pd.DataFrame(
            columns=[
                "Code", "Name", "资产角色", "年度配置项", "买点灯号", "质量评分", "情绪温度",
                "阶段状态", "阶段目标下限", "阶段目标上限", "阶段占用率", "阶段验证",
                "全资产当前仓位", "全资产目标仓位", "全资产剩余额度",
                "阶段推进结论", "下一档单次上限", "阶段阻断原因", "阶段动作说明", "市场权限",
            ]
        )
    result = apply_stage_progression_fields(buy_filter)
    result["市场权限"] = market_permission
    if market_permission != "开放买点复核":
        eligible = result["阶段推进结论"].isin(["允许下一档", "补足当前阶段", "小额防踏空"])
        result.loc[eligible, "阶段阻断原因"] = "市场权限未开放，阶段结论降为等待"
        result.loc[eligible, "阶段动作说明"] = "等待市场权限恢复后再复核。"
        result.loc[eligible, "下一档单次上限"] = 0.0
        result.loc[eligible, "阶段推进结论"] = "等待"

    rank = {
        "允许下一档": 0,
        "补足当前阶段": 1,
        "小额防踏空": 2,
        "等待": 3,
        "暂停/降级": 4,
        "已达目标": 5,
        "不适用": 6,
    }
    result["_stage_rank"] = result["阶段推进结论"].map(rank).fillna(9)
    result["_quality_rank"] = pd.to_numeric(result.get("质量评分"), errors="coerce").fillna(-1)
    result["_remaining_rank"] = pd.to_numeric(result.get("全资产剩余额度"), errors="coerce").fillna(-1)
    result = result.sort_values(
        ["_stage_rank", "_quality_rank", "_remaining_rank"],
        ascending=[True, False, False],
        kind="stable",
    ).drop(columns=["_stage_rank", "_quality_rank", "_remaining_rank"])
    columns = [
        "Code", "Name", "资产角色", "年度配置项", "买点灯号", "质量评分", "情绪温度",
        "阶段状态", "阶段目标下限", "阶段目标上限", "阶段占用率", "阶段验证",
        "全资产当前仓位", "全资产目标仓位", "全资产剩余额度",
        "阶段推进结论", "下一档单次上限", "阶段阻断原因", "阶段动作说明", "市场权限",
    ]
    return result[[column for column in columns if column in result.columns]].reset_index(drop=True)


def build_framework_rules(
    decision_inputs: dict[str, dict[str, str]],
    first_year: pd.DataFrame | None = None,
) -> pd.DataFrame:
    """把关键阈值直接写进工作簿，便于复核和版本迁移。"""
    cash_min = decision_input_number(decision_inputs, "cash_buffer_min") or 300000
    cash_target = decision_input_number(decision_inputs, "cash_buffer_target") or 500000
    rows = [
        ("总决策链", "执行顺序", "产业主线→主线周期→双锚系统→资产质量→量化验证层→爆量/回踩/量质吸结构→买点过滤器→情绪温度→认知防错层→仓位执行→复盘迭代", "V2.8.5-QM 主线周期 + 量化验证 + 认知防错增强版"),
        ("主线周期", "六阶段", MAINLINE_CYCLE_STAGES, "V2.8.5 第五部分"),
        ("量化验证层", "模块定位", "主观框架定方向；量化模块负责验证信号、减少情绪误判、提高执行纪律和风险监控", QUANT_MODULE_SOURCE),
        ("量化验证层", "正期望原则", "期望收益=胜率×平均盈利-失败率×平均亏损；不是固定追求高胜率低赔率，而是追求经复核后的正期望", QUANT_MODULE_SOURCE),
        ("量化验证层", "三项职责", "数据验证层：指数环境/ETF买点/情绪温度；执行纪律层：所有买点打分；风险监控层：退潮、赎回、放量滞涨、次日验证", QUANT_MODULE_SOURCE),
        ("量化验证层", "模型边界", "用模型约束人脑而非替代人脑；人脑提出假设→数据验证假设→模型执行规则→复盘修正模型", QUANT_MODULE_SOURCE),
        ("量化验证层", "产品边界", "ETF主框架可叠加指数增强/Smart Beta作为低波权益补充；个人不做高频、不碰复杂黑箱、不让量化替代主线判断", QUANT_MODULE_SOURCE),
        ("量化验证层", "10分制动作", "8-10标准买点；6-7黄灯观察；4-5不买；低于4风险复审；资金流入不能单独触发买入，必须价格结构确认", QUANT_MODULE_SOURCE),
        ("量化验证层", "核心ETF六项", "分时结构0-2；量价关系0-2；ETF份额0-2；折溢价0-1；龙头同步0-2；次日验证0-1", QUANT_MODULE_SOURCE),
        ("资产角色", "核心成长ETF", "产业逻辑决定方向；双锚决定权限；买点过滤器决定能不能买；量化打分决定买多少；只买缩量回踩不破+次日转强", QUANT_MODULE_SOURCE),
        ("资产角色", "低波权益ETF", "A500/红利低波按配置节奏分批，不追强；用于降低组合波动，不当现金安全垫", QUANT_MODULE_SOURCE),
        ("资产角色", "红利现金流个股", "红利个股归入低波现金流层；买稳定分红和现金流，不买高息幻觉；不能替代半导体/AI硬件核心成长仓", DIVIDEND_STOCK_SOURCE),
        ("红利个股", "行业优先级", "优先公用事业/运营商/水电，其次国有大行，再看能源资源红利，质量消费红利作为增强；不按股息率从高到低机械排序", DIVIDEND_STOCK_SOURCE),
        ("红利个股", "买入绿灯", "分红连续性至少5年稳定分红；分红率30%–75%；经营现金流覆盖分红；盈利不连续明显下滑；股息率安全垫达标；估值历史中低；缩量回踩止跌", DIVIDEND_STOCK_SOURCE),
        ("红利个股", "高息陷阱否决", "分红率>100%、高息源于股价暴跌、盈利/现金流恶化、举债分红、资本开支压力、周期盈利高点或只因快分红而买，一票否决", DIVIDEND_STOCK_SOURCE),
        ("红利个股", "仓位节奏", "红利个股+红利低波初始3%–5%；第一阶段6%–8%；稳定后10%–15%；单只A档初始0.5%–1%，B档0.3%–0.8%", DIVIDEND_STOCK_SOURCE),
        ("红利个股", "复审退出", "跌10%复核风格；跌15%强制复审分红逻辑；跌20%且利润/现金流恶化至少减半；削减分红或现金流低于分红则降级/退出", DIVIDEND_STOCK_SOURCE),
        ("资产角色", "弹性低波ETF", "可转债ETF用于缓冲与反弹弹性；每次小额分批，不在转债大涨日追", QUANT_MODULE_SOURCE),
        ("资产角色", "纯防御ETF", "国债ETF只做安全垫和保险丝；可以与现金共同计入防御层，不因短期跑输科技而频繁切换", QUANT_MODULE_SOURCE),
        ("资产角色", "降级/遗留仓", "云计算、绿电、中望等不补仓；反弹压缩或退出复审，释放资金给核心主线与现金等待买点", QUANT_MODULE_SOURCE),
        ("资产角色", "比亚迪2030战略观察仓", "现有9–10万元作为底仓；基本面确认、估值安全垫、技术买点同时通过后才小额复核；合格加仓后全资产上限3%", "V2.8.5比亚迪2030战略配置完整稿"),
        ("资产角色", "卫星验证个股", "个股不能替代ETF主仓；只做小仓验证，首仓0.3%-0.5%，必须先定义证伪和最大损失", QUANT_MODULE_SOURCE),
        ("趋势买点", "启用条件", "双锚至少黄绿；标的属于主线或核心观察方向；质量评分≥8；情绪不进入狂热；无一票否决", "V2.8.5 第九部分"),
        ("趋势买点", "优先级", TREND_BUY_PRIORITY, "V2.8.5 第九部分"),
        ("爆量/量质吸", "模块定位", "不能单独产生买入信号；必须服从主线周期、双锚、资产质量、买点过滤、情绪和仓位纪律", V285_SUPPLEMENT_SOURCE),
        ("爆量/量质吸", "核心路径", VOLUME_ABSORPTION_SEQUENCE, V285_SUPPLEMENT_SOURCE),
        ("爆量突破", "启用条件", "核心主线；双锚至少黄绿；质量评分≥8；情绪≤4；份额/资金或成交额确认；突破平台/压力位", V285_SUPPLEMENT_SOURCE),
        ("健康爆量", "执行上限", "爆量突破只允许小仓试探：ETF全资产0.5%–1%；个股全资产0.3%–0.5%；不得一次补满目标仓位", V285_SUPPLEMENT_SOURCE),
        ("回踩确认", "最佳买点", "爆量后缩量回踩；不破爆量K低点/5日线/10日线；龙头不破；ETF份额未连续净赎回；次日转强", V285_SUPPLEMENT_SOURCE),
        ("回踩确认", "执行上限", "核心ETF≤1%；强确认核心ETF1%–2%；个股≤0.5%；高弹性个股≤0.3%；后排个股不买", V285_SUPPLEMENT_SOURCE),
        ("量质吸", "成立条件", "核心主线、双锚未红、前期放量、回踩缩量、不再创新低、龙头不破、份额未连续赎回、次日红K/放量转强；≥5观察，≥6候选", V285_SUPPLEMENT_SOURCE),
        ("爆量失败", "复审/停止", "低开低走、跌破爆量K低点、放量滞涨/下跌、龙头分化、ETF连续赎回等；1项暂停新增，2项战术仓复审，3项以上核心仓降温", V285_SUPPLEMENT_SOURCE),
        ("关键支撑", "ETF/个股", "ETF优先爆量阳线低点、5日线、10日线、平台上沿；个股优先爆量K低点、平台下沿、5日线、10日线、前一日低点", V285_SUPPLEMENT_SOURCE),
        ("认知防错层", "模块定位", "不推翻主线周期、双锚、趋势买点、买点过滤、情绪温度和仓位纪律；只用于防止情绪、偏误、沉没成本和机会成本误导", MENTAL_MODEL_SOURCE),
        ("认知防错层", "认知防错八问", "机会成本是什么；是否只看支持观点的信息；是否被近期新闻/涨幅影响；是否因亏损不愿卖；最大亏损是多少；是否有更优全资产选择；是否在能力圈内；10个月后是否仍认可", MENTAL_MODEL_SOURCE),
        ("认知防错层", "两项不清则不买", "机会成本、反证、最大损失、能力圈边界等核心问题中有两项回答不清，禁止新增", MENTAL_MODEL_SOURCE),
        ("思维模型", "机会成本", "买入前必须比较159516、159381、A500/可转债、现金等待买点等替代选择；新增资金优先级默认159516 > 159381 > A500/可转债 > 机器人 > 比亚迪2030观察 > 云计算/非核心", MENTAL_MODEL_SOURCE),
        ("思维模型", "局部最优/全局最优", "证券账户局部最优不等于480万全资产全局最优；所有仓位判断以全资产穿透口径为最终标准", MENTAL_MODEL_SOURCE),
        ("思维模型", "决策树", "每个核心标的至少给出强势延续、缩量回踩、失败转弱三种情景与动作", MENTAL_MODEL_SOURCE),
        ("思维模型", "沉没成本", "亏损不是继续持有或加仓的理由；只有主线、资产质量、买点重新通过，才有资格继续配置", MENTAL_MODEL_SOURCE),
        ("思维模型", "易得性/确认偏误", "利好后必须找相反数据：ETF净赎回、龙头利好不涨、情绪过热、信息是否已被价格反映", MENTAL_MODEL_SOURCE),
        ("思维模型", "损失规避/前景理论", "盈利仓看趋势，亏损仓看逻辑，不用盈亏金额决定买卖，避免亏损仓越跌越补", MENTAL_MODEL_SOURCE),
        ("思维模型", "风险概率", "买入前同时判断胜率、赔率、跌破关键位损失和全资产影响；逻辑好但赔率差不买", MENTAL_MODEL_SOURCE),
        ("思维模型", "反向失败", "先研究它怎么失败，再决定是否买；退潮三因子与失败路径优先于乐观叙事", MENTAL_MODEL_SOURCE),
        ("思维模型", "反脆弱", "核心成长负责收益，低波权益负责稳定，可转债负责弹性缓冲，现金/国债负责回踩买点弹药", MENTAL_MODEL_SOURCE),
        ("思维模型", "安全边际/冗余/断裂点", "不在情绪温度4级以上追高；保留现金国债可转债；个股跌破爆量K低点复审，战术仓止损", MENTAL_MODEL_SOURCE),
        ("思维模型", "能力圈", "AI基础设施、半导体设备、机器人ETF、AI硬件ETF、ETF资金/份额/买点过滤、全资产仓位管理为当前能力圈；圈外只能观察", MENTAL_MODEL_SOURCE),
        ("思维模型", "复利", "提高决策质量而非交易频率；长期主线、正确仓位、持续复盘、少犯大错、关键买点加核心仓", MENTAL_MODEL_SOURCE),
        ("主线退潮", "三因子", f"{RETREAT_THREE_FACTORS}；三者共振时从进攻切换为防守", "V2.8.5 第十二部分"),
        ("全资产目标", "三年目标结构", "核心成长45%-50%；低波权益15%-20%；弹性低波8%-12%；纯防御20%-25%；战术机会5%-8%", "V2.8.5 第十三部分"),
        ("全资产目标", "第一年结构迁移", "核心成长30%-35%；低波权益/混合18%-22%；弹性低波8%-10%；纯防御35%-40%；战术仓3%-5%", "V2.8.5 第十四部分"),
        ("仓位口径", "买入/减仓主口径", "以总资产/第一年配置要求为准；证券账户仓位只作交易集中度参考，不替代年度全资产目标", "第一年配置方案和建议.docx"),
        ("质量准入", "核心候选", f"完整评分 ≥ {QUALITY_CORE_MIN:.0f}", "V2.8.5 第三部分"),
        ("质量准入", "观察候选", f"{QUALITY_OBSERVE_MIN:.0f} ≤ 完整评分 < {QUALITY_CORE_MIN:.0f}", "V2.8.5 第三部分"),
        ("质量准入", "禁止新增", f"完整评分 < {QUALITY_OBSERVE_MIN:.0f}；或缺少完整评分/证据", "V2.8.5 第三部分"),
        ("买点过滤", "标准首批", f"六项中通过 ≥ {BUY_STANDARD_MIN}，且无一票否决", "V2.8.5 第六部分"),
        ("买点过滤", "半额首批", f"六项中通过 = {BUY_HALF_MIN}，且无一票否决", "V2.8.5 第六部分"),
        ("买点过滤", "与爆量/量质吸关系", "先判断趋势结构是否存在，再用买点过滤器确认是否可以买，最后用情绪温度和仓位系统决定买多少", V285_SUPPLEMENT_SOURCE),
        (
            "买点过滤",
            "建议买入区间",
            "完整近期行情下按昨日低点与日内回踩区间复核；仓位已满仅保留观察区间；不自动下单",
            "V2.8.4/V2.8.5 买点与行为纪律",
        ),
        ("当前核心方向", "159516 半导体设备ETF", "第一核心候选；连续上涨不追；缩量回踩5/10日且龙头不破、份额不连续赎回、次日转强后分批", V285_SUPPLEMENT_SOURCE),
        ("当前核心方向", "159381 创业板人工智能ETF", "AI硬件增强仓；不追研报催化高开；等光模块/CPO/AI芯片龙头缩量企稳后的回踩确认", V285_SUPPLEMENT_SOURCE),
        ("当前核心方向", "562500 机器人ETF", "核心持有仓；当前不因回调直接补仓；必须等放量下跌结束、缩量企稳、不再创新低、次日红K确认", V285_SUPPLEMENT_SOURCE),
        ("当前核心方向", "AI/半导体个股篮子", "资产质量≥8、主线未退潮、爆量后缩量回踩、不破爆量K低点、无放量滞涨、情绪≤4；首仓0.3%–0.5%", V285_SUPPLEMENT_SOURCE),
        ("当前核心方向", "比亚迪2030战略观察仓", "不是第一主线进攻仓；跟踪海外销量、净利率、智能化、高端品牌和现金流；不做T不追涨，不挤占核心ETF预算", "V2.8.5比亚迪2030战略配置完整稿"),
        ("当前持仓影响", "中望软件/绿电/云计算", "沉没成本不构成补仓理由；非核心仓位只做反弹压缩或退出复审，释放资金给核心主线和现金等待买点", MENTAL_MODEL_SOURCE),
        ("当前持仓影响", "比亚迪", "从普通遗留仓单列为2030战略观察仓；亏损不是补仓理由，只有三重确认通过后才小额复核", "V2.8.5比亚迪2030战略配置完整稿"),
        ("情绪纪律", "暂停新增", f"情绪温度 ≥ {EMOTION_PAUSE_LEVEL}", "V2.8.5 第八部分"),
        ("执行纪律", "单日动作上限", "一个买入方向 + 两个卖出/减仓方向", "V2.8.4/V2.8.5"),
        ("现金安全垫", "最低值", f"{cash_min:,.0f}元", "V2.8.4 第十三部分"),
        ("现金安全垫", "目标值", f"{cash_target:,.0f}元", "V2.8.4 第十三部分"),
    ]
    for theme, limit in THEME_LIMITS.items():
        rows.append(("产业链簇上限", theme, f"≤ {limit:.0%}", "V2.8.5 第四部分"))
    if first_year is not None and not first_year.empty:
        for _, item in first_year.iterrows():
            rows.append(
                (
                    "第一年全资产配置",
                    str(item.get("配置项", "")),
                    f"目标{to_float(item.get('年度目标占比')) or 0:.0%}；收益观察区间{to_float(item.get('目标收益下限')) or 0:.0%}-{to_float(item.get('目标收益上限')) or 0:.0%}",
                    "1st年配置表.xlsx；缺口不是买入信号",
                )
            )
    return pd.DataFrame(rows, columns=["规则模块", "规则", "阈值/定义", "来源"])


def build_checks(
    watchlist: pd.DataFrame,
    positions: pd.DataFrame,
    positions_sheet: pd.DataFrame,
    account_meta: dict[str, float],
    market_data: pd.DataFrame,
    quality_score: pd.DataFrame,
    first_year: pd.DataFrame,
) -> pd.DataFrame:
    """生成可审计的输入与结果校验表。"""
    market_value_sum = pd.to_numeric(positions.get("Market Value"), errors="coerce").fillna(0).sum()
    weight_sum = pd.to_numeric(positions.get("Weight"), errors="coerce").fillna(0).sum()
    broker_market_value = to_float(account_meta.get("broker_market_value"))
    position_ratio = to_float(account_meta.get("position_ratio"))
    missing_quotes = int(missing_latest_mask(market_data["Latest"]).sum())
    interface_not_ready = int(
        market_data.get("接口状态", pd.Series(index=market_data.index, dtype=object))
        .astype(str)
        .str.contains("未刷新|失败|缺失", regex=True, na=False)
        .sum()
    )
    quality_missing = int((quality_score["质量状态"] == "数据不足，需人工评分").sum())
    watch_codes = watchlist["Code"].astype(str)
    position_codes = positions["Code"].astype(str)
    rows: list[dict[str, object]] = []

    def add(name, actual, expected, difference, tolerance, status, note):
        rows.append({"检查项": name, "实际": actual, "期望": expected, "差异": difference, "容差": tolerance, "状态": status, "修复建议": note})

    mv_diff = market_value_sum - broker_market_value if broker_market_value is not None else None
    add("持仓市值合计", market_value_sum, broker_market_value, mv_diff, 1.0, "OK" if mv_diff is not None and abs(mv_diff) <= 1 else "检查", "核对截图总市值与各持仓市值")
    weight_diff = weight_sum - position_ratio if position_ratio is not None else None
    add("持仓权重合计", weight_sum, position_ratio, weight_diff, 0.001, "OK" if weight_diff is not None and abs(weight_diff) <= 0.001 else "检查", "单只权重四舍五入可能产生小差异")
    add("行情完整性", missing_quotes, 0, missing_quotes, 0, "OK" if missing_quotes == 0 else "检查", "联网重跑或核对缺失代码；详见Market_Data接口状态")
    add("行情接口诊断", interface_not_ready, 0, interface_not_ready, 0, "OK" if interface_not_ready == 0 else "检查", "查看Market_Data失败原因与重试结果，禁止用旧行情替代")
    add("观察池重复代码", int(watch_codes.duplicated().sum()), 0, int(watch_codes.duplicated().sum()), 0, "OK" if not watch_codes.duplicated().any() else "检查", "删除重复观察项")
    add("持仓重复代码", int(position_codes.duplicated().sum()), 0, int(position_codes.duplicated().sum()), 0, "OK" if not position_codes.duplicated().any() else "检查", "合并重复持仓")
    add("质量评分缺失", quality_missing, 0, quality_missing, 0, "待补" if quality_missing else "OK", "在watchlist.csv填写Manual Quality Score与Quality Evidence；不依赖行情接口")
    total_assets = to_float(account_meta.get("total_assets"))
    account_total = to_float(account_meta.get("account_total"))
    add("全资产口径", total_assets, ">=证券账户总资产", None, 0, "OK" if total_assets and account_total and total_assets >= account_total else "检查", "核对total_assets与account_total口径")
    if not first_year.empty:
        target_weight_sum = pd.to_numeric(first_year["年度目标占比"], errors="coerce").fillna(0).sum()
        source_current_sum = pd.to_numeric(first_year["配置表当前金额约"], errors="coerce").fillna(0).sum()
        latest_current_sum = pd.to_numeric(first_year["最新持仓金额"], errors="coerce").fillna(0).sum()
        current_diff = latest_current_sum - source_current_sum
        add("第一年目标比例合计", target_weight_sum, "<=100%", None, 0, "OK" if target_weight_sum <= 1 else "检查", "配置表当前规划58%，其余应明确为现金/防御/未配置")
        add("配置表当前金额复核", latest_current_sum, source_current_sum, current_diff, 2000, "OK" if abs(current_diff) <= 2000 else "检查", "来源表金额以万元四舍五入；差异过大时核对代码映射")
        mapped_codes: list[str] = []
        for value in first_year["映射代码"].astype(str):
            mapped_codes.extend([format_code(code) for code in value.split("、") if code and code != "待选"])
        duplicate_mapped = len(mapped_codes) - len(set(mapped_codes))
        add("年度配置代码重复映射", duplicate_mapped, 0, duplicate_mapped, 0, "OK" if duplicate_mapped == 0 else "检查", "同一持仓不得重复计入多个年度配置项")
    return pd.DataFrame(rows)


# ==================== 双锚灯号（第一版简化规则） ====================
def evaluate_anchor_light(latest, pct_chg, low, prev_day_low) -> tuple[str, str]:
    """
    双锚灯号第一版简化规则：
    1. 跌破昨日低点：红灯；
    2. 跌幅超过 1.5%：黄灯偏弱；
    3. 当日未创新低且涨跌幅为正：黄绿；
    4. 其他情况：观察。
    """
    latest_f = to_float(latest)
    pct_f = to_float(pct_chg)
    low_f = to_float(low)
    prev_day_low_f = to_float(prev_day_low)

    if latest_f is None or pct_f is None:
        return "待确认", "缺少最新价或涨跌幅，无法判灯"

    if prev_day_low_f is not None and latest_f < prev_day_low_f:
        return "红灯", "最新价跌破昨日低点"

    if pct_f <= -1.5:
        return "黄灯偏弱", "指数跌幅超过1.5%"

    if low_f is not None and latest_f > low_f and pct_f > 0:
        note = "当日未创新低，且涨跌幅为正"
        if prev_day_low_f is None:
            note += "；昨日低点缺失，红灯规则待补"
        return "黄绿", note

    if prev_day_low_f is None:
        return "观察", "昨日低点缺失；未触发黄绿或黄灯偏弱"
    return "观察", "未触发黄绿、黄灯偏弱或红灯"


def summarize_anchor_group(lights: list[str]) -> str:
    """把一组锚点灯号压缩为组灯号。"""
    if any(light == "红灯" for light in lights):
        return "红灯"
    if any(light == "黄灯偏弱" for light in lights):
        return "黄灯偏弱"
    if lights and all(light == "黄绿" for light in lights):
        return "黄绿"
    if any(light == "黄绿" for light in lights):
        return "黄绿"
    return "观察"


def summarize_double_anchor(total_light: str, growth_light: str) -> tuple[str, str]:
    """根据总量锚和成长锚输出综合灯号。"""
    if growth_light == "红灯":
        return "成长红", "禁止成长新买入"
    if total_light == "黄绿" and growth_light == "黄绿":
        return "双绿", "允许研究买入，但仍需买点过滤器"
    if growth_light == "黄绿":
        return "黄绿", "仅试探"
    return "观察", "不满足双绿或黄绿条件，继续观察"


def build_double_anchor(market_data: pd.DataFrame) -> pd.DataFrame:
    """生成 Double_Anchor 工作表。"""
    market_by_code = market_data.set_index("Code")
    rows: list[dict] = []
    total_lights: list[str] = []
    growth_lights: list[str] = []

    for group_name, anchors, group_lights in [
        ("总量锚", TOTAL_ANCHORS, total_lights),
        ("成长锚", GROWTH_ANCHORS, growth_lights),
    ]:
        for code, fallback_name in anchors:
            quote = market_by_code.loc[code] if code in market_by_code.index else None
            latest = quote["Latest"] if quote is not None else None
            pct_chg = quote["PctChg"] if quote is not None else None
            low = quote["Low"] if quote is not None else None
            prev_day_low = quote["PrevDayLow"] if quote is not None else None
            light, note = evaluate_anchor_light(latest, pct_chg, low, prev_day_low)
            group_lights.append(light)

            rows.append(
                {
                    "锚点": group_name,
                    "Code": code,
                    "指数名称": quote["Name"] if quote is not None else fallback_name,
                    "最新价": latest,
                    "涨跌幅": pct_chg,
                    "当日最低": low,
                    "昨日低点": prev_day_low,
                    "灯号": light,
                    "说明": note,
                }
            )

    total_group_light = summarize_anchor_group(total_lights)
    growth_group_light = summarize_anchor_group(growth_lights)
    combined_light, combined_note = summarize_double_anchor(total_group_light, growth_group_light)

    rows.extend(
        [
            {
                "锚点": "总量锚综合",
                "Code": "",
                "指数名称": "上证综指 / 沪深300 / 中证A500",
                "最新价": "",
                "涨跌幅": "",
                "当日最低": "",
                "昨日低点": "",
                "灯号": total_group_light,
                "说明": "总量锚观察：上证综指、沪深300、中证A500",
            },
            {
                "锚点": "成长锚综合",
                "Code": "",
                "指数名称": "科创50 / 创业板指",
                "最新价": "",
                "涨跌幅": "",
                "当日最低": "",
                "昨日低点": "",
                "灯号": growth_group_light,
                "说明": "成长锚观察：科创50、创业板指",
            },
            {
                "锚点": "综合灯号",
                "Code": "",
                "指数名称": "双锚综合",
                "最新价": "",
                "涨跌幅": "",
                "当日最低": "",
                "昨日低点": "",
                "灯号": combined_light,
                "说明": combined_note,
            },
        ]
    )

    return pd.DataFrame(rows)


# ==================== 持仓检查 ====================
def build_positions_sheet(
    positions: pd.DataFrame,
    market_data: pd.DataFrame,
    account_meta: dict[str, float],
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    生成 Positions 工作表，并计算：
    - 最新价、市值、盈亏、证券账户仓位与全资产穿透仓位
    - 证券账户目标只作为集中度参考，不直接触发买入/减仓建议
    """
    market_by_code = market_data.set_index("Code")
    account_total = account_meta.get("account_total", 0)
    total_assets = account_meta.get("total_assets", 0)

    rows: list[dict] = []
    alerts: list[dict] = []

    for _, item in positions.iterrows():
        code = str(item["Code"]).strip()
        shares = to_float(item.get("Shares")) or 0
        available_shares = to_float(item.get("Available Shares"))
        cost = to_float(item.get("Cost"))
        broker_latest = to_float(item.get("Latest"))
        broker_market_value = to_float(item.get("Market Value"))
        broker_pl = to_float(item.get("P/L"))
        broker_pl_pct = to_float(item.get("P/L%"))
        today_pl = to_float(item.get("Today P/L"))
        today_pl_pct = to_float(item.get("Today P/L%"))
        buy_avg = to_float(item.get("Buy Avg"))
        broker_weight = to_float(item.get("Weight"))
        target_weight = item.get("Target Weight", 0)
        role = item.get("Role", "")
        action = item.get(f"{FRAMEWORK_VERSION} Action", item.get("V2.8.4 Action", "持有/观察"))

        market_latest = None
        market_source = ""
        if code in market_by_code.index:
            market_row = market_by_code.loc[code]
            market_latest = to_float(market_row["Latest"])
            market_source = str(market_row.get("DataSource", "")).strip()

        latest = first_valid(market_latest, broker_latest)
        latest_source = market_source if market_latest is not None and market_source else "券商截图"

        market_value = None
        pl = None
        pl_pct = None
        weight = None

        try:
            if latest is not None:
                latest = float(latest)
                market_value = round(shares * latest, 2)
                if cost is not None:
                    pl = round((latest - cost) * shares, 2)
                    pl_pct = round((latest - cost) / cost, 4) if cost else 0
                if account_total:
                    weight = round(market_value / float(account_total), 4)
        except (TypeError, ValueError):
            pass

        market_value = first_valid(broker_market_value, market_value)
        pl = first_valid(broker_pl, pl)
        pl_pct = first_valid(broker_pl_pct, pl_pct)
        weight = first_valid(broker_weight, weight)

        discipline = action
        broker_concentration_note = ""
        target_weight_f = to_float(target_weight)
        if target_weight_f is not None and weight is not None:
            try:
                if float(weight) > target_weight_f + 0.005:
                    broker_concentration_note = f"证券账户仓位 {weight:.2%} 高于参考目标 {target_weight_f:.2%}；仅作集中度参考，减仓以全资产第一年配置为准"
            except (TypeError, ValueError):
                pass

        rows.append(
            {
                "Code": code,
                "Name": item.get("Name", ""),
                "Shares": shares,
                "Available Shares": available_shares,
                "Cost": cost,
                "Latest": latest,
                "Latest Source": latest_source,
                "Market Value": market_value,
                "P/L": pl,
                "P/L%": pl_pct,
                "Today P/L": today_pl,
                "Today P/L%": today_pl_pct,
                "Buy Avg": buy_avg,
                "Weight": weight,
                "Full Asset Weight": round(float(market_value) / float(total_assets), 4) if market_value is not None and total_assets else None,
                "Broker Latest": broker_latest,
                "Broker Market Value": broker_market_value,
                "Broker Weight": broker_weight,
                "Role": role,
                "Target Weight": target_weight,
                "Weight Gap": round(float(weight) - target_weight_f, 4) if weight is not None and target_weight_f is not None else None,
                "证券账户集中度提示": broker_concentration_note,
                f"{FRAMEWORK_VERSION} Action": discipline,
            }
        )

    positions_df = pd.DataFrame(rows)
    alerts_df = pd.DataFrame(alerts)
    return positions_df, alerts_df


# ==================== 持仓纪律提醒 ====================
def is_high_volume_pullback(latest, high, low, pct_chg, amount, avg20_amount) -> bool:
    """
    机器人ETF“高位放量回落”的第一版简化识别。
    后续接入20日均额后，可以把 amount / avg20_amount 作为主条件。
    """
    latest_f = to_float(latest)
    high_f = to_float(high)
    low_f = to_float(low)
    pct_f = to_float(pct_chg)
    amount_f = to_float(amount)
    avg20_f = to_float(avg20_amount)

    if latest_f is None or high_f is None or low_f is None or pct_f is None:
        return False
    if high_f <= low_f:
        return False

    intraday_position = (latest_f - low_f) / (high_f - low_f)
    if avg20_f is None or avg20_f <= 0 or amount_f is None:
        return False

    volume_confirmed = amount_f >= avg20_f * 1.2
    return pct_f < 0 and intraday_position <= 0.35 and volume_confirmed


def build_positions_action(positions: pd.DataFrame, market_data: pd.DataFrame) -> pd.DataFrame:
    """生成 Positions_Action 工作表，只做纪律提醒，不自动交易。"""
    market_by_code = market_data.set_index("Code")
    rows: list[dict] = []

    for _, item in positions.iterrows():
        code = str(item["Code"]).strip()
        name = item.get("Name", "")
        quote = market_by_code.loc[code] if code in market_by_code.index else None

        latest = quote["Latest"] if quote is not None else None
        high = quote["High"] if quote is not None else None
        low = quote["Low"] if quote is not None else None
        pct_chg = quote["PctChg"] if quote is not None else None
        amount = quote["Amount"] if quote is not None else None
        avg20_amount = quote["Avg20Amount"] if quote is not None else None

        category = "未纳入专项纪律"
        trigger = "否"
        action = "按原持仓规则观察"
        review = "否"
        note = ""

        if code == "510210":
            category = "宽基重复"
            price = to_float(latest)
            if price is None:
                action = "行情缺失，暂不触发"
                note = "上证ETF价格达到1.01-1.03元时，提示卖出20000份"
            elif 1.01 <= price <= 1.03:
                trigger = "是"
                action = "提示卖出20000份"
                note = "上证ETF属于宽基重复，价格进入1.01-1.03元纪律区间"
            else:
                action = "暂不卖出"
                note = "等待1.01-1.03元纪律区间"

        elif code == "562550":
            category = "非核心新能源"
            price = to_float(latest)
            if price is None:
                action = "行情缺失，暂不触发"
                note = "绿电ETF价格达到1.30元以上时，提示卖出10000份"
            elif price >= 1.30:
                trigger = "是"
                action = "提示卖出10000份"
                note = "绿电ETF属于非核心新能源，价格达到1.30元以上"
            else:
                action = "暂不卖出"
                note = "未达到1.30元纪律价格"

        elif code == "562500":
            category = "核心持有"
            if is_high_volume_pullback(latest, high, low, pct_chg, amount, avg20_amount):
                trigger = "是"
                review = "是"
                action = "提示复审，不自动卖"
                note = "机器人ETF为核心持有；疑似高位放量回落，仅复审"
            else:
                action = "持有不加"
                note = "机器人ETF为核心持有；未确认高位放量回落，持有不加"

        if not note:
            note = "未命中专项纪律，按原持仓规则观察"

        rows.append(
            {
                "Code": code,
                "Name": name,
                "纪律分类": category,
                "Latest": latest,
                "PctChg": pct_chg,
                "High": high,
                "Low": low,
                "Amount": amount,
                "Avg20Amount": avg20_amount,
                "触发提醒": trigger,
                "动作建议": action,
                "是否复审": review,
                "说明": note,
            }
        )

    return pd.DataFrame(rows)


def build_broker_snapshot(positions: pd.DataFrame) -> pd.DataFrame:
    """把券商截图录入的原始持仓数据写入 Broker_Snapshot 工作表。"""
    columns = [
        "Code",
        "Name",
        "Shares",
        "Available Shares",
        "Latest",
        "Cost",
        "Market Value",
        "P/L",
        "P/L%",
        "Today P/L",
        "Today P/L%",
        "Buy Avg",
        "Weight",
        "Role",
        "Target Weight",
        "V2.8.5 Action",
        "V2.8.4 Action",
    ]
    existing_columns = [column for column in columns if column in positions.columns]
    return positions[existing_columns].copy()


def build_dashboard(
    account_meta: dict[str, float],
    alerts: pd.DataFrame,
    market_data: pd.DataFrame,
    double_anchor: pd.DataFrame,
    emotion: pd.DataFrame,
    buy_filter: pd.DataFrame,
    quality_score: pd.DataFrame,
    positions_sheet: pd.DataFrame,
    decision_inputs: dict[str, dict[str, str]],
    first_year_summary: dict[str, float],
) -> pd.DataFrame:
    """生成 Dashboard 摘要页。"""
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    source_series = market_data.get("DataSource", pd.Series(index=market_data.index, dtype=object)).astype(str)
    interface_series = market_data.get("接口状态", pd.Series(index=market_data.index, dtype=object)).astype(str)
    interface_not_ready = interface_series.str.contains("未刷新|缺失|失败", regex=True, na=False)
    missing_count = ((~source_series.isin(LIVE_QUOTE_SOURCES)) | missing_latest_mask(market_data["Latest"]) | interface_not_ready).sum()
    quote_times = market_data.get("QuoteTime", pd.Series(index=market_data.index, dtype=object)).dropna().astype(str)
    market_as_of = quote_times.max() if not quote_times.empty else "未取到当日接口时间"
    source_summary = " / ".join(
        f"{source}:{count}"
        for source, count in source_series.value_counts().sort_index().items()
    )
    interface_summary = " / ".join(
        f"{status}:{count}"
        for status, count in interface_series.value_counts().sort_index().items()
        if status and status.lower() != "nan"
    ) or "未生成接口诊断"

    anchor_row = double_anchor[double_anchor["锚点"] == "综合灯号"]
    anchor_light = anchor_row.iloc[0]["灯号"] if not anchor_row.empty else "待确认"
    temp_row = emotion[emotion["指标"] == "情绪温度"]
    temperature = int(to_float(temp_row.iloc[0]["数值"]) or 1) if not temp_row.empty else None
    emotion_reliability = str(temp_row.iloc[0].get("可靠性", "待确认")) if not temp_row.empty else "待确认"
    green_count = int((buy_filter["买点灯号"] == "绿").sum()) if not buy_filter.empty else 0
    yellow_count = int((buy_filter["买点灯号"] == "黄").sum()) if not buy_filter.empty else 0
    dividend_stock_count = int((buy_filter.get("资产角色", pd.Series(dtype=object)).astype(str) == "红利现金流个股").sum()) if not buy_filter.empty else 0
    quality_ready = int(quality_score["质量状态"].isin(["核心候选", "观察候选"]).sum()) if not quality_score.empty else 0
    quality_missing = int((quality_score["质量状态"] == "数据不足，需人工评分").sum()) if not quality_score.empty else 0
    account_total = to_float(account_meta.get("account_total"))
    total_assets = to_float(account_meta.get("total_assets"))
    broker_market_value = to_float(account_meta.get("broker_market_value"))
    full_asset_equity = broker_market_value / total_assets if broker_market_value is not None and total_assets else None
    available_cash = to_float(account_meta.get("available_cash")) or 0
    government_bond_value = 0.0
    if not positions_sheet.empty:
        bond_rows = positions_sheet[positions_sheet["Name"].astype(str).str.contains("国债", na=False)]
        government_bond_value = pd.to_numeric(bond_rows.get("Market Value"), errors="coerce").fillna(0).sum()
    liquidity_buffer = available_cash + government_bond_value
    cash_min = decision_input_number(decision_inputs, "cash_buffer_min") or 300000
    cash_target = decision_input_number(decision_inputs, "cash_buffer_target") or 500000
    cash_status = "达标" if liquidity_buffer >= cash_min else "低于最低安全垫"

    rows = [
        {"项目": "版本", "内容": f"{FRAMEWORK_DISPLAY_VERSION} {PRODUCT_NAME}"},
        {"项目": "生成时间", "内容": now},
        {"项目": "行情数据时间", "内容": market_as_of},
        {"项目": "行情来源统计", "内容": source_summary},
        {"项目": "行情接口状态摘要", "内容": interface_summary},
        {"项目": "说明", "内容": "本程序只做纪律提醒，不连接券商、不下单、不自动交易。"},
        {"项目": "量化验证层", "内容": "V2.8.5-Q：主观定方向，量化验信号，纪律控仓位，复盘修正模型"},
        {"项目": "红利现金流层", "内容": f"已纳入{dividend_stock_count}只红利个股观察；定位为低波现金流层，不替代核心成长收益发动机"},
        {"项目": "双锚综合", "内容": anchor_light},
        {"项目": "情绪温度", "内容": temperature if temperature is not None else "待确认"},
        {"项目": "情绪数据可靠性", "内容": emotion_reliability},
        {"项目": "标准首批候选数", "内容": green_count},
        {"项目": "半额首批复核数", "内容": yellow_count},
        {"项目": "质量门槛可用标的数", "内容": quality_ready},
        {"项目": "质量评分待补标的数", "内容": quality_missing},
        {"项目": "第一年配置目标占比", "内容": first_year_summary.get("年度目标占比", "未填写")},
        {"项目": "第一年当前已配置占比", "内容": first_year_summary.get("年度当前全资产占比", "未填写")},
        {"项目": "第一年配置完成率", "内容": first_year_summary.get("年度完成率", "未填写")},
        {"项目": "第一年动态目标金额", "内容": first_year_summary.get("年度目标金额", "未填写")},
        {"项目": "第一年当前映射金额", "内容": first_year_summary.get("年度当前金额", "未填写")},
        {"项目": "第一年资金缺口", "内容": first_year_summary.get("年度资金缺口", "未填写")},
        {"项目": "第一年未配置目标占比", "内容": first_year_summary.get("未配置目标占比", "未填写")},
        {"项目": "证券账户总资产", "内容": account_meta.get("account_total", "未填写")},
        {"项目": "全资产总额", "内容": account_meta.get("total_assets", "未填写")},
        {"项目": "券商截图总市值", "内容": account_meta.get("broker_market_value", "未填写")},
        {"项目": "券商截图今日盈亏", "内容": account_meta.get("today_pl", "未填写")},
        {"项目": "券商截图持仓收益", "内容": account_meta.get("position_pl", "未填写")},
        {"项目": "券商截图可用资金", "内容": account_meta.get("available_cash", "未填写")},
        {"项目": "券商截图可取资金", "内容": account_meta.get("withdrawable_cash", "未填写")},
        {"项目": "券商截图总仓位", "内容": account_meta.get("position_ratio", "未填写")},
        {"项目": "全资产权益穿透仓位", "内容": full_asset_equity if full_asset_equity is not None else "未填写"},
        {"项目": "证券账户可用现金比例", "内容": (to_float(account_meta.get("available_cash")) / account_total) if account_total and to_float(account_meta.get("available_cash")) is not None else "未填写"},
        {"项目": "全资产流动性安全垫代理", "内容": liquidity_buffer},
        {"项目": "其中：证券可用现金", "内容": available_cash},
        {"项目": "其中：国债ETF市值", "内容": government_bond_value},
        {"项目": "现金安全垫最低值", "内容": cash_min},
        {"项目": "现金安全垫目标值", "内容": cash_target},
        {"项目": "现金安全垫状态", "内容": cash_status},
        {"项目": "未取到行情数量", "内容": int(missing_count)},
        {"项目": "仓位/纪律提醒数量", "内容": len(alerts)},
    ]

    if not alerts.empty:
        for _, alert in alerts.iterrows():
            rows.append({"项目": f"提醒-{alert['Code']}", "内容": alert["提醒"]})

    return pd.DataFrame(rows)


def build_portfolio_overview(dashboard: pd.DataFrame) -> pd.DataFrame:
    """组合总览只保留账户、现金和配置指标；逐只提醒移至持仓风险页。"""
    if dashboard.empty or "项目" not in dashboard.columns:
        return dashboard.copy()
    mask = ~dashboard["项目"].astype(str).str.startswith("提醒-")
    return dashboard.loc[mask].reset_index(drop=True)


def build_decision_center(
    double_anchor: pd.DataFrame,
    emotion: pd.DataFrame,
    quality_score: pd.DataFrame,
    buy_filter: pd.DataFrame,
    positions_action: pd.DataFrame,
    first_year_summary: dict[str, float],
    decision_inputs: dict[str, dict[str, str]] | None = None,
) -> pd.DataFrame:
    inputs = decision_inputs or {}
    anchor_row = double_anchor[double_anchor["锚点"] == "综合灯号"]
    anchor_light = anchor_row.iloc[0]["灯号"] if not anchor_row.empty else "待确认"
    anchor_note = anchor_row.iloc[0]["说明"] if not anchor_row.empty else "缺少双锚数据"
    temp_row = emotion[emotion["指标"] == "情绪温度"]
    temperature = int(to_float(temp_row.iloc[0]["数值"]) or 1) if not temp_row.empty else None
    emotion_note = temp_row.iloc[0]["解释"] if not temp_row.empty else "缺少情绪数据"
    emotion_reliability = str(temp_row.iloc[0].get("可靠性", "待确认")) if not temp_row.empty else "待确认"
    actionable = buy_filter[(buy_filter["买点灯号"].isin(["绿", "黄"])) & (buy_filter["一票否决"] == "否")]
    dividend_stock_count = int((buy_filter.get("资产角色", pd.Series(dtype=object)).astype(str) == "红利现金流个股").sum()) if not buy_filter.empty else 0
    risk_actions = positions_action[positions_action["触发提醒"] == "是"] if not positions_action.empty else positions_action
    quality_missing = int((quality_score["质量状态"] == "数据不足，需人工评分").sum()) if not quality_score.empty else 0
    quality_ready = int(quality_score["质量状态"].isin(["核心候选", "观察候选"]).sum()) if not quality_score.empty else 0
    market_permission = "开放买点复核" if anchor_light == "双绿" and (temperature or 5) <= 3 else "暂停标准新增"
    mainline_stage = decision_input_text(inputs, "mainline_stage") or "主线聚焦期后段/加速初期（框架默认）"
    trend_buy_type = decision_input_text(inputs, "trend_buy_type") or "待确认"
    quant_module_status = decision_input_text(inputs, "quant_module_status") or "V2.8.5-Q启用-需人工复核缺失项"
    retreat_status = decision_input_text(inputs, "retreat_factor_status") or "待人工复核"
    stage_ready = 0
    stage_watch = 0
    if not buy_filter.empty and "阶段推进结论" in buy_filter.columns:
        stage_ready = int(buy_filter["阶段推进结论"].isin(["允许下一档", "补足当前阶段", "小额防踏空"]).sum())
        stage_watch = int((buy_filter["阶段推进结论"] == "等待").sum())
    rows = [
        {"层级": "市场权限", "状态": market_permission, "证据": f"双锚={anchor_light}；情绪温度={temperature}", "动作": anchor_note},
        {"层级": "主线周期", "状态": mainline_stage, "证据": MAINLINE_CYCLE_STAGES, "动作": "核心持有、回踩加仓、强势不追、过热冷却、退潮复审"},
        {
            "层级": "量化验证层",
            "状态": quant_module_status,
            "证据": "指数环境、ETF买点、情绪温度、正期望、资产角色分层",
            "动作": "主观定方向，量化验信号；所有买点先打分，再按仓位纪律执行",
        },
        {
            "层级": "红利现金流层",
            "状态": f"{dividend_stock_count}只红利个股观察",
            "证据": "分红连续性、现金流覆盖、股息率安全垫、估值中低位和缩量回踩",
            "动作": "作为低波现金流层慢建；高息陷阱、举债分红、盈利下滑或只因快分红均否决",
        },
        {"层级": "趋势买点", "状态": trend_buy_type, "证据": TREND_BUY_PRIORITY, "动作": "只有主线/质量/情绪/买点/无否决共振时才复核"},
        {"层级": "阶段推进", "状态": f"{stage_ready}项可复核 / {stage_watch}项等待", "证据": "观察仓→验证仓→核心仓→目标仓；年度缺口不是买入信号", "动作": "达到当前阶段后，必须通过阶段验证才打开下一档"},
        {"层级": "退潮三因子", "状态": retreat_status, "证据": RETREAT_THREE_FACTORS, "动作": "三因子共振则降低进攻仓位或降级退出"},
        {
            "层级": "认知防错层",
            "状态": "V2.8.5-QM",
            "证据": "机会成本、全局最优、确认偏误、沉没成本、损失规避、风险概率、能力圈、复利",
            "动作": "若机会成本/反证/最大损失/能力圈中两项答不清，不买",
        },
        {"层级": "标准首批", "状态": f"{int((actionable['买点灯号'] == '绿').sum())}只", "证据": "买点通过≥5项、无否决、质量与仓位门槛通过", "动作": "仅在标准首批候选中复核"},
        {"层级": "半额首批", "状态": f"{int((actionable['买点灯号'] == '黄').sum())}只", "证据": "买点通过4项、无否决", "动作": "最多半额首批，且不得跨越剩余额度"},
        {"层级": "质量准入", "状态": f"{quality_ready}只通过 / {quality_missing}只待补", "证据": "部分代理分不折算为完整10分", "动作": "缺少完整评分与证据时禁止新增"},
        {
            "层级": "第一年配置",
            "状态": f"完成{first_year_summary.get('年度完成率', 0):.1%}",
            "证据": f"全资产目标{first_year_summary.get('年度目标占比', 0):.0%}；当前{first_year_summary.get('年度当前全资产占比', 0):.1%}；缺口{first_year_summary.get('年度资金缺口', 0):,.0f}元",
            "动作": "买入/减仓以全资产第一年目标为主；证券账户仓位仅作集中度参考",
        },
        {"层级": "情绪纪律", "状态": f"{temperature}级·{emotion_reliability}" if temperature else "待确认", "证据": emotion_note, "动作": "温度≥4暂停新增；低可靠性不得视为全市场结论"},
        {"层级": "持仓风控", "状态": f"{len(risk_actions)}项触发", "证据": "仓位、专项纪律与退出条件", "动作": "风控 > 清旧 > 补新 > 再平衡"},
        {"层级": "执行上限", "状态": "单日最多3类动作", "证据": "一个买入方向 + 两个卖出/减仓方向", "动作": "系统性风险日只执行风控"},
        {"层级": "数据边界", "状态": "条件式", "证据": "ETF份额、IOPV、龙头同步、次日验证、个股财务/估值未全量接入", "动作": "缺失项不得被代理数据自动判绿"},
    ]
    return pd.DataFrame(rows)


def build_execution_plan(
    buy_filter: pd.DataFrame,
    positions_action: pd.DataFrame,
    positions_sheet: pd.DataFrame,
    first_year: pd.DataFrame,
) -> pd.DataFrame:
    """把“一个买入方向 + 两个减仓方向”落实成清晰的当日动作预算。"""
    rows: list[dict[str, object]] = []
    annual_lookup = first_year_lookup(first_year)
    stage_executable = {"允许下一档", "补足当前阶段", "小额防踏空"}
    if "阶段推进结论" in buy_filter.columns:
        candidates = buy_filter[
            buy_filter["阶段推进结论"].isin(stage_executable)
            & (buy_filter["一票否决"] == "否")
        ].copy()
    else:
        candidates = buy_filter[
            buy_filter["买点灯号"].isin(["绿", "黄"])
            & (buy_filter["一票否决"] == "否")
        ].copy()
    used_themes: set[str] = set()
    buy_slots = 1
    for _, item in candidates.iterrows():
        theme = classify_theme(str(item.get("Name", "")), str(item.get("Role", "")))
        annual = annual_lookup.get(format_code(item.get("Code")), {})
        allowed = buy_slots > 0 and theme not in used_themes
        rows.append(
            {
                "优先级": 3,
                "动作类型": "买入复核",
                "标的": f"{item.get('Name')}({item.get('Code')})",
                "主题": theme,
                "是否占用额度": "是" if allowed else "否",
                "动作预算": f"{item.get('阶段推进结论') or '阶段待确认'}；最多1个买入方向；{item.get('单次仓位上限') or '遵守单次上限'}",
                "依据": item.get("建议", ""),
                "趋势结构": item.get("趋势结构"),
                "单次仓位上限": item.get("单次仓位上限"),
                "资产角色": item.get("资产角色"),
                "量化验证状态": item.get("量化验证状态"),
                "量化买点规则": item.get("量化买点规则"),
                "正期望检查": item.get("正期望检查"),
                "量化风控规则": item.get("量化风控规则"),
                "红利配置角色": item.get("红利配置角色"),
                "分红质量检查": item.get("分红质量检查"),
                "红利买点规则": item.get("红利买点规则"),
                "高息陷阱否决": item.get("高息陷阱否决"),
                "机会成本检查": item.get("机会成本检查"),
                "认知防错": item.get("认知防错"),
                "状态": "可进入复核" if allowed else "顺延",
                "阶段状态": item.get("阶段状态"),
                "阶段推进结论": item.get("阶段推进结论"),
                "阶段验证": item.get("阶段验证"),
                "阶段占用率": item.get("阶段占用率"),
                "下一档单次上限": item.get("下一档单次上限"),
                "阶段阻断原因": item.get("阶段阻断原因"),
                "阶段动作说明": item.get("阶段动作说明"),
                "年度配置项": annual.get("配置项"),
                "年度全资产目标": annual.get("年度目标占比"),
                "年度资金缺口": annual.get("年度资金缺口"),
                "年度完成率": annual.get("年度完成率"),
            }
        )
        if allowed:
            buy_slots -= 1
            used_themes.add(theme)

    sell_candidates: list[dict[str, object]] = []
    if not positions_action.empty:
        for _, item in positions_action[positions_action["触发提醒"] == "是"].iterrows():
            sell_candidates.append({"Code": item.get("Code"), "Name": item.get("Name"), "说明": item.get("说明", item.get("动作建议", "")), "超额": 999.0})
    if not positions_sheet.empty:
        for _, item in positions_sheet.iterrows():
            weight = to_float(first_valid(item.get("Full Asset Weight"), item.get("全资产当前仓位")))
            target = to_float(item.get("年度全资产目标"))
            if weight is not None and target is not None and weight > target:
                sell_candidates.append(
                    {
                        "Code": item.get("Code"),
                        "Name": item.get("Name"),
                        "说明": f"全资产仓位{weight:.2%}，超过第一年目标{target:.2%}",
                        "超额": weight - target,
                    }
                )
    seen_codes: set[str] = set()
    sell_rows: list[dict[str, object]] = []
    for item in sorted(sell_candidates, key=lambda row: float(row.get("超额", 0)), reverse=True):
        code = format_code(item.get("Code"))
        if code in seen_codes:
            continue
        seen_codes.add(code)
        sell_rows.append(item)
        if len(sell_rows) >= 2:
            break
    for priority, item in enumerate(sell_rows, start=1):
        annual = annual_lookup.get(format_code(item.get("Code")), {})
        rows.append(
            {
                "优先级": priority,
                "动作类型": "风控/减仓复核",
                "标的": f"{item.get('Name')}({item.get('Code')})",
                "主题": classify_theme(str(item.get("Name", "")), ""),
                "是否占用额度": "是",
                "动作预算": "最多2个卖出/减仓方向",
                "依据": item.get("说明", ""),
                "状态": "优先处理",
                "年度配置项": annual.get("配置项"),
                "年度全资产目标": annual.get("年度目标占比"),
                "年度资金缺口": annual.get("年度资金缺口"),
                "年度完成率": annual.get("年度完成率"),
            }
        )

    if not rows:
        rows.append({"优先级": 1, "动作类型": "观察", "标的": "无可执行候选", "主题": "—", "是否占用额度": "否", "动作预算": "0", "依据": "质量/情绪/买点/仓位未形成完整证据链", "状态": "仅观察", "年度配置项": None, "年度全资产目标": None, "年度资金缺口": None, "年度完成率": None})
    return pd.DataFrame(rows).sort_values(["优先级", "动作类型"], ignore_index=True)


def build_buy_candidates_view(
    buy_filter: pd.DataFrame,
    market_permission: str,
) -> pd.DataFrame:
    """保留全部关注标的，并按绿、黄、红、灰排序显示阻断原因。"""
    if buy_filter.empty:
        return buy_filter.copy()
    result = buy_filter.copy()
    result["市场权限"] = market_permission

    def blocker(row: pd.Series) -> str:
        reason = clean_blocker_reason(row.get("否决原因", ""))
        signal = str(row.get("买点灯号", "灰"))
        if reason:
            return reason
        if market_permission != "开放买点复核" and signal in {"绿", "黄"}:
            return f"市场权限：{market_permission}"
        if signal == "红":
            return str(row.get("建议", "未通过买点门槛") or "未通过买点门槛")
        if signal == "灰":
            return "证据不足或仅观察"
        return "无硬性阻断，仍需人工复核"

    result["阻断原因"] = result.apply(blocker, axis=1)
    result["数据状态"] = "有效"
    range_issues = pd.Series("", index=result.index, dtype=object)
    range_context_columns = {"High", "Low", "PrevClose", "PrevDayLow", "QuoteTime", "DataSource"}
    has_range_context = bool(range_context_columns.intersection(result.columns))
    if has_range_context:
        range_issues = result.apply(
            lambda row: buy_range_data_issue(
                latest=row.get("Latest"),
                high=row.get("High"),
                low=row.get("Low"),
                prev_close=row.get("PrevClose"),
                prev_day_low=row.get("PrevDayLow"),
                quote_time=row.get("QuoteTime"),
                data_source=row.get("DataSource", ""),
                prev_day_low_source=row.get("PrevDayLowSource", ""),
            ),
            axis=1,
        )
        result.loc[range_issues.astype(str).str.strip().ne(""), "数据状态"] = "行情未刷新"
    if "接口状态" in result.columns:
        interface_status = result["接口状态"].astype(str)
        result.loc[interface_status.str.contains("未刷新|缺失|失败", regex=True, na=False), "数据状态"] = "行情未刷新"
    if "Latest" in result.columns:
        latest = pd.to_numeric(result["Latest"], errors="coerce")
        result.loc[latest.isna(), "数据状态"] = "行情未刷新"
    if "PctChg" in result.columns:
        pct_chg = pd.to_numeric(result["PctChg"], errors="coerce")
        result.loc[pct_chg.isna(), "数据状态"] = "行情未刷新"
    if "建议买入区间" in result.columns:
        range_text = result["建议买入区间"].astype(str)
        numeric_range = range_text.str.contains(
            r"\d+(?:\.\d+)?–\d+(?:\.\d+)?",
            regex=True,
        )
        result.loc[range_text.str.contains("行情未刷新|数据不完整", regex=True, na=False), "数据状态"] = "行情未刷新"
        unsafe_data_range = numeric_range & result["数据状态"].eq("行情未刷新")
        if unsafe_data_range.any():
            downgraded_ranges: list[str] = []
            for issue in range_issues.loc[unsafe_data_range].astype(str).tolist():
                issue_text = f"：{issue}" if issue else ""
                downgraded_ranges.append(f"暂不建议买入（行情未刷新或数据不完整{issue_text}）")
            result.loc[unsafe_data_range, "建议买入区间"] = downgraded_ranges
            range_text = result["建议买入区间"].astype(str)
            numeric_range = range_text.str.contains(
                r"\d+(?:\.\d+)?–\d+(?:\.\d+)?",
                regex=True,
            )
    if market_permission != "开放买点复核" and "建议买入区间" in result.columns:
        range_text = result["建议买入区间"].astype(str)
        numeric_range = range_text.str.contains(
            r"\d+(?:\.\d+)?–\d+(?:\.\d+)?",
            regex=True,
        )
        permission_block = numeric_range & result["买点灯号"].isin(["绿", "黄"])
        result.loc[permission_block, "建议买入区间"] = f"暂不建议买入（市场权限：{market_permission}）"

    def merge_data_blocker(row: pd.Series) -> str:
        data_reason = quote_data_blocker_reason(row, range_issues.get(row.name, ""))
        current_reason = clean_blocker_reason(row.get("阻断原因", ""))
        if not data_reason:
            return current_reason or str(row.get("阻断原因", "") or "").strip()
        if not current_reason:
            return data_reason
        if data_reason in current_reason:
            return current_reason
        if current_reason in data_reason:
            return data_reason
        return f"{data_reason}；{current_reason}"

    result["阻断原因"] = result.apply(merge_data_blocker, axis=1)
    if market_permission != "开放买点复核" and "阶段推进结论" in result.columns:
        executable_stage = result["阶段推进结论"].isin(["允许下一档", "补足当前阶段", "小额防踏空"])
        result.loc[executable_stage, "阶段推进结论"] = "等待"
        result.loc[executable_stage, "下一档单次上限"] = 0.0
        result.loc[executable_stage, "阶段阻断原因"] = f"市场权限：{market_permission}"
        result.loc[executable_stage, "阶段动作说明"] = "等待市场权限恢复后再复核。"

    signal_rank = {"绿": 0, "黄": 1, "红": 2, "灰": 3}
    result["_signal_rank"] = result["买点灯号"].map(signal_rank).fillna(4)
    result["_quality_rank"] = pd.to_numeric(result.get("质量评分"), errors="coerce").fillna(-1)
    result["_pass_rank"] = pd.to_numeric(result.get("通过项"), errors="coerce").fillna(-1)
    result["_room_rank"] = pd.to_numeric(result.get("剩余额度"), errors="coerce").fillna(-1)
    result = result.sort_values(
        ["_signal_rank", "_quality_rank", "_pass_rank", "_room_rank"],
        ascending=[True, False, False, False],
        kind="stable",
    ).drop(columns=["_signal_rank", "_quality_rank", "_pass_rank", "_room_rank"])

    priority_columns = [
        "买点灯号", "Code", "Name", "建议买入区间", "趋势结构", "单次仓位上限", "建议", "阻断原因",
        "阶段状态", "阶段推进结论", "阶段验证", "阶段占用率", "下一档单次上限", "阶段阻断原因", "阶段动作说明",
        "资产角色", "量化验证分", "量化验证状态", "量化买点规则", "正期望检查", "量化风控规则",
        "红利配置角色", "分红质量检查", "红利买点规则", "高息陷阱否决",
        "机会成本检查", "认知防错", "Latest", "PctChg", "QuoteTime", "DataSource", "接口状态", "失败原因", "重试结果",
        "质量评分", "质量状态", "评分输入状态", "评分缺失项", "通过项", "市场权限",
        "仓位决策口径", "全资产当前仓位", "全资产目标仓位", "全资产剩余额度",
        "证券账户仓位", "年度配置项", "年度资金缺口", "年度完成率", "数据状态",
    ]
    for column in [
        "资产角色", "量化验证分", "量化验证状态", "量化买点规则", "正期望检查", "量化风控规则",
        "红利配置角色", "分红质量检查", "红利买点规则", "高息陷阱否决",
    ]:
        if column not in result.columns:
            result[column] = ""
    ordered = [column for column in priority_columns if column in result.columns]
    return result[ordered].reset_index(drop=True)


def build_position_risk_view(
    positions_action: pd.DataFrame,
    positions_sheet: pd.DataFrame,
) -> pd.DataFrame:
    """把持仓、盈亏、超配和专项纪律合并成按风险优先的工作表。"""
    if positions_sheet.empty:
        return positions_action.copy()
    base = positions_sheet.copy()
    action_columns = [
        column for column in ["Code", "触发提醒", "动作建议", "说明", "纪律分类", "是否复审"]
        if column in positions_action.columns
    ]
    actions = positions_action[action_columns].copy() if action_columns else pd.DataFrame(columns=["Code"])
    result = base.merge(actions, on="Code", how="left")
    result["全资产当前仓位"] = [
        to_float(first_valid(row.get("Full Asset Weight"), row.get("当前仓位")))
        for _, row in result.iterrows()
    ]
    result["全资产目标仓位"] = [
        to_float(row.get("年度全资产目标"))
        for _, row in result.iterrows()
    ]
    result["全资产剩余额度"] = [
        max(target - current, 0.0) if current is not None and target is not None else None
        for current, target in zip(result["全资产当前仓位"], result["全资产目标仓位"])
    ]
    result["全资产仓位状态"] = [
        describe_weight_status(current, target, label="第一年目标")
        for current, target in zip(result["全资产当前仓位"], result["全资产目标仓位"])
    ]
    result["证券账户仓位"] = [to_float(row.get("Weight")) for _, row in result.iterrows()]
    result["证券账户目标仓位"] = [to_float(row.get("Target Weight")) for _, row in result.iterrows()]
    result["证券账户仓位状态"] = [
        describe_weight_status(current, target, label="证券账户目标")
        for current, target in zip(result["证券账户仓位"], result["证券账户目标仓位"])
    ]

    risk_levels: list[str] = []
    reasons: list[str] = []
    for _, row in result.iterrows():
        weight = to_float(row.get("全资产当前仓位"))
        target = to_float(row.get("全资产目标仓位"))
        triggered = str(row.get("触发提醒", "否")) == "是"
        overweight = weight is not None and target is not None and weight > target
        loss = to_float(first_valid(row.get("Unrealized PnL"), row.get("P/L"))) or 0
        reason_parts: list[str] = []
        note = str(row.get("说明", "") or "").strip()
        if triggered and note:
            reason_parts.append(note)
        if overweight:
            reason_parts.append(f"全资产仓位{weight:.2%}高于第一年目标{target:.2%}")
        if loss < 0 and not reason_parts:
            reason_parts.append(f"当前浮亏{loss:,.0f}元")
        if triggered or overweight:
            risk_levels.append("红")
        elif loss < 0:
            risk_levels.append("黄")
        else:
            risk_levels.append("灰")
        reasons.append("；".join(reason_parts) if reason_parts else "未触发减仓或超配条件")

    result["风险级别"] = risk_levels
    result["触发原因"] = reasons
    if "动作建议" not in result.columns:
        result["动作建议"] = result.get(f"{FRAMEWORK_VERSION} Action", "观察")
    result["动作建议"] = result["动作建议"].fillna(result.get(f"{FRAMEWORK_VERSION} Action", "观察"))
    risk_rank = {"红": 0, "黄": 1, "灰": 2}
    result["_risk_rank"] = result["风险级别"].map(risk_rank).fillna(3)
    result["_weight_rank"] = pd.to_numeric(result.get("全资产当前仓位"), errors="coerce").fillna(0)
    result = result.sort_values(["_risk_rank", "_weight_rank"], ascending=[True, False], kind="stable")
    result = result.drop(columns=["_risk_rank", "_weight_rank"])
    priority_columns = [
        "风险级别", "Code", "Name", "动作建议", "触发原因",
        "全资产当前仓位", "全资产目标仓位", "全资产剩余额度", "全资产仓位状态",
        "证券账户仓位", "证券账户集中度提示",
        "P/L", "P/L%", "Market Value", "Latest", "纪律分类", "触发提醒",
        "年度配置项", "年度资金缺口", "年度完成率",
    ]
    ordered = [column for column in priority_columns if column in result.columns]
    return result[ordered].reset_index(drop=True)


def build_action_plan_view(
    execution_plan: pd.DataFrame,
    buy_filter: pd.DataFrame,
    positions_sheet: pd.DataFrame,
) -> pd.DataFrame:
    """给动作计划补充全资产年度主口径与证券账户参考口径。"""
    if execution_plan.empty:
        return execution_plan.copy()
    result = execution_plan.copy()
    if "Code" not in result.columns:
        result["Code"] = result.get("标的", "").astype(str).str.extract(r"\(([^()]+)\)\s*$")[0]
    result["Code"] = result["Code"].map(format_code)

    position_lookup: dict[str, dict[str, object]] = {}
    for _, row in positions_sheet.iterrows():
        code = format_code(row.get("Code"))
        weight = to_float(first_valid(row.get("Full Asset Weight"), row.get("当前仓位")))
        target = to_float(first_valid(row.get("年度全资产目标"), row.get("目标仓位")))
        broker_weight = to_float(row.get("Weight"))
        broker_target = to_float(row.get("Target Weight"))
        position_lookup[code] = {
            "仓位决策口径": "全资产/第一年配置" if target is not None else "缺少第一年配置映射",
            "当前仓位": weight,
            "目标仓位": target,
            "剩余额度": max(target - weight, 0) if weight is not None and target is not None else None,
            "全资产当前仓位": weight,
            "全资产目标仓位": target,
            "全资产剩余额度": max(target - weight, 0) if weight is not None and target is not None else None,
            "证券账户仓位": broker_weight,
            "证券账户目标仓位": broker_target,
        }
    for _, row in buy_filter.iterrows():
        code = format_code(row.get("Code"))
        position_lookup[code] = {
            "仓位决策口径": row.get("仓位决策口径"),
            "当前仓位": to_float(row.get("当前仓位")),
            "目标仓位": to_float(row.get("目标仓位")),
            "剩余额度": to_float(row.get("剩余额度")),
            "全资产当前仓位": to_float(row.get("全资产当前仓位")),
            "全资产目标仓位": to_float(row.get("全资产目标仓位")),
            "全资产剩余额度": to_float(row.get("全资产剩余额度")),
            "证券账户仓位": to_float(row.get("证券账户仓位")),
            "证券账户目标仓位": to_float(row.get("证券账户目标仓位")),
            "趋势结构": row.get("趋势结构"),
            "单次仓位上限": row.get("单次仓位上限"),
            "阶段状态": row.get("阶段状态"),
            "阶段推进结论": row.get("阶段推进结论"),
            "阶段验证": row.get("阶段验证"),
            "阶段占用率": row.get("阶段占用率"),
            "下一档单次上限": row.get("下一档单次上限"),
            "阶段阻断原因": row.get("阶段阻断原因"),
            "阶段动作说明": row.get("阶段动作说明"),
            "资产角色": row.get("资产角色"),
            "量化验证状态": row.get("量化验证状态"),
            "量化买点规则": row.get("量化买点规则"),
            "正期望检查": row.get("正期望检查"),
            "量化风控规则": row.get("量化风控规则"),
            "红利配置角色": row.get("红利配置角色"),
            "分红质量检查": row.get("分红质量检查"),
            "红利买点规则": row.get("红利买点规则"),
            "高息陷阱否决": row.get("高息陷阱否决"),
            "机会成本检查": row.get("机会成本检查"),
            "认知防错": row.get("认知防错"),
        }

    for column in [
        "仓位决策口径", "当前仓位", "目标仓位", "剩余额度",
        "全资产当前仓位", "全资产目标仓位", "全资产剩余额度",
        "证券账户仓位", "证券账户目标仓位", "趋势结构", "单次仓位上限",
        "阶段状态", "阶段推进结论", "阶段验证", "阶段占用率", "下一档单次上限", "阶段阻断原因", "阶段动作说明",
        "资产角色", "量化验证状态", "量化买点规则", "正期望检查", "量化风控规则",
        "红利配置角色", "分红质量检查", "红利买点规则", "高息陷阱否决",
        "机会成本检查", "认知防错",
    ]:
        result[column] = result["Code"].map(lambda code: position_lookup.get(code, {}).get(column))
    priority_columns = [
        "优先级", "动作类型", "Code", "标的", "状态", "动作预算", "依据",
        "趋势结构", "单次仓位上限", "资产角色", "量化验证状态", "量化买点规则",
        "阶段状态", "阶段推进结论", "阶段验证", "阶段占用率", "下一档单次上限", "阶段阻断原因", "阶段动作说明",
        "正期望检查", "量化风控规则", "红利配置角色", "分红质量检查",
        "红利买点规则", "高息陷阱否决", "机会成本检查", "认知防错",
        "仓位决策口径", "全资产当前仓位", "全资产目标仓位", "全资产剩余额度",
        "证券账户仓位", "证券账户目标仓位",
        "年度配置项", "年度资金缺口", "年度完成率",
    ]
    ordered = [column for column in priority_columns if column in result.columns]
    return result[ordered].sort_values(["优先级", "动作类型"], kind="stable").reset_index(drop=True)


def build_action_dashboard_view(
    decision_center: pd.DataFrame,
    action_plan: pd.DataFrame,
    buy_candidates: pd.DataFrame,
    position_risk: pd.DataFrame,
    dashboard: pd.DataFrame,
) -> pd.DataFrame:
    """生成卡片式首页需要的四类结构化记录。"""
    rows: list[dict[str, object]] = []

    def decision_status(level: str, fallback: str) -> str:
        if decision_center.empty or "层级" not in decision_center.columns:
            return fallback
        hit = decision_center[decision_center["层级"] == level]
        return str(hit.iloc[0].get("状态", fallback)) if not hit.empty else fallback

    def dashboard_value(item: str, fallback: str) -> str:
        if dashboard.empty or not {"项目", "内容"}.issubset(dashboard.columns):
            return fallback
        hit = dashboard[dashboard["项目"] == item]
        return str(hit.iloc[0].get("内容", fallback)) if not hit.empty else fallback

    market_permission = decision_status("市场权限", "待确认")
    stage_permission = decision_status("阶段推进", "待确认")
    candidate_count = 0
    if not buy_candidates.empty and "买点灯号" in buy_candidates.columns:
        if "阶段推进结论" in buy_candidates.columns:
            candidate_count = int(buy_candidates["阶段推进结论"].isin(["允许下一档", "补足当前阶段", "小额防踏空"]).sum())
        else:
            candidate_count = int(buy_candidates["买点灯号"].isin(["绿", "黄"]).sum())
        if market_permission != "开放买点复核":
            candidate_count = 0
    risk_count = int((position_risk.get("风险级别") == "红").sum()) if "风险级别" in position_risk.columns else 0
    cash_status = dashboard_value("现金安全垫状态", "待确认")
    core_statuses = [
        ("市场权限", market_permission, "双锚、情绪和数据共同决定"),
        ("阶段推进", stage_permission if stage_permission != "待确认" else f"{candidate_count}只", "只统计阶段结论允许复核的标的"),
        ("减仓复核", f"{risk_count}项", "触发专项纪律或全资产第一年超配"),
        ("现金安全垫", cash_status, "全资产与证券账户双口径"),
    ]
    for item, status, evidence in core_statuses:
        rows.append({"区域": "核心状态", "项目": item, "状态": status, "证据": evidence, "动作": "查看对应明细"})

    for _, row in action_plan.head(3).iterrows():
        rows.append(
            {
                "区域": "今日动作",
                "项目": row.get("标的", "未命名动作"),
                "状态": row.get("动作类型", "观察"),
                "证据": row.get("依据", ""),
                "动作": row.get("状态", "待复核"),
            }
        )

    if not buy_candidates.empty and "阻断原因" in buy_candidates.columns:
        blockers = buy_candidates[
            buy_candidates["阻断原因"].astype(str).str.strip().ne("")
        ].head(3)
        for _, row in blockers.iterrows():
            rows.append(
                {
                    "区域": "主要阻断",
                    "项目": row.get("Name", row.get("Code", "候选")),
                    "状态": row.get("买点灯号", "灰"),
                    "证据": row.get("阻断原因", ""),
                    "动作": row.get("建议", "继续观察"),
                }
            )

    quote_time = dashboard_value("行情数据时间", "未取到当日接口时间")
    interface_summary = dashboard_value("行情接口状态摘要", "未生成接口诊断")
    if any(key in quote_time for key in ("未取到", "未刷新", "缺失")):
        data_status = "行情未刷新"
    elif any(key in interface_summary for key in ("未刷新", "缺失", "失败")):
        data_status = "行情字段不完整"
    else:
        data_status = "行情已刷新"
    rows.append(
        {
            "区域": "数据状态",
            "项目": "行情时间",
            "状态": data_status,
            "证据": f"{quote_time}；{interface_summary}",
            "动作": "未刷新或字段不完整时禁止新增" if data_status != "行情已刷新" else "按纪律复核",
        }
    )
    return pd.DataFrame(rows, columns=["区域", "项目", "状态", "证据", "动作"])


def build_buy_point_plan_view(
    buy_point_plan: pd.DataFrame,
    positions_sheet: pd.DataFrame,
    buy_filter: pd.DataFrame | None = None,
) -> pd.DataFrame:
    """把人工买点计划与当前持仓上下文合并成前台可读表。"""
    columns = [
        "Code", "Name", "买点灯号", "资产角色", "当前价", "持仓市值", "成本", "当前盈亏",
        "轻仓观察区间", "标准买点区间", "强买点区间", "不买/暂停区",
        "下一笔合理金额", "成立条件", "风险暂停条件", "来源日期",
    ]
    if buy_point_plan is None or buy_point_plan.empty:
        return pd.DataFrame(columns=columns)

    position_lookup = {}
    if positions_sheet is not None and not positions_sheet.empty:
        position_lookup = positions_sheet.set_index("Code").to_dict("index")
    buy_filter_lookup = {}
    if buy_filter is not None and not buy_filter.empty:
        buy_filter_lookup = buy_filter.set_index("Code").to_dict("index")

    rows: list[dict[str, object]] = []
    for _, item in buy_point_plan.iterrows():
        code = format_code(item.get("Code"))
        position = position_lookup.get(code, {})
        filter_row = buy_filter_lookup.get(code, {})
        current_price = first_valid(
            to_float(position.get("Latest")),
            to_float(item.get("当前价参考")),
        )
        rows.append(
            {
                "Code": code,
                "Name": first_valid_text(position.get("Name"), item.get("Name")),
                "买点灯号": filter_row.get("买点灯号"),
                "资产角色": item.get("资产角色"),
                "当前价": current_price,
                "持仓市值": to_float(position.get("Market Value")),
                "成本": to_float(position.get("Cost")),
                "当前盈亏": to_float(position.get("P/L%")),
                "轻仓观察区间": item.get("轻仓观察区间"),
                "标准买点区间": item.get("标准买点区间"),
                "强买点区间": item.get("强买点区间"),
                "不买/暂停区": item.get("不买/暂停区"),
                "下一笔合理金额": item.get("下一笔合理金额"),
                "成立条件": item.get("成立条件"),
                "风险暂停条件": item.get("风险暂停条件"),
                "来源日期": item.get("来源日期"),
            }
        )
    result = pd.DataFrame(rows, columns=columns)
    result["_has_position"] = result["持仓市值"].notna().astype(int)
    result["_value_rank"] = pd.to_numeric(result["持仓市值"], errors="coerce").fillna(-1)
    result = result.sort_values(["_has_position", "_value_rank", "Code"], ascending=[False, False, True], kind="stable")
    return result.drop(columns=["_has_position", "_value_rank"]).reset_index(drop=True)


def is_long_term_tracking_row(row) -> bool:
    """识别观察池中的长期跟踪个股。"""
    text = " ".join(
        str(row.get(column, "") or "")
        for column in ["Role", "Action", "Notes", "Thesis"]
    )
    asset_type = str(row.get("Asset Type", "") or "").strip().lower()
    return asset_type == "stock" and ("长期跟踪" in text or "个股观察" in text)


def build_long_term_tracking_view(
    watchlist: pd.DataFrame,
    market_data: pd.DataFrame,
    quality_score: pd.DataFrame,
    buy_filter: pd.DataFrame,
    first_year: pd.DataFrame,
) -> pd.DataFrame:
    """专门汇总长期跟踪个股，避免在超宽买入候选页里查找。"""
    market_lookup = market_data.set_index("Code").to_dict("index") if not market_data.empty else {}
    quality_lookup = quality_score.set_index("Code").to_dict("index") if not quality_score.empty else {}
    buy_lookup = buy_filter.set_index("Code").to_dict("index") if not buy_filter.empty else {}
    year_lookup = first_year_lookup(first_year)
    rows: list[dict[str, object]] = []

    for _, item in watchlist.iterrows():
        if not is_long_term_tracking_row(item):
            continue
        code = format_code(item.get("Code"))
        market = market_lookup.get(code, {})
        quality = quality_lookup.get(code, {})
        buy = buy_lookup.get(code, {})
        year = year_lookup.get(code, {})
        quality_status = str(quality.get("质量状态", "待确认"))
        quality_input_status = str(quality.get("评分输入状态", "待确认"))
        quality_missing_items = str(quality.get("评分缺失项", "未生成质量评分诊断"))
        blocker = clean_blocker_reason(buy.get("阻断原因", "")) or clean_blocker_reason(buy.get("否决原因", ""))
        interface_status = str(market.get("接口状态", "行情未刷新"))
        data_blocker = quote_data_blocker_reason(market)
        if data_blocker:
            if blocker and data_blocker not in blocker:
                blocker = f"{data_blocker}；{blocker}"
            else:
                blocker = data_blocker
        next_steps: list[str] = []
        if "未刷新" in interface_status or "失败" in interface_status or "缺失" in interface_status:
            next_steps.append("联网重跑行情接口")
        if "数据不足" in quality_status:
            next_steps.append("补完整质量评分、估值/筹码、龙头同步和次日验证")
        elif blocker:
            next_steps.append("先处理阻断原因，再进入人工复核")
        if not next_steps:
            next_steps.append("仅限人工复核，不自动交易")
        rows.append(
            {
                "Code": code,
                "Name": first_valid(market.get("Name"), item.get("Name")),
                "跟踪分组": item.get("Role"),
                "Theme": item.get("Theme"),
                "Latest": market.get("Latest"),
                "PctChg": market.get("PctChg"),
                "Open": market.get("Open"),
                "High": market.get("High"),
                "Low": market.get("Low"),
                "PrevClose": market.get("PrevClose"),
                "Amount": market.get("Amount"),
                "QuoteTime": market.get("QuoteTime"),
                "DataSource": market.get("DataSource"),
                "接口状态": market.get("接口状态"),
                "失败原因": market.get("失败原因"),
                "重试结果": market.get("重试结果"),
                "质量状态": quality_status,
                "质量评分": quality.get("折算质量分"),
                "评分输入状态": quality_input_status,
                "评分缺失项": quality_missing_items,
                "买点灯号": buy.get("买点灯号"),
                "阻断原因": blocker or "无",
                "建议买入区间": buy.get("建议买入区间"),
                "趋势结构": buy.get("趋势结构"),
                "单次仓位上限": buy.get("单次仓位上限"),
                "资产角色": buy.get("资产角色"),
                "量化验证状态": buy.get("量化验证状态"),
                "量化买点规则": buy.get("量化买点规则"),
                "正期望检查": buy.get("正期望检查"),
                "量化风控规则": buy.get("量化风控规则"),
                "红利配置角色": buy.get("红利配置角色"),
                "分红质量检查": buy.get("分红质量检查"),
                "红利买点规则": buy.get("红利买点规则"),
                "高息陷阱否决": buy.get("高息陷阱否决"),
                "机会成本检查": buy.get("机会成本检查"),
                "反向失败路径": buy.get("反向失败路径"),
                "建议": buy.get("建议"),
                "仓位决策口径": buy.get("仓位决策口径"),
                "全资产当前仓位": buy.get("全资产当前仓位"),
                "全资产目标仓位": buy.get("全资产目标仓位"),
                "全资产剩余额度": buy.get("全资产剩余额度"),
                "证券账户仓位": buy.get("证券账户仓位"),
                "证券账户目标仓位": buy.get("证券账户目标仓位"),
                "年度配置项": year.get("配置项"),
                "年度资金缺口": year.get("年度资金缺口"),
                "年度完成率": year.get("年度完成率"),
                "Thesis": item.get("Thesis"),
                "Invalidation": item.get("Invalidation"),
                "下一步": "；".join(next_steps),
            }
        )

    return pd.DataFrame(
        rows,
        columns=[
            "Code", "Name", "跟踪分组", "Theme",
            "Latest", "PctChg", "Open", "High", "Low", "PrevClose", "Amount", "QuoteTime", "DataSource",
            "接口状态", "失败原因", "重试结果",
            "质量状态", "质量评分", "评分输入状态", "评分缺失项", "买点灯号",
            "阻断原因", "建议买入区间", "趋势结构", "单次仓位上限",
            "资产角色", "量化验证状态", "量化买点规则", "正期望检查", "量化风控规则",
            "红利配置角色", "分红质量检查", "红利买点规则", "高息陷阱否决",
            "机会成本检查", "反向失败路径", "建议",
            "仓位决策口径", "全资产当前仓位", "全资产目标仓位", "全资产剩余额度",
            "证券账户仓位",
            "年度配置项", "年度资金缺口", "年度完成率",
            "Thesis", "Invalidation", "下一步",
        ],
    )


def build_output_sheet_order(sheets: dict[str, pd.DataFrame]) -> list[str]:
    """返回固定的前台六页顺序，并把明细页稳定地排在后面。"""
    names = list(sheets)
    ordered = [name for name in FRONT_SHEET_ORDER if name in sheets]
    ordered += [name for name in DETAIL_SHEET_ORDER if name in sheets and name not in ordered]
    ordered += [name for name in names if name not in ordered]
    return ordered


def style_excel_worksheet(ws, sheet_name: str) -> None:
    if ws.max_row < 1:
        return

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin_fill = PatternFill("solid", fgColor="EAF3F8")
    good_fill = PatternFill("solid", fgColor="C6EFCE")
    watch_fill = PatternFill("solid", fgColor="FFEB9C")
    bad_fill = PatternFill("solid", fgColor="FFC7CE")
    neutral_fill = PatternFill("solid", fgColor="D9EAD3")
    grey_fill = PatternFill("solid", fgColor="E7E6E6")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = {cell.value: cell.column for cell in ws[1]}
    default_widths = {
        "Code": 12,
        "Name": 18,
        "建议买入区间": 38,
        "Role": 14,
        "质量评分": 11,
        "质量状态": 22,
        "情绪温度": 11,
        "反馈状态": 12,
        "趋势结构": 34,
        "关键支撑纪律": 42,
        "单次仓位上限": 32,
        "爆量/量质吸纪律": 42,
        "机会成本检查": 42,
        "认知防错": 44,
        "决策树情景": 42,
        "赔率/最大损失": 36,
        "反向失败路径": 42,
        "能力圈边界": 34,
        "资产角色": 18,
        "量化验证分": 12,
        "量化验证状态": 16,
        "量化买点规则": 48,
        "正期望检查": 46,
        "量化风控规则": 46,
        "红利配置角色": 18,
        "分红质量检查": 48,
        "红利买点规则": 48,
        "高息陷阱否决": 52,
        "当前仓位": 12,
        "目标仓位": 12,
        "剩余额度": 12,
        "仓位状态": 22,
        "阶段状态": 14,
        "阶段目标下限": 14,
        "阶段目标上限": 14,
        "阶段占用率": 14,
        "阶段验证": 11,
        "阶段推进结论": 16,
        "下一档单次上限": 16,
        "阶段阻断原因": 42,
        "阶段动作说明": 42,
        "仓位决策口径": 20,
        "全资产当前仓位": 16,
        "全资产目标仓位": 16,
        "全资产剩余额度": 16,
        "全资产仓位状态": 24,
        "证券账户目标仓位": 17,
        "证券账户仓位状态": 24,
        "证券账户集中度提示": 42,
        "Latest": 11,
        "PctChg": 10,
        "Open": 10,
        "High": 10,
        "Low": 10,
        "PrevClose": 11,
        "Amount": 15,
        "Avg20Amount": 15,
        "Avg20AmountSource": 18,
        "PrevDayLowSource": 18,
        "日内位置": 13,
        "量能倍数": 11,
        "分时结构": 16,
        "量价关系": 18,
        "份额变动": 16,
        "折溢价": 14,
        "龙头同步": 16,
        "次日验证": 18,
        "通过项": 9,
        "通过明细": 36,
        "未通过项": 36,
        "待确认项": 36,
        "一票否决": 11,
        "否决原因": 34,
        "买点灯号": 12,
        "建议": 44,
        "说明": 42,
        "动作建议": 24,
        "全资产穿透仓位": 16,
        "证券账户仓位": 16,
        "数据完整度": 12,
        "评分明细": 42,
        "数据边界": 42,
        "证据": 42,
        "动作": 38,
        "修复建议": 42,
        "阈值/定义": 34,
        "来源": 22,
        "依据": 42,
        "主题上限": 14,
        "上限剩余额度": 16,
        "上限状态": 14,
        "风控口径": 18,
        "配置项": 26,
        "年度配置项": 26,
        "资产层级": 18,
        "映射代码": 24,
        "年度配置状态": 18,
        "进度状态": 18,
        "执行约束": 46,
        "跟踪分组": 22,
        "Thesis": 44,
        "Invalidation": 44,
        "下一步": 36,
        "Section": 14,
        "Key": 28,
        "Value": 42,
        "Notes": 56,
        "当前生效值": 42,
        "维护文件": 24,
        "Module": 14,
        "DataType": 26,
        "PrimarySource": 28,
        "FallbackSource": 38,
        "FreshnessRule": 56,
        "FailureBehavior": 56,
        "ConfigFile": 26,
        "当前价": 11,
        "持仓市值": 14,
        "成本": 11,
        "当前盈亏": 12,
        "当前价参考": 12,
        "轻仓观察区间": 18,
        "标准买点区间": 18,
        "强买点区间": 18,
        "不买/暂停区": 28,
        "下一笔合理金额": 24,
        "成立条件": 58,
        "风险暂停条件": 52,
        "来源日期": 14,
    }

    for col_idx in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col_idx).value
        width = default_widths.get(header, 14)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
        for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2, max_row=ws.max_row):
            for item in cell:
                item.alignment = Alignment(vertical="center", wrap_text=width >= 18)

    if sheet_name in ("Buy_Filter", "03_买入候选"):
        for row in range(2, ws.max_row + 1):
            ws.row_dimensions[row].height = 38
        for col_name in [
            "建议买入区间", "趋势结构", "关键支撑纪律", "单次仓位上限", "爆量/量质吸纪律",
            "阶段阻断原因", "阶段动作说明",
            "机会成本检查", "认知防错", "决策树情景", "赔率/最大损失", "反向失败路径", "能力圈边界",
            "量化买点规则", "正期望检查", "量化风控规则",
            "分红质量检查", "红利买点规则", "高息陷阱否决",
            "通过明细", "未通过项", "待确认项", "否决原因", "建议",
        ]:
            col_idx = headers.get(col_name)
            if col_idx:
                for row in range(2, ws.max_row + 1):
                    ws.cell(row=row, column=col_idx).alignment = Alignment(vertical="center", wrap_text=True)

        lamp_col = headers.get("买点灯号")
        if lamp_col:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=lamp_col)
                text = str(cell.value or "")
                if "绿" in text:
                    cell.fill = good_fill
                elif "黄" in text:
                    cell.fill = watch_fill
                elif "红" in text:
                    cell.fill = bad_fill
                else:
                    cell.fill = grey_fill
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        veto_col = headers.get("一票否决")
        if veto_col:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=veto_col)
                if str(cell.value) == "是":
                    cell.fill = bad_fill
                    cell.font = Font(bold=True)
                elif str(cell.value) == "否":
                    cell.fill = neutral_fill

        quant_col = headers.get("量化验证状态")
        if quant_col:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=quant_col)
                text = str(cell.value or "")
                if "标准" in text:
                    cell.fill = good_fill
                elif "观察" in text:
                    cell.fill = watch_fill
                elif "阻断" in text or "风险" in text:
                    cell.fill = bad_fill
                else:
                    cell.fill = grey_fill
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        stage_col = headers.get("阶段推进结论")
        if stage_col:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=stage_col)
                text = str(cell.value or "")
                if "允许" in text or "补足" in text:
                    cell.fill = good_fill
                elif "防踏空" in text or "等待" in text:
                    cell.fill = watch_fill
                elif "暂停" in text or "降级" in text:
                    cell.fill = bad_fill
                else:
                    cell.fill = grey_fill
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    if sheet_name in ("Investment_Profile", "Data_Sources"):
        for row in range(2, ws.max_row + 1):
            ws.row_dimensions[row].height = 44

    if sheet_name in ("买点计划", "Buy_Point_Plan_Source"):
        for row in range(2, ws.max_row + 1):
            ws.row_dimensions[row].height = 56
        lamp_col = headers.get("买点灯号")
        if lamp_col:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=lamp_col)
                text = str(cell.value or "")
                if "绿" in text:
                    cell.fill = good_fill
                elif "黄" in text:
                    cell.fill = watch_fill
                elif "红" in text:
                    cell.fill = bad_fill
                else:
                    cell.fill = grey_fill
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col_name in ["成立条件", "风险暂停条件", "不买/暂停区", "下一笔合理金额"]:
            col_idx = headers.get(col_name)
            if col_idx:
                for row in range(2, ws.max_row + 1):
                    ws.cell(row=row, column=col_idx).alignment = Alignment(vertical="center", wrap_text=True)

    if sheet_name == "长期跟踪个股":
        for row in range(2, ws.max_row + 1):
            ws.row_dimensions[row].height = 42
        lamp_col = headers.get("买点灯号")
        if lamp_col:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=lamp_col)
                text = str(cell.value or "")
                if "绿" in text:
                    cell.fill = good_fill
                elif "黄" in text:
                    cell.fill = watch_fill
                elif "红" in text:
                    cell.fill = bad_fill
                else:
                    cell.fill = grey_fill

    if sheet_name == "04_阶段推进":
        ws.freeze_panes = "A2"
        for row in range(2, ws.max_row + 1):
            ws.row_dimensions[row].height = 42
        stage_col = headers.get("阶段推进结论")
        if stage_col:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=stage_col)
                text = str(cell.value or "")
                if "允许" in text or "补足" in text:
                    cell.fill = good_fill
                elif "防踏空" in text or "等待" in text:
                    cell.fill = watch_fill
                elif "暂停" in text or "降级" in text:
                    cell.fill = bad_fill
                else:
                    cell.fill = grey_fill
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col_name in ["阶段阻断原因", "阶段动作说明"]:
            col_idx = headers.get(col_name)
            if col_idx:
                for row in range(2, ws.max_row + 1):
                    ws.cell(row, col_idx).alignment = Alignment(vertical="center", wrap_text=True)

    if sheet_name in ("Buy_Filter", "Positions_Action", "03_买入候选", "04_阶段推进", "05_持仓风险"):
        for row in range(2, ws.max_row + 1):
            if row % 2 == 0:
                for col_idx in range(1, ws.max_column + 1):
                    if ws.cell(row=row, column=col_idx).fill.fill_type is None:
                        ws.cell(row=row, column=col_idx).fill = thin_fill

    if sheet_name == "Decision_Center":
        ws.freeze_panes = "A2"
        for row in range(2, ws.max_row + 1):
            ws.row_dimensions[row].height = 34
        for col_name in ("证据", "动作"):
            col_idx = headers.get(col_name)
            if col_idx:
                for row in range(2, ws.max_row + 1):
                    ws.cell(row, col_idx).alignment = Alignment(vertical="center", wrap_text=True)

    if sheet_name == "05_持仓风险" and headers.get("风险级别"):
        risk_col = headers["风险级别"]
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row, risk_col)
            cell.fill = bad_fill if str(cell.value) == "红" else watch_fill if str(cell.value) == "黄" else grey_fill
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[row].height = 34

    if sheet_name == "02_今日动作" and headers.get("动作类型"):
        action_col = headers["动作类型"]
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row, action_col)
            text = str(cell.value or "")
            cell.fill = bad_fill if "减仓" in text or "风控" in text else good_fill if "买入" in text else grey_fill
            cell.font = Font(bold=True)

    if sheet_name == "Quality_Score" and headers.get("折算质量分"):
        col_letter = get_column_letter(headers["折算质量分"])
        ws.conditional_formatting.add(
            f"{col_letter}2:{col_letter}{ws.max_row}",
            ColorScaleRule(start_type="num", start_value=0, start_color="F8696B", mid_type="num", mid_value=6, mid_color="FFEB84", end_type="num", end_value=10, end_color="63BE7B"),
        )

    if sheet_name == "Checks" and headers.get("状态"):
        status_col = headers["状态"]
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row, status_col)
            cell.fill = good_fill if str(cell.value) == "OK" else watch_fill if str(cell.value) == "待补" else bad_fill
            cell.font = Font(bold=True)

    if sheet_name == "Exposure" and headers.get("上限状态"):
        status_col = headers["上限状态"]
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row, status_col)
            if str(cell.value) == "超限":
                cell.fill = bad_fill
                cell.font = Font(bold=True)
            elif str(cell.value) == "范围内":
                cell.fill = good_fill

    for header, number_format in {
        "Latest": "0.000",
        "Open": "0.000",
        "High": "0.000",
        "Low": "0.000",
        "PrevClose": "0.000",
        "PctChg": "0.00",
        "当前仓位": "0.00%",
        "目标仓位": "0.00%",
        "剩余额度": "0.00%",
        "阶段目标下限": "0.00%",
        "阶段目标上限": "0.00%",
        "阶段占用率": "0.00%",
        "下一档单次上限": "0.00%",
        "全资产当前仓位": "0.00%",
        "全资产目标仓位": "0.00%",
        "全资产剩余额度": "0.00%",
        "量能倍数": "0.00",
        "质量评分": "0.0",
        "折算质量分": "0.0",
        "量化验证分": "0.0",
        "数据完整度": "0%",
        "证券账户仓位": "0.00%",
        "证券账户目标仓位": "0.00%",
        "全资产穿透仓位": "0.00%",
        "主题上限": "0.00%",
        "上限剩余额度": "0.00%",
        "Full Asset Weight": "0.00%",
        "Weight Gap": "0.00%",
        "年度目标占比": "0.00%",
        "年度全资产目标": "0.00%",
        "年度完成率": "0.00%",
        "当前全资产占比": "0.00%",
        "目标收益下限": "0.00%",
        "目标收益上限": "0.00%",
        "Amount": "#,##0",
        "Avg20Amount": "#,##0",
        "配置表目标金额": "#,##0;[Red](#,##0);-",
        "按当前全资产目标金额": "#,##0;[Red](#,##0);-",
        "最新持仓金额": "#,##0;[Red](#,##0);-",
        "配置表当前金额约": "#,##0;[Red](#,##0);-",
        "年度目标金额": "#,##0;[Red](#,##0);-",
        "年度组内当前金额": "#,##0;[Red](#,##0);-",
        "年度资金缺口": "#,##0;[Red](#,##0);-",
        "High20": "0.000",
        "Low20": "0.000",
        "Avg20Amount": "#,##0",
    }.items():
        col_idx = headers.get(header)
        if col_idx:
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=col_idx).number_format = number_format


def semantic_fill(value: object) -> PatternFill:
    """首页状态卡的语义色。"""
    text = str(value or "")
    if any(key in text for key in ("暂停", "禁止", "未刷新", "超配", "红")):
        return PatternFill("solid", fgColor="F4CCCC")
    if any(key in text for key in ("复核", "观察", "待确认", "黄")):
        return PatternFill("solid", fgColor="FFF2CC")
    if any(key in text for key in ("开放", "达标", "已刷新", "绿")):
        return PatternFill("solid", fgColor="D9EAD3")
    return PatternFill("solid", fgColor="E7E6E6")


def style_action_dashboard(ws, frame: pd.DataFrame) -> None:
    """把结构化首页记录渲染为一屏可读的卡片式决策工作台。"""
    ws.delete_rows(1, ws.max_row)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A8"
    for column, width in {"A": 11, "B": 14, "C": 18, "D": 15, "E": 16, "F": 16, "G": 16, "H": 18}.items():
        ws.column_dimensions[column].width = width

    ws.merge_cells("A1:H1")
    ws["A1"] = f"{FRAMEWORK_DISPLAY_VERSION} 今日行动决策中心"
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=18)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    core = frame[frame["区域"] == "核心状态"] if not frame.empty else pd.DataFrame()
    permission = "待确认"
    if not core.empty:
        hit = core[core["项目"] == "市场权限"]
        if not hit.empty:
            permission = str(hit.iloc[0].get("状态", "待确认"))
    action_count = int((frame["区域"] == "今日动作").sum()) if not frame.empty else 0
    conclusion = "只做风控，不新增" if "暂停" in permission or "禁止" in permission else "允许按纪律复核买点"
    ws.merge_cells("A2:H2")
    ws["A2"] = f"今日结论：{conclusion}｜市场权限：{permission}｜动作队列：{action_count}项"
    ws["A2"].fill = semantic_fill(permission)
    ws["A2"].font = Font(bold=True, size=13, color="9C0006" if "暂停" in permission else "006100")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 26

    card_ranges = [("A4:B4", "A5:B6"), ("C4:D4", "C5:D6"), ("E4:F4", "E5:F6"), ("G4:H4", "G5:H6")]
    core_records = core.to_dict("records")[:4]
    for index, (label_range, value_range) in enumerate(card_ranges):
        ws.merge_cells(label_range)
        ws.merge_cells(value_range)
        label_cell = ws[label_range.split(":")[0]]
        value_cell = ws[value_range.split(":")[0]]
        record = core_records[index] if index < len(core_records) else {"项目": "待补", "状态": "待确认"}
        label_cell.value = record.get("项目")
        value_cell.value = record.get("状态")
        label_cell.fill = PatternFill("solid", fgColor="D9EAF7")
        label_cell.font = Font(bold=True, color="1F1F1F")
        label_cell.alignment = Alignment(horizontal="center", vertical="center")
        value_cell.fill = semantic_fill(record.get("状态"))
        value_cell.font = Font(bold=True, size=15)
        value_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.merge_cells("A8:H8")
    ws["A8"] = "今日动作（风控优先，最多一个买入方向 + 两个减仓方向）"
    ws["A8"].fill = PatternFill("solid", fgColor="1F4E78")
    ws["A8"].font = Font(color="FFFFFF", bold=True)
    action_headers = ["顺序", "动作类型", "标的", "状态", "核心依据"]
    for column, value in enumerate(action_headers, start=1):
        cell = ws.cell(9, column, value)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("E9:H9")
    action_rows = frame[frame["区域"] == "今日动作"].head(3) if not frame.empty else pd.DataFrame()
    for offset, (_, record) in enumerate(action_rows.iterrows(), start=10):
        values = [offset - 9, record.get("状态"), record.get("项目"), record.get("动作"), record.get("证据")]
        for column, value in enumerate(values[:4], start=1):
            ws.cell(offset, column, value)
        ws.merge_cells(start_row=offset, start_column=5, end_row=offset, end_column=8)
        ws.cell(offset, 5, values[4])
        for column in range(1, 9):
            ws.cell(offset, column).alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[offset].height = 30

    blocker_start = 14
    ws.merge_cells(start_row=blocker_start, start_column=1, end_row=blocker_start, end_column=8)
    ws.cell(blocker_start, 1, "主要阻断原因")
    ws.cell(blocker_start, 1).fill = PatternFill("solid", fgColor="7F6000")
    ws.cell(blocker_start, 1).font = Font(color="FFFFFF", bold=True)
    blockers = frame[frame["区域"] == "主要阻断"].head(3) if not frame.empty else pd.DataFrame()
    for offset, (_, record) in enumerate(blockers.iterrows(), start=blocker_start + 1):
        ws.cell(offset, 1, record.get("项目"))
        ws.cell(offset, 2, record.get("状态"))
        ws.merge_cells(start_row=offset, start_column=3, end_row=offset, end_column=6)
        ws.cell(offset, 3, record.get("证据"))
        ws.merge_cells(start_row=offset, start_column=7, end_row=offset, end_column=8)
        ws.cell(offset, 7, record.get("动作"))
        for column in range(1, 9):
            ws.cell(offset, column).alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[offset].height = 28

    data_row = frame[frame["区域"] == "数据状态"].head(1) if not frame.empty else pd.DataFrame()
    ws.merge_cells("A19:H20")
    if not data_row.empty:
        record = data_row.iloc[0]
        ws["A19"] = f"数据状态：{record.get('状态')}｜{record.get('证据')}｜{record.get('动作')}"
        ws["A19"].fill = semantic_fill(record.get("状态"))
    else:
        ws["A19"] = "数据状态：待确认"
        ws["A19"].fill = semantic_fill("待确认")
    ws["A19"].font = Font(bold=True)
    ws["A19"].alignment = Alignment(vertical="center", wrap_text=True)


def add_dashboard_chart(wb: Workbook) -> None:
    """Add one decision-useful native Excel chart: full-asset current versus first-year target."""
    dashboard_name = "07_组合总览" if "07_组合总览" in wb.sheetnames else "06_组合总览" if "06_组合总览" in wb.sheetnames else "Dashboard"
    if dashboard_name not in wb.sheetnames or "Positions" not in wb.sheetnames:
        return
    dashboard = wb[dashboard_name]
    positions = wb["Positions"]
    headers = {cell.value: cell.column for cell in positions[1]}
    if {"Name", "Full Asset Weight", "年度全资产目标"}.issubset(headers):
        current_column = "Full Asset Weight"
        target_column = "年度全资产目标"
        chart_title = "全资产当前仓位 vs 第一年目标"
        x_axis_title = "全资产权重"
        series_labels = ("全资产当前", "第一年目标")
    elif {"Name", "Weight", "Target Weight"}.issubset(headers):
        current_column = "Weight"
        target_column = "Target Weight"
        chart_title = "证券账户仓位 vs 目标仓位"
        x_axis_title = "证券账户权重"
        series_labels = ("证券账户当前", "证券账户目标")
    else:
        return
    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = chart_title
    chart.y_axis.title = "标的"
    chart.x_axis.title = x_axis_title
    chart.height = 8.2
    chart.width = 16.5
    categories = Reference(positions, min_col=headers["Name"], min_row=2, max_row=positions.max_row)
    for column_name in (current_column, target_column):
        data = Reference(
            positions,
            min_col=headers[column_name],
            max_col=headers[column_name],
            min_row=1,
            max_row=positions.max_row,
        )
        chart.add_data(data, titles_from_data=True, from_rows=False)
    if len(chart.series) >= 2:
        chart.series[0].tx = SeriesLabel(v=series_labels[0])
        chart.series[1].tx = SeriesLabel(v=series_labels[1])
    chart.set_categories(categories)
    chart.legend.position = "b"
    chart.dLbls = DataLabelList()
    chart.dLbls.showVal = False
    dashboard.add_chart(chart, "D2")


def add_first_year_chart(wb: Workbook) -> None:
    """在Dashboard加入第一年动态目标金额与当前金额对比。"""
    dashboard_name = "07_组合总览" if "07_组合总览" in wb.sheetnames else "06_组合总览" if "06_组合总览" in wb.sheetnames else "Dashboard"
    allocation_name = "06_年度配置" if "06_年度配置" in wb.sheetnames else "05_年度配置" if "05_年度配置" in wb.sheetnames else "FirstYear_Allocation"
    if dashboard_name not in wb.sheetnames or allocation_name not in wb.sheetnames:
        return
    dashboard = wb[dashboard_name]
    allocation = wb[allocation_name]
    headers = {cell.value: cell.column for cell in allocation[1]}
    required = {"配置项", "按当前全资产目标金额", "最新持仓金额"}
    if not required.issubset(headers):
        return
    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = "第一年配置：当前金额 vs 动态目标"
    chart.y_axis.title = "配置项"
    chart.x_axis.title = "金额（元）"
    chart.height = 8.2
    chart.width = 16.5
    categories = Reference(allocation, min_col=headers["配置项"], min_row=2, max_row=allocation.max_row)
    for column_name in ("最新持仓金额", "按当前全资产目标金额"):
        data = Reference(allocation, min_col=headers[column_name], max_col=headers[column_name], min_row=1, max_row=allocation.max_row)
        chart.add_data(data, titles_from_data=True)
    if len(chart.series) >= 2:
        chart.series[0].tx = SeriesLabel(v="当前金额")
        chart.series[1].tx = SeriesLabel(v="动态目标")
    chart.set_categories(categories)
    chart.legend.position = "b"
    dashboard.add_chart(chart, "D20")


def write_excel(output_path: Path, sheets: dict[str, pd.DataFrame]) -> None:
    """把多个 DataFrame 写入一个 Excel 文件。"""
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    for sheet_name in build_output_sheet_order(sheets):
        df = sheets[sheet_name]
        ws = wb.create_sheet(title=sheet_name[:31])
        export_df = df.copy()
        if "Code" in export_df.columns:
            export_df["Code"] = export_df["Code"].map(format_code)

        rows = list(dataframe_to_rows(export_df, index=False, header=True))
        for row_idx, row in enumerate(rows, start=1):
            ws.append(row)
            # 把 Code 列设成文本，防止前导零丢失
            if row_idx == 1 or "Code" not in export_df.columns:
                continue
            code_col = list(export_df.columns).index("Code") + 1
            cell = ws.cell(row=row_idx, column=code_col)
            cell.number_format = "@"
            if cell.value is not None:
                cell.value = format_code(cell.value)

        if sheet_name == "01_今日决策":
            style_action_dashboard(ws, export_df)
        else:
            style_excel_worksheet(ws, sheet_name)

    add_dashboard_chart(wb)
    add_first_year_chart(wb)
    dashboard_name = "07_组合总览" if "07_组合总览" in wb.sheetnames else "06_组合总览" if "06_组合总览" in wb.sheetnames else "Dashboard"
    if dashboard_name in wb.sheetnames:
        ws = wb[dashboard_name]
        percentage_labels = {
            "券商截图总仓位", "全资产权益穿透仓位", "证券账户可用现金比例",
            "第一年配置目标占比", "第一年当前已配置占比", "第一年配置完成率", "第一年未配置目标占比",
        }
        amount_labels = {
            "第一年动态目标金额", "第一年当前映射金额", "第一年资金缺口",
            "证券账户总资产", "全资产总额", "券商截图总市值", "券商截图今日盈亏",
            "券商截图持仓收益", "券商截图可用资金", "券商截图可取资金",
            "全资产流动性安全垫代理", "其中：证券可用现金", "其中：国债ETF市值",
            "现金安全垫最低值", "现金安全垫目标值",
        }
        for row in range(2, ws.max_row + 1):
            label = str(ws.cell(row, 1).value or "")
            if label in percentage_labels:
                ws.cell(row, 2).number_format = "0.00%"
            elif label in amount_labels:
                ws.cell(row, 2).number_format = "#,##0;[Red](#,##0);-"
        ws.column_dimensions["A"].width = 26
        ws.column_dimensions["B"].width = 48
    wb.save(output_path)


# ==================== 主程序 ====================
def volume_ratio(amount, avg20_amount) -> float | None:
    amount_f = to_float(amount)
    avg20_f = to_float(avg20_amount)
    if amount_f is None or avg20_f is None or avg20_f <= 0:
        return None
    return round(amount_f / avg20_f, 2)


def describe_intraday_position(pos: float | None) -> str:
    if pos is None:
        return "缺少日内数据"
    if pos >= 0.85:
        return "接近日内高位"
    if pos <= 0.25:
        return "接近日内低位"
    return "日内中位震荡"


def describe_price_volume(pct_chg, amount, avg20_amount) -> str:
    pct = to_float(pct_chg)
    vr = volume_pace_ratio(amount, avg20_amount)
    if pct is None:
        return "缺少涨跌幅"
    if pct > 0 and (vr is None or vr <= 1.8):
        return "温和上涨"
    if pct > 0:
        return "放量上涨，防追高"
    if pct < 0 and vr is not None and vr >= 1.2:
        return "放量下跌"
    return "弱势震荡"


def normalize_manual_signal(value) -> str:
    """把人工输入归一为绿/红/待确认。"""
    text = str(value or "").strip().lower()
    if text in {"绿", "通过", "是", "ok", "pass", "1"}:
        return "绿"
    if text in {"红", "不通过", "否", "fail", "0"}:
        return "红"
    return "待确认"


def format_buy_range_recommendation(
    code: str,
    name: str,
    technical_signal: str,
    latest,
    high,
    low,
    prev_close,
    prev_day_low,
    quote_time,
    data_source: str,
    hard_block_kind: str,
    hard_block_reason: str,
    veto_reason: str,
    prev_day_low_source: str = "",
    now: datetime | None = None,
) -> str:
    """把可复算价格区间转换为带权限语义的候选表文本。"""
    current = now or datetime.now()

    def reference_range_text(label: str = "观察区间") -> str:
        reference_range = calculate_reference_buy_range(
            code=code,
            name=name,
            latest=latest,
            high=high,
            low=low,
            prev_close=prev_close,
            prev_day_low=prev_day_low,
        )
        if reference_range is None:
            return ""
        return f"{label} {format_price_range(reference_range, code, name)}"

    def blocked_with_reference(reason: str, label: str = "观察区间") -> str:
        range_text = reference_range_text(label)
        return f"暂不建议买入（{reason}）" + (f"；{range_text}" if range_text else "")

    def blocked_without_range(reason: str) -> str:
        return f"暂不建议买入（{reason}）"

    data_issue = buy_range_data_issue(
        latest=latest,
        high=high,
        low=low,
        prev_close=prev_close,
        prev_day_low=prev_day_low,
        quote_time=quote_time,
        data_source=data_source,
        prev_day_low_source=prev_day_low_source,
        now=current,
    )
    if data_issue:
        return blocked_without_range(f"行情未刷新或数据不完整：{data_issue}")

    if str(veto_reason or "").strip() not in {"", "无"}:
        return blocked_without_range(str(veto_reason).strip())
    if hard_block_kind and hard_block_kind != "position_full":
        return blocked_without_range(hard_block_reason or "存在硬性阻断")

    range_signal = technical_signal
    if hard_block_kind == "position_full" and range_signal not in {"绿", "黄"}:
        range_signal = "黄"
    price_range = calculate_buy_range(
        code=code,
        name=name,
        signal=range_signal,
        latest=latest,
        high=high,
        low=low,
        prev_close=prev_close,
        prev_day_low=prev_day_low,
    )
    if price_range is None:
        prev_low_text = format_price_value(prev_day_low, code, name)
        if prev_low_text and to_float(latest) is not None and to_float(prev_day_low) is not None and to_float(latest) < to_float(prev_day_low):
            return f"暂不建议买入（等待重新站稳昨日低点 {prev_low_text}）"
        return blocked_without_range("买点过滤器未通过")

    range_text = format_price_range(price_range, code, name)
    quote_dt = pd.to_datetime(quote_time).to_pydatetime()
    is_next_session_reference = quote_dt.date() < current.date()
    if hard_block_kind == "position_full":
        label = "下一交易日观察区间" if is_next_session_reference else "观察区间"
        return f"暂不建议买入；{label} {range_text}"
    if technical_signal == "绿":
        label = "下一交易日标准复核区间" if is_next_session_reference else "标准复核区间"
        return f"{label} {range_text}"
    if technical_signal == "黄":
        label = "下一交易日半额复核区间" if is_next_session_reference else "半额复核区间"
        return f"{label} {range_text}"
    return blocked_without_range("买点过滤器未通过")


def build_buy_filter(
    watchlist: pd.DataFrame,
    market_data: pd.DataFrame,
    positions: pd.DataFrame | None = None,
    quality_score: pd.DataFrame | None = None,
    emotion: pd.DataFrame | None = None,
    first_year: pd.DataFrame | None = None,
) -> pd.DataFrame:
    """V2.8.5 buy filter with quality, emotion, full-asset position and role constraints."""
    market_by_code = market_data.set_index("Code")
    quality_by_code = quality_score.set_index("Code") if quality_score is not None and not quality_score.empty else None
    annual_lookup = first_year_lookup(first_year) if first_year is not None else {}
    emotion_temperature = 1
    if emotion is not None and not emotion.empty:
        temperature_row = emotion[emotion["指标"] == "情绪温度"]
        if not temperature_row.empty:
            emotion_temperature = int(to_float(temperature_row.iloc[0]["数值"]) or 1)
    broker_weight_by_code: dict[str, float] = {}
    broker_target_by_code: dict[str, float] = {}
    if positions is not None and not positions.empty:
        for _, position in positions.iterrows():
            code = format_code(position.get("Code"))
            weight = to_float(position.get("Weight"))
            if weight is not None:
                broker_weight_by_code[code] = weight
            target = to_float(position.get("Target Weight"))
            if target is not None:
                broker_target_by_code[code] = target
    rows: list[dict] = []

    for _, item in watchlist.iterrows():
        code = format_code(item["Code"])
        role = item.get("Role", "")
        theme = item.get("Theme", "")
        asset_type = item.get("Asset Type", "")
        notes = str(item.get("Notes", "") or "").strip()
        volume_profile = volume_structure_profile(code, item.get("Name", ""), role, theme, asset_type, notes)
        cognitive_profile = cognitive_guardrail_profile(code, item.get("Name", ""), role, theme, asset_type, notes)
        quant_role = quant_role_profile(code, item.get("Name", ""), role, theme, asset_type, notes)
        dividend_profile = dividend_stock_profile(code, item.get("Name", ""), role, theme, asset_type, notes)
        quote = market_by_code.loc[code] if code in market_by_code.index else None
        quality_row = quality_by_code.loc[code] if quality_by_code is not None and code in quality_by_code.index else None
        quality_value = to_float(quality_row.get("折算质量分")) if quality_row is not None else None
        quality_status = str(quality_row.get("质量状态", "待确认")) if quality_row is not None else "待确认"
        quality_input_status = str(quality_row.get("评分输入状态", "待确认")) if quality_row is not None else "待确认"
        quality_missing_items = str(quality_row.get("评分缺失项", "未生成质量评分诊断")) if quality_row is not None else "未生成质量评分诊断"

        broker_target_weight = first_valid(to_float(item.get("Target Weight")), broker_target_by_code.get(code))
        broker_current_weight = first_valid(
            broker_weight_by_code.get(code),
            to_float(item.get("Current Weight")),
        )
        broker_position_status = describe_weight_status(
            to_float(broker_current_weight),
            to_float(broker_target_weight),
            label="证券账户目标",
        )

        annual = annual_lookup.get(code, {})
        annual_target_weight = to_float(annual.get("年度目标占比"))
        annual_current_weight = to_float(annual.get("当前全资产占比"))
        has_annual_mapping = bool(annual)
        target_weight = annual_target_weight
        current_weight = annual_current_weight
        decision_basis = "全资产/第一年配置"
        if is_index_code(code):
            decision_basis = "指数锚点，不适用"
        elif not has_annual_mapping:
            decision_basis = "缺少第一年配置映射"
        if current_weight is None and target_weight is not None and not is_index_code(code):
            current_weight = 0.0
        remaining_weight = None
        if target_weight is not None and current_weight is not None:
            remaining_weight = max(target_weight - current_weight, 0.0)

        hard_block_kind = ""
        hard_block_reason = ""
        if is_index_code(code):
            position_status = "指数锚点（不可交易）"
            hard_block_kind = "index"
            hard_block_reason = "指数仅作市场锚点，不执行买入"
        elif target_weight is None:
            position_status = "未映射第一年全资产目标"
            hard_block_kind = "missing_annual_target"
            hard_block_reason = "缺少第一年配置目标，禁止按证券账户口径替代买入额度"
        elif current_weight is None:
            position_status = "当前仓位缺失，需人工确认"
            hard_block_kind = "missing_weight"
            hard_block_reason = "全资产仓位缺失，禁止自动给出买入建议"
        else:
            position_status = describe_weight_status(current_weight, target_weight, label="第一年目标")

        if not hard_block_kind and "遗留仓" in str(role):
            hard_block_kind = "legacy"
            hard_block_reason = "遗留仓禁止新买，按减仓/退出纪律处理"
        elif not hard_block_kind and "持有不加" in notes:
            hard_block_kind = "hold_no_add"
            hard_block_reason = "持有不加，禁止新增仓位"
        elif (
            not hard_block_kind
            and target_weight is not None
            and current_weight is not None
            and current_weight >= target_weight
        ):
            hard_block_kind = "position_full"
            hard_block_reason = "全资产仓位已达或超过第一年配置目标"
        elif "数据不足" in quality_status:
            hard_block_kind = "quality_missing"
            hard_block_reason = "缺少完整质量评分或证据，禁止新增"
        elif "低于6分" in quality_status:
            hard_block_kind = "quality"
            hard_block_reason = "质量评分低于6分，禁止新增"
        elif emotion_temperature >= 4:
            hard_block_kind = "emotion"
            hard_block_reason = f"情绪温度{emotion_temperature}级，按V2.8.5暂停新增"

        latest = quote["Latest"] if quote is not None else None
        pct_chg = quote["PctChg"] if quote is not None else None
        open_price = quote["Open"] if quote is not None else None
        high = quote["High"] if quote is not None else None
        low = quote["Low"] if quote is not None else None
        prev_close = quote["PrevClose"] if quote is not None else None
        prev_day_low = quote["PrevDayLow"] if quote is not None else None
        prev_day_low_source = quote["PrevDayLowSource"] if quote is not None and "PrevDayLowSource" in quote.index else ""
        quote_time = quote["QuoteTime"] if quote is not None else None
        data_source = quote["DataSource"] if quote is not None else ""
        interface_status = quote["接口状态"] if quote is not None and "接口状态" in quote.index else "行情未刷新"
        failure_reason = quote["失败原因"] if quote is not None and "失败原因" in quote.index else "未生成接口诊断"
        retry_result = quote["重试结果"] if quote is not None and "重试结果" in quote.index else "未生成重试诊断"
        amount = quote["Amount"] if quote is not None else None
        avg20_amount = quote["Avg20Amount"] if quote is not None else None
        avg20_source = quote["Avg20AmountSource"] if quote is not None and "Avg20AmountSource" in quote.index else ""

        intraday_pos = calc_intraday_position(latest, low, high)
        vr = volume_pace_ratio(amount, avg20_amount)
        name = quote["Name"] if quote is not None and first_valid(quote["Name"]) is not None else item.get("Name", "")
        share_proxy = describe_share_proxy(amount, avg20_amount, avg20_source)
        premium_proxy = describe_premium_proxy(pct_chg, intraday_pos)
        leader_proxy = describe_leader_proxy(code, name, role, pct_chg)
        feedback = feedback_state(pct_chg, intraday_pos, vr)

        veto = "否"
        veto_reason = "无"
        manual_veto = str(item.get("One-Vote Veto", "") or "").strip().lower()
        if manual_veto in {"是", "红", "否决", "yes", "true", "1"}:
            veto = "是"
            veto_reason = "人工一票否决"
        open_f = to_float(open_price)
        prev_f = to_float(prev_close)
        if open_f is not None and prev_f is not None and intraday_pos is not None:
            if open_f > prev_f and intraday_pos <= 0.20:
                veto = "是"
                veto_reason = "高开后接近日内低位" if veto_reason == "无" else f"{veto_reason}；高开后接近日内低位"

        pass_items: list[str] = []
        fail_items: list[str] = []
        pending_items: list[str] = []
        pct = to_float(pct_chg)

        # 六项检查 1：分时结构。
        if intraday_pos is None:
            pending_items.append("分时结构")
        elif intraday_pos <= 0.20:
            fail_items.append("分时结构走弱")
        elif intraday_pos < 0.85:
            pass_items.append("分时结构")
        else:
            fail_items.append("分时位置过高，防追涨")

        # 六项检查 2：量价关系（使用交易时段校正后的成交节奏）。
        avg20_is_proxy = "替代" in str(avg20_source or "")
        if pct is None or vr is None:
            pending_items.append("量价关系")
        elif avg20_is_proxy:
            pending_items.append("量价关系（均额基准为替代）")
        elif pct < 0 and vr >= 1.2:
            fail_items.append("放量下跌")
            veto = "是"
            veto_reason = "放量下跌" if veto_reason == "无" else f"{veto_reason}；放量下跌"
        elif (pct >= 0 and 0.6 <= vr <= 2.5) or (-1.0 <= pct < 0 and vr < 1.0 and (intraday_pos or 0) >= 0.35):
            pass_items.append("量价关系")
        else:
            fail_items.append("量价关系不合格")

        # 六项检查 3：ETF份额或个股筹码。
        share_signal = normalize_manual_signal(item.get("Share/Chip Signal"))
        share_change = to_float(quote["ETFShareChg"] if quote is not None else None)
        if share_signal == "绿" or (share_signal == "待确认" and share_change is not None and share_change >= -0.02):
            pass_items.append("份额/筹码")
        elif share_signal == "红" or (share_change is not None and share_change < -0.02):
            fail_items.append("份额/筹码恶化")
            veto = "是"
            veto_reason = "份额/筹码恶化" if veto_reason == "无" else f"{veto_reason}；份额/筹码恶化"
        else:
            pending_items.append("份额/筹码")

        # 六项检查 4：ETF折溢价；个股可人工填估值/安全边际信号。
        premium_signal = normalize_manual_signal(item.get("Premium Signal"))
        premium = to_float(quote["Premium"] if quote is not None else None)
        if premium_signal == "绿" or (premium_signal == "待确认" and premium is not None and abs(premium) < 0.5):
            pass_items.append("折溢价/估值")
        elif premium_signal == "红" or (premium is not None and abs(premium) >= 0.5):
            fail_items.append("折溢价/估值无安全垫")
        else:
            pending_items.append("折溢价/估值")

        # 六项检查 5：龙头同步。
        leader_signal = normalize_manual_signal(item.get("Leader Signal"))
        if leader_signal == "绿":
            pass_items.append("龙头同步")
        elif leader_signal == "红":
            fail_items.append("核心龙头破位或不同步")
            veto = "是"
            veto_reason = "核心龙头破位" if veto_reason == "无" else f"{veto_reason}；核心龙头破位"
        else:
            pending_items.append("龙头同步")

        # 六项检查 6：次日验证。
        next_day_signal = normalize_manual_signal(item.get("Next-Day Signal"))
        if next_day_signal == "绿":
            pass_items.append("次日验证")
        elif next_day_signal == "红":
            fail_items.append("次日验证失败")
            veto = "是"
            veto_reason = "次日跌穿前日低点" if veto_reason == "无" else f"{veto_reason}；次日跌穿前日低点"
        else:
            pending_items.append("次日验证")

        pass_count = len(pass_items)
        if veto == "是":
            technical_signal = "红"
        elif pass_count >= BUY_STANDARD_MIN:
            technical_signal = "绿"
        elif pass_count >= BUY_HALF_MIN:
            technical_signal = "黄"
        else:
            technical_signal = "灰"
        if avg20_is_proxy and technical_signal == "绿":
            technical_signal = "黄"

        quant_validation = quant_validation_result(
            pass_items=pass_items,
            fail_items=fail_items,
            pending_items=pending_items,
            quality_value=quality_value,
            emotion_temperature=emotion_temperature,
            veto=veto,
            hard_block_kind=hard_block_kind,
        )

        range_block_kind = hard_block_kind
        range_block_reason = hard_block_reason
        if hard_block_kind == "position_full":
            if "数据不足" in quality_status:
                range_block_kind = "quality_missing"
                range_block_reason = "缺少完整质量评分或证据，禁止新增"
            elif "低于6分" in quality_status:
                range_block_kind = "quality"
                range_block_reason = "质量评分低于6分，禁止新增"
            elif emotion_temperature >= 4:
                range_block_kind = "emotion"
                range_block_reason = f"情绪温度{emotion_temperature}级，按V2.8.5暂停新增"
        buy_range_recommendation = format_buy_range_recommendation(
            code=code,
            name=name,
            technical_signal=technical_signal,
            latest=latest,
            high=high,
            low=low,
            prev_close=prev_close,
            prev_day_low=prev_day_low,
            quote_time=quote_time,
            data_source=data_source,
            hard_block_kind=range_block_kind,
            hard_block_reason=range_block_reason,
            veto_reason=veto_reason,
            prev_day_low_source=prev_day_low_source,
        )
        if hard_block_kind:
            veto = "是"
            veto_reason = hard_block_reason if veto_reason == "无" else f"{hard_block_reason}；{veto_reason}"
            decision_level = "红"
            if hard_block_kind == "index":
                suggestion = "指数锚点，仅观察，不执行买入"
            elif hard_block_kind == "legacy":
                suggestion = f"遗留仓禁止新买；{position_status}"
            elif hard_block_kind == "hold_no_add":
                suggestion = f"持有不加，禁止新增；{position_status}"
            elif hard_block_kind == "missing_weight":
                suggestion = "仓位数据缺失，先核对持仓，禁止买入"
            elif hard_block_kind == "missing_annual_target":
                suggestion = "缺少第一年全资产配置目标，先补映射，禁止新增"
            elif hard_block_kind == "quality":
                suggestion = "质量评分低于6分，禁止新增"
            elif hard_block_kind == "quality_missing":
                suggestion = "质量评分或证据不完整，先补评分，禁止新增"
            elif hard_block_kind == "emotion":
                suggestion = f"情绪温度{emotion_temperature}级，暂停新增"
            else:
                suggestion = f"{position_status}，禁止加仓"
        elif veto == "是":
            suggestion = "不买/观察"
            decision_level = "红"
        elif pass_count >= BUY_STANDARD_MIN:
            if remaining_weight is not None:
                suggestion = f"可复核标准首批；累计加仓不超过剩余额度 {remaining_weight:.2%}"
            else:
                suggestion = "可复核标准首批；需人工确认仓位上限"
            decision_level = "绿"
        elif pass_count >= BUY_HALF_MIN:
            if remaining_weight is not None:
                suggestion = f"仅可复核半额首批；剩余额度 {remaining_weight:.2%}"
            else:
                suggestion = "仅可复核半额首批；需人工确认仓位上限"
            decision_level = "黄"
        else:
            suggestion = "不买/观察（买点不合格；仓位额度不等于买入信号）"
            decision_level = "灰"
        if avg20_is_proxy and decision_level == "绿":
            suggestion = "可观察（量能基准为替代，需复核）"
            decision_level = "黄"

        row = {
            "买点灯号": decision_level,
            "Code": code,
            "Name": name,
            "建议买入区间": buy_range_recommendation,
            "Role": role,
            "质量评分": quality_value,
            "质量状态": quality_status,
            "情绪温度": emotion_temperature,
            "反馈状态": feedback,
            "趋势结构": volume_profile["趋势结构"],
            "关键支撑纪律": volume_profile["关键支撑纪律"],
            "单次仓位上限": volume_profile["单次仓位上限"],
            "爆量/量质吸纪律": volume_profile["爆量/量质吸纪律"],
            "机会成本检查": cognitive_profile["机会成本检查"],
            "认知防错": cognitive_profile["认知防错"],
            "决策树情景": cognitive_profile["决策树情景"],
            "赔率/最大损失": cognitive_profile["赔率/最大损失"],
            "反向失败路径": cognitive_profile["反向失败路径"],
            "能力圈边界": cognitive_profile["能力圈边界"],
            "资产角色": quant_role["资产角色"],
            "量化验证分": quant_validation["量化验证分"],
            "量化验证状态": quant_validation["量化验证状态"],
            "量化买点规则": quant_role["量化买点规则"],
            "正期望检查": quant_validation["正期望检查"],
            "量化风控规则": quant_role["量化风控规则"],
            "红利配置角色": dividend_profile["红利配置角色"],
            "分红质量检查": dividend_profile["分红质量检查"],
            "红利买点规则": dividend_profile["红利买点规则"],
            "高息陷阱否决": dividend_profile["高息陷阱否决"],
            "当前仓位": current_weight,
            "目标仓位": target_weight,
            "剩余额度": remaining_weight,
            "仓位状态": position_status,
            "仓位决策口径": decision_basis,
            "全资产当前仓位": current_weight,
            "全资产目标仓位": target_weight,
            "全资产剩余额度": remaining_weight,
            "全资产仓位状态": position_status,
            "证券账户仓位": to_float(broker_current_weight),
            "证券账户目标仓位": to_float(broker_target_weight),
            "证券账户仓位状态": broker_position_status,
            "Latest": latest,
            "PctChg": pct_chg,
            "Open": open_price,
            "High": high,
            "Low": low,
            "PrevClose": prev_close,
            "PrevDayLow": prev_day_low,
            "PrevDayLowSource": prev_day_low_source,
            "QuoteTime": quote_time,
            "DataSource": data_source,
            "接口状态": interface_status,
            "失败原因": failure_reason,
            "重试结果": retry_result,
            "评分输入状态": quality_input_status,
            "评分缺失项": quality_missing_items,
            "日内位置": intraday_pos,
            "量能倍数": vr,
            "分时结构": describe_intraday_position(intraday_pos),
            "量价关系": describe_price_volume(pct_chg, amount, avg20_amount),
            "份额变动": first_valid_text(quote["ETFShareChg"] if quote is not None else None, fallback=share_proxy),
            "折溢价": first_valid_text(quote["Premium"] if quote is not None else None, fallback=premium_proxy),
            "龙头同步": first_valid_text(item.get("Leader Signal"), fallback="待人工确认"),
            "次日验证": first_valid_text(item.get("Next-Day Signal"), fallback="待次日确认"),
            "通过项": pass_count,
            "通过明细": "；".join(pass_items) if pass_items else "无",
            "未通过项": "；".join(fail_items) if fail_items else "无",
            "待确认项": "；".join(pending_items) if pending_items else "无",
            "一票否决": veto,
            "否决原因": veto_reason,
            "建议": suggestion,
        }
        row.update(evaluate_stage_progression(pd.Series(row)))
        rows.append(row)

    return pd.DataFrame(rows, columns=BUY_FILTER_COLUMNS)


def main() -> int:
    print("=" * 60)
    print(f"{FRAMEWORK_DISPLAY_VERSION} {PRODUCT_NAME}")
    print("=" * 60)
    print("提示：本程序只生成 Excel 纪律提醒，不会下单。")
    print()

    watchlist = load_watchlist()
    decision_inputs = load_decision_inputs()
    buy_point_plan_source = load_buy_point_plan()
    first_year_source = load_first_year_allocation()
    account_meta, positions = load_positions()
    print(f"已读取观察池 {len(watchlist)} 条，持仓 {len(positions)} 条，第一年配置 {len(first_year_source)} 项。")

    spot_map = fetch_all_spot()

    market_data = build_market_data(watchlist, spot_map, positions)
    market_data = enrich_market_data_with_eastmoney(market_data)
    positions_sheet, alerts = build_positions_sheet(positions, market_data, account_meta)
    first_year = build_first_year_allocation(first_year_source, positions_sheet, account_meta)
    first_year_summary = summarize_first_year_allocation(first_year, account_meta)
    double_anchor = build_double_anchor(market_data)
    quality_score = build_quality_score(watchlist, market_data)
    emotion = build_emotion_thermometer(market_data, decision_inputs)
    buy_filter = build_buy_filter(watchlist, market_data, positions, quality_score, emotion, first_year)
    positions_action = build_positions_action(positions, market_data)
    broker_snapshot = build_broker_snapshot(positions)
    exposure = build_exposure_summary(positions_sheet, account_meta)
    market_data = attach_first_year_fields(market_data, first_year)
    quality_score = attach_first_year_fields(quality_score, first_year)
    buy_filter = attach_first_year_fields(buy_filter, first_year)
    positions_sheet = attach_first_year_fields(positions_sheet, first_year)
    positions_action = attach_first_year_fields(positions_action, first_year)
    broker_snapshot = attach_first_year_fields(broker_snapshot, first_year)
    exposure = attach_first_year_fields(exposure, first_year)
    watchlist_output = attach_first_year_fields(watchlist, first_year)
    rules = build_framework_rules(decision_inputs, first_year)
    investment_profile = build_investment_profile(decision_inputs)
    data_sources = build_data_sources()
    checks = build_checks(watchlist, positions, positions_sheet, account_meta, market_data, quality_score, first_year)
    execution_plan = build_execution_plan(buy_filter, positions_action, positions_sheet, first_year)
    decision_center = build_decision_center(double_anchor, emotion, quality_score, buy_filter, positions_action, first_year_summary, decision_inputs)
    dashboard = build_dashboard(account_meta, alerts, market_data, double_anchor, emotion, buy_filter, quality_score, positions_sheet, decision_inputs, first_year_summary)
    portfolio_overview = build_portfolio_overview(dashboard)
    permission_row = decision_center[decision_center["层级"] == "市场权限"]
    market_permission = str(permission_row.iloc[0]["状态"]) if not permission_row.empty else "待确认"
    stage_progression = build_stage_progression_view(buy_filter, market_permission)
    buy_candidates = build_buy_candidates_view(buy_filter, market_permission)
    buy_point_plan = build_buy_point_plan_view(buy_point_plan_source, positions_sheet, buy_filter)
    position_risk = build_position_risk_view(positions_action, positions_sheet)
    action_plan = build_action_plan_view(execution_plan, buy_filter, positions_sheet)
    action_dashboard = build_action_dashboard_view(
        decision_center,
        action_plan,
        buy_candidates,
        position_risk,
        dashboard,
    )
    long_term_tracking = build_long_term_tracking_view(
        watchlist_output,
        market_data,
        quality_score,
        buy_filter,
        first_year,
    )

    # 把原始 watchlist 也写进去，方便对照
    output_xlsx = make_output_xlsx_path()
    sheets = {
        "01_今日决策": action_dashboard,
        "02_今日动作": action_plan,
        "03_买入候选": buy_candidates,
        "04_阶段推进": stage_progression,
        "买点计划": buy_point_plan,
        "05_持仓风险": position_risk,
        "06_年度配置": first_year,
        "07_组合总览": portfolio_overview,
        "Double_Anchor": double_anchor,
        "Emotion": emotion,
        "Quality_Score": quality_score,
        "Exposure": exposure,
        "Market_Data": market_data,
        "Buy_Filter": buy_filter,
        "Positions": positions_sheet,
        "Broker_Snapshot": broker_snapshot,
        "Watchlist": watchlist_output,
        "长期跟踪个股": long_term_tracking,
        "Investment_Profile": investment_profile,
        "Data_Sources": data_sources,
        "Buy_Point_Plan_Source": buy_point_plan_source,
        "Framework_Rules": rules,
        "Checks": checks,
        "使用说明": pd.DataFrame(
            [
                {"步骤": "1", "操作": "编辑 watchlist.csv", "说明": "维护观察池代码、角色、目标权重"},
                {"步骤": "2", "操作": "补充质量与六项信号", "说明": "在watchlist.csv填写完整质量分、评分证据、份额/筹码、折溢价/估值、龙头同步和次日验证"},
                {"步骤": "3", "操作": "补充量化验证口径", "说明": "按资产角色确认指数环境、ETF买点、情绪温度、正期望和风险监控；资金流入不能单独触发买入"},
                {"步骤": "4", "操作": "编辑 decision_inputs.csv", "说明": "可选：填写1-5级人工情绪温度、主线周期、趋势买点类型、量化模块状态和退潮三因子状态；空白时只显示框架默认或低可靠性代理"},
                {"步骤": "5", "操作": "编辑 first_year_allocation.csv", "说明": "维护第一年全资产目标、代码映射、收益观察区间与执行约束；CASH代表证券账户可用现金，年度缺口不是买入信号"},
                {"步骤": "6", "操作": "编辑 investment_profile.csv", "说明": "维护投资偏好、纪律开关、建议刷新时间和展示偏好"},
                {"步骤": "7", "操作": "查看 Data_Sources", "说明": "确认行情、持仓、年度配置、量化验证和盘前情报的数据来源与失败处理"},
                {"步骤": "8", "操作": "查看 长期跟踪个股", "说明": "长期跟踪股票进入每日行情监控；未补完整评分前只观察不自动买入"},
                {"步骤": "9", "操作": "编辑 positions.csv", "说明": "维护持仓数量、成本、账户总资产和券商截图持仓快照"},
                {"步骤": "10", "操作": "运行 python main.py", "说明": "抓取行情并生成 Excel"},
                {"步骤": "11", "操作": f"打开 output 文件夹里最新的 {FRAMEWORK_VERSION}_每日行情输出_日期时间.xlsx", "说明": "按顺序查看 01_今日决策、02_今日动作、03_买入候选、04_阶段推进、买点计划、05_持仓风险、06_年度配置和07_组合总览"},
                {"步骤": "12", "操作": "理解双口径", "说明": "买入和减仓建议使用全资产/第一年配置口径；证券账户口径只作交易集中度参考"},
                {"步骤": "13", "操作": "理解数据边界", "说明": "缺失项不得自动判绿；代理分不折算为完整评分"},
            ]
        ),
    }

    write_excel(output_xlsx, sheets)

    print()
    print(f"已生成 Excel：{output_xlsx}")
    print("请先查看 01_今日决策、02_今日动作、03_买入候选、04_阶段推进、买点计划和05_持仓风险，再查看年度配置与组合总览。")
    if missing_latest_mask(market_data["Latest"]).any():
        print("注意：部分代码未取到行情，可能是网络问题、代码错误或接口暂时不可用。")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except KeyboardInterrupt:
        print("\n用户中断。")
        raise SystemExit(1)
    except Exception as exc:
        print(f"\n【程序出错】{exc}")
        raise SystemExit(1)
