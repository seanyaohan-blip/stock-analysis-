from pathlib import Path
import sys
import tempfile
import unittest
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook


sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import main


class ActionFirstDashboardTests(unittest.TestCase):
    def test_calculate_buy_range_uses_pullback_band_for_etf(self):
        result = main.calculate_buy_range(
            code="159516",
            name="半导体设备ETF",
            signal="绿",
            latest=1.520,
            high=1.550,
            low=1.480,
            prev_close=1.510,
            prev_day_low=1.490,
        )

        self.assertEqual(result, (1.498, 1.515))

    def test_calculate_buy_range_uses_stricter_yellow_band_for_stock(self):
        result = main.calculate_buy_range(
            code="002594",
            name="比亚迪",
            signal="黄",
            latest=90.00,
            high=92.00,
            low=86.00,
            prev_close=89.00,
            prev_day_low=86.50,
        )

        self.assertEqual(result, (86.90, 88.10))

    def test_quote_is_recent_accepts_friday_on_saturday(self):
        self.assertTrue(
            main.is_recent_complete_quote(
                "2026-06-19 15:00:00",
                now=datetime(2026, 6, 20, 10, 0, 0),
            )
        )

    def test_quote_is_recent_rejects_missing_or_old_timestamp(self):
        now = datetime(2026, 6, 20, 10, 0, 0)

        self.assertFalse(main.is_recent_complete_quote(None, now=now))
        self.assertFalse(main.is_recent_complete_quote("2026-06-15 15:00:00", now=now))

    def test_format_buy_range_allows_full_position_observation_exception(self):
        text = main.format_buy_range_recommendation(
            code="159516",
            name="半导体设备ETF",
            technical_signal="绿",
            latest=1.520,
            high=1.550,
            low=1.480,
            prev_close=1.510,
            prev_day_low=1.490,
            quote_time="2026-06-19 15:00:00",
            data_source="东方财富补齐",
            hard_block_kind="position_full",
            hard_block_reason="当前仓位已达目标仓位",
            veto_reason="无",
            now=datetime(2026, 6, 20, 10, 0, 0),
        )

        self.assertEqual(text, "暂不建议买入；下一交易日观察区间 1.498–1.515")

    def test_format_buy_range_hides_numeric_reference_for_non_position_hard_blocks(self):
        text = main.format_buy_range_recommendation(
            code="159516",
            name="半导体设备ETF",
            technical_signal="绿",
            latest=1.520,
            high=1.550,
            low=1.480,
            prev_close=1.510,
            prev_day_low=1.490,
            quote_time="2026-06-19 15:00:00",
            data_source="东方财富补齐",
            hard_block_kind="quality",
            hard_block_reason="质量评分不足",
            veto_reason="无",
            now=datetime(2026, 6, 20, 10, 0, 0),
        )

        self.assertEqual(text, "暂不建议买入（质量评分不足）")
        self.assertNotRegex(text, r"\d+\.\d+–\d+\.\d+")

    def test_format_buy_range_reports_incomplete_price_fields(self):
        text = main.format_buy_range_recommendation(
            code="159516",
            name="半导体设备ETF",
            technical_signal="绿",
            latest=1.520,
            high=None,
            low=1.480,
            prev_close=1.510,
            prev_day_low=1.490,
            quote_time="2026-06-24 14:00:00",
            data_source="东方财富补齐",
            hard_block_kind="",
            hard_block_reason="",
            veto_reason="无",
            now=datetime(2026, 6, 24, 14, 30, 0),
        )

        self.assertEqual(text, "暂不建议买入（行情未刷新或数据不完整：最高）")
        self.assertNotRegex(text, r"\d+\.\d+–\d+\.\d+")

    def test_format_buy_range_reports_prev_low_value_when_waiting_to_reclaim(self):
        text = main.format_buy_range_recommendation(
            code="159516",
            name="半导体设备ETF",
            technical_signal="红",
            latest=1.480,
            high=1.500,
            low=1.470,
            prev_close=1.510,
            prev_day_low=1.490,
            quote_time="2026-06-24 14:00:00",
            data_source="东方财富补齐",
            hard_block_kind="",
            hard_block_reason="",
            veto_reason="无",
            now=datetime(2026, 6, 24, 14, 30, 0),
        )

        self.assertEqual(text, "暂不建议买入（等待重新站稳昨日低点 1.490）")
        self.assertNotRegex(text, r"\d+\.\d+–\d+\.\d+")

    def test_market_interface_diagnostics_records_quote_failure(self):
        row = {
            "Code": "688012",
            "Latest": None,
            "PctChg": None,
            "Open": None,
            "High": None,
            "Low": None,
            "PrevClose": None,
            "Amount": None,
            "Avg20Amount": None,
            "PrevDayLow": None,
            "QuoteTime": None,
            "DataSource": "券商截图",
        }

        diagnostics = main.build_market_interface_diagnostics(row, now=datetime(2026, 6, 24, 11, 0, 0))

        self.assertEqual(diagnostics["接口状态"], "行情未刷新")
        self.assertIn("东方财富/腾讯", diagnostics["失败原因"])
        self.assertIn("日K首轮失败后已重试", diagnostics["重试结果"])

    def test_market_interface_diagnostics_rejects_substitute_prev_day_low(self):
        row = {
            "Code": "159516",
            "Latest": 1.52,
            "PctChg": 0.6,
            "Open": 1.50,
            "High": 1.55,
            "Low": 1.48,
            "PrevClose": 1.51,
            "Amount": 100000000,
            "Avg20Amount": 100000000,
            "Avg20AmountSource": "东方财富日K统计",
            "PrevDayLow": 1.48,
            "PrevDayLowSource": "当日低点替代",
            "QuoteTime": "2026-06-24 14:00:00",
            "DataSource": "东方财富补齐",
        }

        diagnostics = main.build_market_interface_diagnostics(row, now=datetime(2026, 6, 24, 14, 30, 0))

        self.assertEqual(diagnostics["接口状态"], "行情字段缺失")
        self.assertIn("昨日低点来源", diagnostics["重试结果"])

    def test_market_interface_diagnostics_rejects_dash_price_fields(self):
        row = {
            "Code": "688072",
            "Latest": "-",
            "PctChg": "-",
            "Open": "-",
            "High": "-",
            "Low": "-",
            "PrevClose": 832,
            "Amount": "-",
            "Avg20Amount": 6503737847.6,
            "PrevDayLow": 799.99,
            "PrevDayLowSource": "东方财富日K统计",
            "QuoteTime": "2026-06-29 08:30:14",
            "DataSource": "行情接口",
        }

        diagnostics = main.build_market_interface_diagnostics(row, now=datetime(2026, 6, 29, 9, 55, 0))

        self.assertEqual(diagnostics["接口状态"], "行情字段缺失")
        self.assertIn("最新价", diagnostics["失败原因"])
        self.assertIn("涨跌幅", diagnostics["失败原因"])
        self.assertNotEqual(diagnostics["失败原因"], "无")
        self.assertNotEqual(diagnostics["重试结果"], "实时行情与日K补齐可用")

    def test_checks_count_dash_latest_as_missing_quote(self):
        watchlist = pd.DataFrame([{"Code": "688072"}])
        positions = pd.DataFrame(columns=["Code", "Market Value", "Weight"])
        positions_sheet = pd.DataFrame()
        account_meta = {"broker_market_value": 0, "position_ratio": 0, "total_assets": 1000000, "account_total": 100000}
        market_data = pd.DataFrame(
            [
                {
                    "Code": "688072",
                    "Latest": "-",
                    "接口状态": "行情字段缺失",
                }
            ]
        )
        quality_score = pd.DataFrame([{"质量状态": "数据不足，需人工评分"}])
        first_year = pd.DataFrame()

        checks = main.build_checks(
            watchlist,
            positions,
            positions_sheet,
            account_meta,
            market_data,
            quality_score,
            first_year,
        )

        quote_check = checks.loc[checks["检查项"] == "行情完整性"].iloc[0]
        interface_check = checks.loc[checks["检查项"] == "行情接口诊断"].iloc[0]
        self.assertEqual(quote_check["实际"], 1)
        self.assertEqual(quote_check["状态"], "检查")
        self.assertEqual(interface_check["实际"], 1)
        self.assertEqual(interface_check["状态"], "检查")

    def test_quality_score_manual_input_is_independent_from_quote_refresh(self):
        watchlist = pd.DataFrame(
            [
                {
                    "Code": "688012",
                    "Name": "中微公司",
                    "Role": "半导体设备核心个股",
                    "Asset Type": "Stock",
                    "Manual Quality Score": 8.5,
                    "Quality Evidence": "订单与设备平台竞争力符合核心候选要求",
                }
            ]
        )
        market_data = pd.DataFrame(
            [
                {
                    "Code": "688012",
                    "Name": "中微公司",
                    "Latest": None,
                    "PctChg": None,
                    "Amount": None,
                    "Avg20Amount": None,
                    "DataSource": "券商截图",
                    "QuoteTime": None,
                }
            ]
        )

        result = main.build_quality_score(watchlist, market_data)

        self.assertEqual(result.iloc[0]["质量状态"], "核心候选")
        self.assertEqual(result.iloc[0]["评分输入状态"], "完整")
        self.assertEqual(result.iloc[0]["评分缺失项"], "无")

    def test_market_permission_hides_numeric_range_with_blocker_prefix(self):
        source = pd.DataFrame(
            [
                {
                    "买点灯号": "绿",
                    "Code": "159516",
                    "Name": "半导体设备ETF",
                    "建议买入区间": "标准复核区间 1.498–1.515",
                    "建议": "复核",
                    "否决原因": "",
                    "通过项": 5,
                    "质量评分": 9.0,
                    "剩余额度": 0.02,
                    "Latest": 1.52,
                    "PctChg": 0.6,
                }
            ]
        )

        result = main.build_buy_candidates_view(source, market_permission="暂停标准新增")

        self.assertEqual(result.iloc[0]["建议买入区间"], "暂不建议买入（市场权限：暂停标准新增）")
        self.assertNotRegex(result.iloc[0]["建议买入区间"], r"\d+\.\d+–\d+\.\d+")

    def test_buy_candidates_hides_numeric_range_when_range_data_incomplete(self):
        source = pd.DataFrame(
            [
                {
                    "买点灯号": "绿",
                    "Code": "159516",
                    "Name": "半导体设备ETF",
                    "建议买入区间": "标准复核区间 1.498–1.515",
                    "建议": "复核",
                    "否决原因": "",
                    "通过项": 5,
                    "质量评分": 9.0,
                    "剩余额度": 0.02,
                    "Latest": 1.52,
                    "PctChg": 0.6,
                    "High": 1.55,
                    "Low": 1.48,
                    "PrevClose": 1.51,
                    "PrevDayLow": None,
                    "QuoteTime": datetime.now().strftime("%Y-%m-%d 14:00:00"),
                    "DataSource": "东方财富补齐",
                    "接口状态": "当日行情已刷新",
                }
            ]
        )

        result = main.build_buy_candidates_view(source, market_permission="开放买点复核")

        self.assertEqual(result.iloc[0]["数据状态"], "行情未刷新")
        self.assertIn("昨日低点", result.iloc[0]["建议买入区间"])
        self.assertNotIn("待刷新参考区间", result.iloc[0]["建议买入区间"])
        self.assertNotRegex(result.iloc[0]["建议买入区间"], r"\d+\.\d+–\d+\.\d+")

    def test_buy_range_column_follows_name(self):
        source = pd.DataFrame(
            [
                {
                    "买点灯号": "绿",
                    "Code": "159516",
                    "Name": "半导体设备ETF",
                    "建议买入区间": "标准复核区间 1.498–1.515",
                    "建议": "复核",
                    "否决原因": "",
                    "通过项": 5,
                    "质量评分": 9.0,
                    "剩余额度": 0.02,
                    "Latest": 1.52,
                    "PctChg": 0.6,
                }
            ]
        )

        result = main.build_buy_candidates_view(source, market_permission="开放买点复核")

        self.assertEqual(result.columns[result.columns.get_loc("Name") + 1], "建议买入区间")

    def test_buy_candidates_keep_all_rows_and_sort_by_signal(self):
        buy_filter = pd.DataFrame(
            [
                {"买点灯号": "灰", "Code": "000004", "Name": "灰标的", "建议": "观察", "否决原因": "", "通过项": 1, "质量评分": None, "剩余额度": 0.04, "全资产剩余额度": 0.04, "日内位置": 0.5},
                {"买点灯号": "红", "Code": "000003", "Name": "红标的", "建议": "不买", "否决原因": "质量评分缺失", "通过项": 2, "质量评分": None, "剩余额度": 0.03, "全资产剩余额度": 0.03, "日内位置": 0.5},
                {"买点灯号": "绿", "Code": "000001", "Name": "绿标的", "建议": "复核", "否决原因": "", "通过项": 6, "质量评分": 9.0, "剩余额度": 0.02, "全资产剩余额度": 0.02, "日内位置": 0.5},
                {"买点灯号": "黄", "Code": "000002", "Name": "黄标的", "建议": "半额复核", "否决原因": "", "通过项": 4, "质量评分": 8.0, "剩余额度": 0.01, "全资产剩余额度": 0.01, "日内位置": 0.5},
            ]
        )

        result = main.build_buy_candidates_view(buy_filter, market_permission="暂停标准新增")

        self.assertEqual(result["买点灯号"].tolist(), ["绿", "黄", "红", "灰"])
        self.assertEqual(len(result), len(buy_filter))
        red_reason = result.loc[result["买点灯号"] == "红", "阻断原因"].iloc[0]
        self.assertTrue(str(red_reason).strip())
        self.assertLessEqual(len(result.columns), 42)
        self.assertNotIn("日内位置", result.columns)
        self.assertNotIn("当前仓位", result.columns)
        self.assertIn("资产角色", result.columns)
        self.assertIn("量化验证状态", result.columns)

    def test_buy_candidates_put_incomplete_quote_reason_in_blocker_column(self):
        buy_filter = pd.DataFrame(
            [
                {
                    "买点灯号": "灰",
                    "Code": "688072",
                    "Name": "拓荆科技",
                    "建议": "观察",
                    "否决原因": "无",
                    "建议买入区间": "暂不建议买入（行情未刷新或数据不完整：最新价、最高、最低）",
                    "通过项": 1,
                    "质量评分": 8.5,
                    "剩余额度": 0.01,
                    "全资产剩余额度": 0.01,
                    "Latest": None,
                    "PctChg": None,
                    "High": None,
                    "Low": None,
                    "PrevClose": 832,
                    "PrevDayLow": 799.99,
                    "QuoteTime": "2026-06-30 08:30:28",
                    "DataSource": "行情接口",
                    "接口状态": "行情字段缺失",
                    "失败原因": "实时行情已返回但字段不完整；缺失字段：最新价、涨跌幅、今开、最高、最低、成交额",
                    "重试结果": "实时行情补齐已尝试，未取得当日完整行情",
                }
            ]
        )

        result = main.build_buy_candidates_view(buy_filter, market_permission="开放买点复核")

        self.assertEqual(result.iloc[0]["数据状态"], "行情未刷新")
        self.assertIn("行情未刷新或数据不完整", result.iloc[0]["阻断原因"])
        self.assertIn("最新价", result.iloc[0]["阻断原因"])
        self.assertNotEqual(result.iloc[0]["阻断原因"], "无")

    def test_long_term_tracking_puts_incomplete_quote_reason_in_blocker_column(self):
        watchlist = pd.DataFrame(
            [
                {
                    "Code": "688072",
                    "Name": "拓荆科技",
                    "Role": "长期跟踪个股观察",
                    "Asset Type": "stock",
                    "Theme": "半导体设备",
                    "Thesis": "测试",
                    "Invalidation": "测试",
                }
            ]
        )
        market_data = pd.DataFrame(
            [
                {
                    "Code": "688072",
                    "Name": "拓荆科技",
                    "Latest": None,
                    "PctChg": None,
                    "Open": None,
                    "High": None,
                    "Low": None,
                    "PrevClose": 832,
                    "Amount": None,
                    "QuoteTime": "2026-06-30 08:30:28",
                    "DataSource": "行情接口",
                    "接口状态": "行情字段缺失",
                    "失败原因": "实时行情已返回但字段不完整；缺失字段：最新价、涨跌幅、今开、最高、最低、成交额",
                    "重试结果": "实时行情补齐已尝试，未取得当日完整行情",
                }
            ]
        )
        quality_score = pd.DataFrame(
            [
                {
                    "Code": "688072",
                    "质量状态": "核心候选",
                    "折算质量分": 8.5,
                    "评分输入状态": "完整",
                    "评分缺失项": "无",
                }
            ]
        )
        buy_filter = pd.DataFrame(
            [
                {
                    "Code": "688072",
                    "买点灯号": "灰",
                    "否决原因": "无",
                    "建议买入区间": "暂不建议买入（行情未刷新或数据不完整：最新价、最高、最低）",
                }
            ]
        )
        first_year = pd.DataFrame(
            [
                {
                    "配置项": "AI/半导体高质量个股篮子",
                    "映射代码": "688072",
                    "年度资金缺口": 100000,
                    "年度完成率": 0,
                }
            ]
        )

        result = main.build_long_term_tracking_view(
            watchlist, market_data, quality_score, buy_filter, first_year
        )

        self.assertIn("行情字段缺失", result.iloc[0]["阻断原因"])
        self.assertIn("最新价", result.iloc[0]["阻断原因"])
        self.assertNotEqual(result.iloc[0]["阻断原因"], "无")
        self.assertIn("联网重跑行情接口", result.iloc[0]["下一步"])

    def test_long_term_tracking_keeps_explicit_no_blocker_text(self):
        watchlist = pd.DataFrame(
            [
                {
                    "Code": "688012",
                    "Name": "中微公司",
                    "Role": "长期跟踪个股观察",
                    "Asset Type": "stock",
                    "Theme": "半导体设备",
                    "Thesis": "测试",
                    "Invalidation": "测试",
                }
            ]
        )
        market_data = pd.DataFrame(
            [
                {
                    "Code": "688012",
                    "Name": "中微公司",
                    "Latest": 471.55,
                    "PctChg": 3.25,
                    "Open": 457,
                    "High": 474.07,
                    "Low": 451,
                    "PrevClose": 456.71,
                    "Amount": 5116439346,
                    "QuoteTime": "2026-06-30 09:52:09",
                    "DataSource": "行情接口",
                    "接口状态": "当日行情已刷新",
                    "失败原因": "无",
                    "重试结果": "实时行情与日K补齐可用",
                }
            ]
        )
        quality_score = pd.DataFrame(
            [
                {
                    "Code": "688012",
                    "质量状态": "核心候选",
                    "折算质量分": 8.5,
                    "评分输入状态": "完整",
                    "评分缺失项": "无",
                }
            ]
        )
        buy_filter = pd.DataFrame(
            [
                {
                    "Code": "688012",
                    "买点灯号": "灰",
                    "否决原因": "无",
                    "建议买入区间": "暂不建议买入（买点过滤器未通过）",
                }
            ]
        )
        first_year = pd.DataFrame(
            [
                {
                    "配置项": "AI/半导体高质量个股篮子",
                    "映射代码": "688012",
                    "年度资金缺口": 100000,
                    "年度完成率": 0,
                }
            ]
        )

        result = main.build_long_term_tracking_view(
            watchlist, market_data, quality_score, buy_filter, first_year
        )

        self.assertEqual(result.iloc[0]["阻断原因"], "无")
        self.assertEqual(result.iloc[0]["下一步"], "仅限人工复核，不自动交易")

    def test_buy_filter_uses_first_year_full_asset_weights_not_broker_weight(self):
        watchlist = pd.DataFrame(
            [
                {
                    "Code": "159516",
                    "Name": "半导体设备ETF",
                    "Role": "核心观察",
                    "Target Weight": 0.05,
                    "Current Weight": "",
                    "Share/Chip Signal": "绿",
                    "Premium Signal": "绿",
                    "Leader Signal": "绿",
                    "Next-Day Signal": "绿",
                }
            ]
        )
        market_data = pd.DataFrame(
            [
                {
                    "Code": "159516",
                    "Name": "半导体设备ETF",
                    "Latest": 1.52,
                    "PctChg": 0.6,
                    "Open": 1.50,
                    "High": 1.55,
                    "Low": 1.48,
                    "PrevClose": 1.51,
                    "PrevDayLow": 1.49,
                    "QuoteTime": "2026-06-22 15:00:00",
                    "DataSource": "腾讯行情补齐",
                    "Amount": 100000000,
                    "Avg20Amount": 100000000,
                    "Avg20AmountSource": "东方财富日K",
                    "ETFShareChg": 0,
                    "Premium": 0,
                }
            ]
        )
        positions = pd.DataFrame(
            [
                {"Code": "159516", "Weight": 0.20, "Target Weight": 0.05},
            ]
        )
        quality = pd.DataFrame(
            [
                {"Code": "159516", "折算质量分": 9.0, "质量状态": "核心候选"},
            ]
        )
        emotion = pd.DataFrame([{"指标": "情绪温度", "数值": 1}])
        first_year = pd.DataFrame(
            [
                {
                    "配置项": "半导体设备ETF",
                    "映射代码": "159516",
                    "年度目标占比": 0.30,
                    "当前全资产占比": 0.02,
                    "年度资金缺口": 280000,
                    "年度完成率": 0.067,
                }
            ]
        )

        result = main.build_buy_filter(watchlist, market_data, positions, quality, emotion, first_year)
        row = result.iloc[0]

        self.assertEqual(row["仓位决策口径"], "全资产/第一年配置")
        self.assertAlmostEqual(row["当前仓位"], 0.02)
        self.assertAlmostEqual(row["目标仓位"], 0.30)
        self.assertAlmostEqual(row["剩余额度"], 0.28)
        self.assertAlmostEqual(row["证券账户仓位"], 0.20)
        self.assertIn("爆量突破", str(row["趋势结构"]))
        self.assertIn("0.5%–1%", str(row["单次仓位上限"]))
        self.assertEqual(row["资产角色"], "核心成长ETF")
        self.assertIn("量化验证", str(row["量化验证状态"]))
        self.assertIn("正期望", str(row["正期望检查"]))
        self.assertIn("第一优先", str(row["机会成本检查"]))
        self.assertIn("确认偏误", str(row["认知防错"]))
        self.assertNotIn("全资产仓位已达", str(row["否决原因"]))

    def test_dividend_stock_gets_cashflow_buy_point_rules(self):
        watchlist = pd.DataFrame(
            [
                {
                    "Code": "600900",
                    "Name": "长江电力",
                    "Role": "红利个股观察-A档",
                    "Theme": "红利现金流",
                    "Asset Type": "Stock",
                    "Target Weight": 0.015,
                    "Current Weight": "",
                    "Share/Chip Signal": "绿",
                    "Premium Signal": "绿",
                    "Leader Signal": "绿",
                    "Next-Day Signal": "绿",
                    "Notes": "长期稳定分红",
                }
            ]
        )
        market_data = pd.DataFrame(
            [
                {
                    "Code": "600900",
                    "Name": "长江电力",
                    "Latest": 30.0,
                    "PctChg": 0.2,
                    "Open": 29.8,
                    "High": 30.3,
                    "Low": 29.6,
                    "PrevClose": 29.9,
                    "PrevDayLow": 29.5,
                    "QuoteTime": "2026-06-25 15:00:00",
                    "DataSource": "腾讯行情补齐",
                    "Amount": 100000000,
                    "Avg20Amount": 120000000,
                    "Avg20AmountSource": "东方财富日K",
                    "ETFShareChg": 0,
                    "Premium": 0,
                    "接口状态": "当日行情已刷新",
                    "失败原因": "",
                    "重试结果": "",
                }
            ]
        )
        quality = pd.DataFrame(
            [
                {"Code": "600900", "折算质量分": 8.8, "质量状态": "核心候选", "评分输入状态": "完整", "评分缺失项": "无"},
            ]
        )
        emotion = pd.DataFrame([{"指标": "情绪温度", "数值": 2}])
        first_year = pd.DataFrame(
            [
                {
                    "配置项": "长期稳定分红个股篮子",
                    "映射代码": "600900",
                    "年度目标占比": 0.05,
                    "当前全资产占比": 0,
                    "年度资金缺口": 239150,
                    "年度完成率": 0,
                }
            ]
        )

        result = main.build_buy_filter(watchlist, market_data, pd.DataFrame(), quality, emotion, first_year)
        row = result.iloc[0]

        self.assertEqual(row["资产角色"], "红利现金流个股")
        self.assertEqual(row["红利配置角色"], "A档红利底仓候选")
        self.assertIn("分红", row["分红质量检查"])
        self.assertIn("股息率安全垫", row["红利买点规则"])
        self.assertIn("高息", row["高息陷阱否决"])

    def test_position_risk_prioritizes_triggered_and_overweight(self):
        positions_action = pd.DataFrame(
            [
                {"Code": "000001", "Name": "触发标的", "触发提醒": "是", "动作建议": "减仓复核", "说明": "进入纪律区间"},
                {"Code": "000002", "Name": "普通标的", "触发提醒": "否", "动作建议": "观察", "说明": "未触发"},
            ]
        )
        positions_sheet = pd.DataFrame(
            [
                {"Code": "000001", "Name": "触发标的", "Weight": 0.08, "Target Weight": 0.05, "Market Value": 80000, "Unrealized PnL": -10000},
                {"Code": "000002", "Name": "普通标的", "Weight": 0.03, "Target Weight": 0.05, "Market Value": 30000, "Unrealized PnL": 1000},
            ]
        )

        result = main.build_position_risk_view(positions_action, positions_sheet)

        self.assertEqual(result.iloc[0]["风险级别"], "红")
        self.assertTrue(str(result.iloc[0]["触发原因"]).strip())
        self.assertEqual(result.iloc[0]["Code"], "000001")

    def test_position_risk_does_not_treat_broker_overweight_as_sell_signal(self):
        positions_action = pd.DataFrame(
            [
                {"Code": "000001", "Name": "测试标的", "触发提醒": "否", "动作建议": "观察", "说明": "未触发"},
            ]
        )
        positions_sheet = pd.DataFrame(
            [
                {
                    "Code": "000001",
                    "Name": "测试标的",
                    "Weight": 0.20,
                    "Target Weight": 0.05,
                    "Full Asset Weight": 0.02,
                    "年度全资产目标": 0.30,
                    "Market Value": 20000,
                    "Unrealized PnL": 1000,
                },
            ]
        )

        result = main.build_position_risk_view(positions_action, positions_sheet)

        self.assertEqual(result.iloc[0]["风险级别"], "灰")
        self.assertEqual(result.iloc[0]["全资产当前仓位"], 0.02)
        self.assertEqual(result.iloc[0]["全资产目标仓位"], 0.30)
        self.assertEqual(result.iloc[0]["证券账户仓位"], 0.20)
        self.assertIn("未触发", result.iloc[0]["触发原因"])

    def test_positions_sheet_marks_broker_overweight_as_reference_only(self):
        positions = pd.DataFrame(
            [
                {
                    "Code": "000001",
                    "Name": "测试标的",
                    "Shares": 1000,
                    "Latest": 20,
                    "Cost": 18,
                    "Market Value": 20000,
                    "Weight": 0.20,
                    "Target Weight": 0.05,
                    "V2.8.5 Action": "持有/观察",
                },
            ]
        )
        market_data = pd.DataFrame(
            [
                {"Code": "000001", "Latest": 20, "DataSource": "券商截图"},
            ]
        )
        account_meta = {"account_total": 100000, "total_assets": 1000000}

        sheet, alerts = main.build_positions_sheet(positions, market_data, account_meta)

        self.assertEqual(sheet.iloc[0]["V2.8.5 Action"], "持有/观察")
        self.assertIn("仅作集中度参考", sheet.iloc[0]["证券账户集中度提示"])
        self.assertTrue(alerts.empty)

    def test_execution_plan_sell_review_uses_full_asset_first_year_basis(self):
        buy_filter = pd.DataFrame(
            columns=["买点灯号", "一票否决", "Name", "Code", "Role", "建议"]
        )
        positions_action = pd.DataFrame(
            [
                {"Code": "000001", "Name": "测试标的", "触发提醒": "否", "说明": "未触发"},
            ]
        )
        positions_sheet = pd.DataFrame(
            [
                {
                    "Code": "000001",
                    "Name": "测试标的",
                    "Weight": 0.20,
                    "Target Weight": 0.05,
                    "Full Asset Weight": 0.02,
                    "年度全资产目标": 0.30,
                },
            ]
        )
        first_year = pd.DataFrame(
            [
                {
                    "配置项": "测试配置",
                    "映射代码": "000001",
                    "年度目标占比": 0.30,
                    "年度资金缺口": 280000,
                    "年度完成率": 0.067,
                }
            ]
        )

        result = main.build_execution_plan(buy_filter, positions_action, positions_sheet, first_year)

        self.assertEqual(result.iloc[0]["动作类型"], "观察")
        self.assertNotIn("风控/减仓复核", set(result["动作类型"]))

    def test_front_sheet_order_is_fixed(self):
        sheets = {
            "Checks": pd.DataFrame(),
            "Investment_Profile": pd.DataFrame(),
            "Data_Sources": pd.DataFrame(),
            "03_买入候选": pd.DataFrame(),
            "04_阶段推进": pd.DataFrame(),
            "买点计划": pd.DataFrame(),
            "01_今日决策": pd.DataFrame(),
            "07_组合总览": pd.DataFrame(),
            "05_持仓风险": pd.DataFrame(),
            "02_今日动作": pd.DataFrame(),
            "06_年度配置": pd.DataFrame(),
            "Framework_Rules": pd.DataFrame(),
            "长期跟踪个股": pd.DataFrame(),
            "Buy_Filter": pd.DataFrame(),
        }

        names = main.build_output_sheet_order(sheets)

        self.assertEqual(names[:len(main.FRONT_SHEET_ORDER)], main.FRONT_SHEET_ORDER)
        self.assertLess(names.index("Investment_Profile"), names.index("Framework_Rules"))
        self.assertLess(names.index("Data_Sources"), names.index("Framework_Rules"))
        self.assertGreater(names.index("Buy_Filter"), names.index("07_组合总览"))

    def test_long_term_tracking_sheet_summarizes_tracking_stocks(self):
        watchlist = pd.DataFrame(
            [
                {
                    "Code": "688012",
                    "Name": "中微公司",
                    "Role": "半导体设备个股观察",
                    "Theme": "半导体设备",
                    "Asset Type": "Stock",
                    "Action": "长期跟踪",
                    "Notes": "不自动买入",
                    "Thesis": "国产设备",
                    "Invalidation": "订单下修",
                }
            ]
        )
        market_data = pd.DataFrame(
            [
                {
                    "Code": "688012",
                    "Name": "中微公司",
                    "Latest": 200,
                    "PctChg": 1.2,
                    "Open": 198,
                    "High": 202,
                    "Low": 196,
                    "PrevClose": 197,
                    "Amount": 100000000,
                    "QuoteTime": "2026-06-22 15:00:00",
                    "DataSource": "腾讯行情补齐",
                }
            ]
        )
        quality = pd.DataFrame([{"Code": "688012", "质量状态": "数据不足，需人工评分", "折算质量分": None}])
        buy_filter = pd.DataFrame(
            [
                {
                    "Code": "688012",
                    "买点灯号": "红",
                    "阻断原因": "缺少完整质量评分或证据，禁止新增",
                    "建议买入区间": "暂不建议买入（缺少完整质量评分或证据，禁止新增）",
                    "建议": "质量评分或证据不完整，先补评分，禁止新增",
                }
            ]
        )
        first_year = pd.DataFrame(
            [
                {
                    "配置项": "AI/半导体高质量个股篮子",
                    "映射代码": "688012",
                    "年度资金缺口": 143490,
                    "年度完成率": 0,
                }
            ]
        )

        result = main.build_long_term_tracking_view(watchlist, market_data, quality, buy_filter, first_year)

        self.assertEqual(len(result), 1)
        self.assertEqual(result.iloc[0]["Code"], "688012")
        self.assertIn("补完整质量评分", result.iloc[0]["下一步"])
        self.assertEqual(result.iloc[0]["年度配置项"], "AI/半导体高质量个股篮子")

    def test_profile_and_data_source_tables_are_loaded(self):
        profile = main.build_investment_profile(
            {
                "cash_buffer_min": {"Value": "300000"},
                "cash_buffer_target": {"Value": "500000"},
            }
        )
        sources = main.build_data_sources()

        self.assertIn("buy_candidates_view", set(profile["Key"]))
        self.assertIn("当前生效值", profile.columns)
        self.assertIn("每日行情", set(sources["Module"]))
        self.assertIn("FailureBehavior", sources.columns)

    def test_long_term_stock_tracking_pool_is_in_watchlist_and_allocation(self):
        expected_codes = {
            "688012", "002371", "688072", "688120", "300308",
            "300502", "300394", "002463", "688008", "601138",
        }
        watchlist = main.load_watchlist()
        allocation = main.load_first_year_allocation()
        basket = allocation[allocation["AllocationKey"] == "quality_stock_basket"].iloc[0]
        basket_codes = set(str(basket["Codes"]).split("|"))

        self.assertTrue(expected_codes.issubset(set(watchlist["Code"])))
        self.assertEqual(expected_codes, basket_codes)
        tracking_rows = watchlist[watchlist["Code"].isin(expected_codes)]
        self.assertTrue(tracking_rows["Manual Quality Score"].notna().all())
        self.assertTrue(tracking_rows["Quality Evidence"].astype(str).str.strip().ne("").all())
        self.assertTrue((pd.to_numeric(tracking_rows["Manual Quality Score"]) >= 8).all())

    def test_cleared_etfs_removed_and_dividend_pool_added(self):
        watchlist = main.load_watchlist()
        allocation = main.load_first_year_allocation()
        cleared_codes = {"159" + "566", "159" + "819"}
        dividend_codes = {"600900", "600941", "601398", "601939", "601088", "000333"}
        allocation_codes = "|".join(allocation["Codes"].astype(str).tolist())

        self.assertFalse(cleared_codes.intersection(set(watchlist["Code"].astype(str))))
        self.assertNotIn("159" + "566", allocation_codes)
        self.assertNotIn("159" + "819", allocation_codes)
        self.assertTrue(dividend_codes.issubset(set(watchlist["Code"].astype(str))))
        dividend_row = allocation[allocation["AllocationKey"] == "dividend_cashflow_stocks"].iloc[0]
        self.assertEqual(set(str(dividend_row["Codes"]).split("|")), dividend_codes)
        self.assertAlmostEqual(float(dividend_row["TargetWeight"]), 0.05)

    def test_byd_is_2030_strategy_observation_not_legacy_basket(self):
        watchlist = main.load_watchlist()
        allocation = main.load_first_year_allocation()

        byd = watchlist[watchlist["Code"] == "002594"].iloc[0]
        self.assertEqual(byd["Role"], "2030战略观察仓")
        self.assertIn("2030", str(byd["Notes"]))
        self.assertGreaterEqual(float(byd["Manual Quality Score"]), 7.0)

        byd_bucket = allocation[allocation["AllocationKey"] == "byd_2030_strategy"].iloc[0]
        self.assertEqual(byd_bucket["Codes"], "002594")
        self.assertAlmostEqual(float(byd_bucket["TargetWeight"]), 0.03)
        legacy_bucket = allocation[allocation["AllocationKey"] == "new_energy_legacy"].iloc[0]
        self.assertNotIn("002594", str(legacy_bucket["Codes"]).split("|"))

    def test_buy_point_plan_loads_key_etf_ranges(self):
        plan = main.load_buy_point_plan()

        a500 = plan[plan["Code"] == "159338"].iloc[0]
        dividend = plan[plan["Code"] == "512890"].iloc[0]
        convertible = plan[plan["Code"] == "511180"].iloc[0]

        self.assertEqual(a500["轻仓观察区间"], "1.305–1.315")
        self.assertEqual(a500["标准买点区间"], "1.285–1.300")
        self.assertIn("4–5万", a500["下一笔合理金额"])
        self.assertEqual(dividend["标准买点区间"], "1.075–1.085")
        self.assertEqual(convertible["轻仓观察区间"], "12.45–12.55")

    def test_build_buy_point_plan_view_adds_current_position_context(self):
        plan = pd.DataFrame(
            [
                {
                    "Code": "159338",
                    "Name": "中证A500ETF",
                    "资产角色": "低波权益底仓",
                    "当前价参考": "1.323",
                    "轻仓观察区间": "1.305–1.315",
                    "标准买点区间": "1.285–1.300",
                    "强买点区间": "1.260–1.280",
                    "不买/暂停区": "1.335以上不追",
                    "下一笔合理金额": "4–5万，分2–3次",
                    "成立条件": "指数不破位",
                    "风险暂停条件": "权重集体破位",
                }
            ]
        )
        positions_sheet = pd.DataFrame(
            [
                {"Code": "159338", "Market Value": 86524.20, "Latest": 1.323, "Cost": 1.295, "P/L%": 0.0218}
            ]
        )
        buy_filter = pd.DataFrame(
            [
                {"Code": "159338", "买点灯号": "红"}
            ]
        )

        result = main.build_buy_point_plan_view(plan, positions_sheet, buy_filter)

        self.assertEqual(result.iloc[0]["买点灯号"], "红")
        self.assertEqual(result.iloc[0]["当前价"], 1.323)
        self.assertEqual(result.iloc[0]["持仓市值"], 86524.20)
        self.assertEqual(result.iloc[0]["当前盈亏"], 0.0218)
        self.assertEqual(result.iloc[0]["下一笔合理金额"], "4–5万，分2–3次")

    def test_action_plan_adds_position_limits(self):
        execution_plan = pd.DataFrame(
            [
                {"优先级": 1, "Code": "000001", "动作类型": "风控/减仓复核", "标的": "测试标的(000001)", "依据": "超配", "状态": "优先处理"},
            ]
        )
        buy_filter = pd.DataFrame(
            [
                {
                    "Code": "000001",
                    "仓位决策口径": "全资产/第一年配置",
                    "当前仓位": 0.08,
                    "目标仓位": 0.05,
                    "剩余额度": 0.0,
                    "全资产当前仓位": 0.08,
                    "全资产目标仓位": 0.05,
                    "全资产剩余额度": 0.0,
                    "证券账户仓位": 0.20,
                    "证券账户目标仓位": 0.10,
                },
            ]
        )
        positions_sheet = pd.DataFrame(
            [
                {"Code": "000001", "Weight": 0.08, "Target Weight": 0.05},
            ]
        )

        result = main.build_action_plan_view(execution_plan, buy_filter, positions_sheet)

        self.assertEqual(result.iloc[0]["全资产当前仓位"], 0.08)
        self.assertEqual(result.iloc[0]["全资产目标仓位"], 0.05)
        self.assertEqual(result.iloc[0]["全资产剩余额度"], 0.0)
        self.assertEqual(result.iloc[0]["证券账户仓位"], 0.20)
        self.assertNotIn("当前仓位", result.columns)

    def test_action_dashboard_contains_status_actions_blockers_and_data(self):
        decision_center = pd.DataFrame(
            [
                {"层级": "市场权限", "状态": "暂停标准新增", "证据": "双锚观察", "动作": "继续观察"},
            ]
        )
        action_plan = pd.DataFrame(
            [
                {"优先级": 1, "动作类型": "风控/减仓复核", "标的": "测试标的", "依据": "超配", "状态": "优先处理"},
            ]
        )
        buy_candidates = pd.DataFrame(
            [
                {"买点灯号": "红", "Name": "红标的", "阻断原因": "质量评分缺失"},
            ]
        )
        position_risk = pd.DataFrame([{"风险级别": "红", "Name": "测试标的"}])
        dashboard = pd.DataFrame(
            [
                {"项目": "现金安全垫状态", "内容": "达标"},
                {"项目": "行情数据时间", "内容": "未取到当日接口时间"},
            ]
        )

        result = main.build_action_dashboard_view(
            decision_center,
            action_plan,
            buy_candidates,
            position_risk,
            dashboard,
        )

        self.assertEqual(set(result["区域"]), {"核心状态", "今日动作", "主要阻断", "数据状态"})
        core_items = set(result.loc[result["区域"] == "核心状态", "项目"])
        self.assertEqual(core_items, {"市场权限", "阶段推进", "减仓复核", "现金安全垫"})
        data_status = result.loc[result["区域"] == "数据状态", "状态"].iloc[0]
        self.assertIn("未刷新", data_status)

    def test_write_excel_places_front_sheets_first(self):
        dashboard_view = pd.DataFrame(
            [
                {"区域": "核心状态", "项目": "市场权限", "状态": "暂停标准新增", "证据": "双锚观察", "动作": "继续观察"},
                {"区域": "数据状态", "项目": "行情时间", "状态": "行情未刷新", "证据": "未取到当日接口时间", "动作": "禁止新增"},
            ]
        )
        sheets = {
            "Checks": pd.DataFrame([{"检查项": "测试", "状态": "OK"}]),
            "03_买入候选": pd.DataFrame(
                [
                    {
                        "买点灯号": "红",
                        "Code": "000001",
                        "Name": "测试标的",
                        "建议买入区间": "暂不建议买入（测试）",
                        "阻断原因": "测试",
                    }
                ]
            ),
            "01_今日决策": dashboard_view,
            "04_阶段推进": pd.DataFrame([{"Code": "000001", "Name": "测试标的", "阶段推进结论": "等待"}]),
            "买点计划": pd.DataFrame([{"Code": "000001", "Name": "测试标的"}]),
            "07_组合总览": pd.DataFrame([{"项目": "测试", "内容": "测试"}]),
            "05_持仓风险": pd.DataFrame([{"风险级别": "红", "Code": "000001", "触发原因": "测试"}]),
            "02_今日动作": pd.DataFrame([{"优先级": 1, "动作类型": "观察", "标的": "测试"}]),
            "06_年度配置": pd.DataFrame([{"配置项": "测试", "年度目标占比": 0.1}]),
        }
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "dashboard.xlsx"

            main.write_excel(path, sheets)

            workbook = load_workbook(path, read_only=False)
            self.assertEqual(workbook.sheetnames[:len(main.FRONT_SHEET_ORDER)], main.FRONT_SHEET_ORDER)
            self.assertIn("A1:H1", {str(item) for item in workbook["01_今日决策"].merged_cells.ranges})
            candidate_headers = [cell.value for cell in workbook["03_买入候选"][1]]
            self.assertEqual(candidate_headers[3], "建议买入区间")
            self.assertGreaterEqual(workbook["03_买入候选"].column_dimensions["D"].width, 24)

    def test_framework_rules_document_buy_range_as_review_only(self):
        rules = main.build_framework_rules({})
        matching = rules[rules["规则"] == "建议买入区间"]

        self.assertEqual(len(matching), 1)
        definition = str(matching.iloc[0]["阈值/定义"])
        self.assertIn("复核", definition)
        self.assertIn("不自动", definition)

    def test_framework_rules_include_volume_absorption_module(self):
        rules = main.build_framework_rules({})
        text = "\n".join(rules.astype(str).agg(" ".join, axis=1).tolist())

        self.assertIn("爆量/量质吸", text)
        self.assertIn("爆量突破→缩量回踩", text)
        self.assertIn("强确认核心ETF1%–2%", text)

    def test_framework_rules_include_cognitive_guardrails(self):
        rules = main.build_framework_rules({})
        text = "\n".join(rules.astype(str).agg(" ".join, axis=1).tolist())

        self.assertIn("V2.8.5-QM", text)
        self.assertIn("认知防错八问", text)
        self.assertIn("机会成本", text)
        self.assertIn("两项回答不清", text)

    def test_framework_rules_include_quant_validation_module(self):
        rules = main.build_framework_rules({})
        text = "\n".join(rules.astype(str).agg(" ".join, axis=1).tolist())

        self.assertIn("V2.8.5-Q", text)
        self.assertIn("正期望", text)
        self.assertIn("资产角色", text)
        self.assertIn("高频", text)

    def test_framework_rules_include_dividend_cashflow_layer(self):
        rules = main.build_framework_rules({})
        text = "\n".join(rules.astype(str).agg(" ".join, axis=1).tolist())

        self.assertIn("红利现金流个股", text)
        self.assertIn("高息陷阱", text)
        self.assertIn("分红连续性", text)

    def test_portfolio_overview_removes_per_position_alert_rows(self):
        dashboard = pd.DataFrame(
            [
                {"项目": "全资产总额", "内容": 1000000},
                {"项目": "提醒-000001", "内容": "当前仓位超目标"},
                {"项目": "现金安全垫状态", "内容": "达标"},
            ]
        )

        result = main.build_portfolio_overview(dashboard)

        self.assertEqual(result["项目"].tolist(), ["全资产总额", "现金安全垫状态"])


if __name__ == "__main__":
    unittest.main()
