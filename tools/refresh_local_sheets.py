#!/usr/bin/env python3
"""不联网刷新工作簿中的本地决策页，用于规则/版式迭代。"""
from __future__ import annotations

import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

import main


def replace_sheet(wb, name: str, frame: pd.DataFrame, index: int) -> None:
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name, index)
    for row in dataframe_to_rows(frame, index=False, header=True):
        ws.append(row)
    main.style_excel_worksheet(ws, name)


def run(source: Path) -> Path:
    sheets = pd.read_excel(
        source,
        sheet_name=["03_买入候选", "04_持仓风险", "Positions", "05_年度配置"],
        dtype={"Code": str},
    )
    execution = main.build_execution_plan(
        sheets["03_买入候选"],
        sheets["04_持仓风险"],
        sheets["Positions"],
        sheets["05_年度配置"],
    )
    execution = main.build_action_plan_view(execution, sheets["03_买入候选"], sheets["Positions"])
    wb = load_workbook(source)
    replace_sheet(wb, "02_今日动作", execution, 1)
    output = source.with_name(f"{main.FRAMEWORK_VERSION}_每日行情输出_{datetime.now():%Y%m%d_%H%M%S}.xlsx")
    wb.save(output)
    return output


if __name__ == "__main__":
    if len(sys.argv) != 2:
        raise SystemExit("Usage: refresh_local_sheets.py <daily-output.xlsx>")
    print(run(Path(sys.argv[1])))
