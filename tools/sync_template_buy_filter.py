#!/usr/bin/env python3
from __future__ import annotations

from copy import copy
from pathlib import Path
import sys

from openpyxl import load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.workbook.properties import CalcProperties

import main


TEMPLATE = ROOT / "V2.8.4_ETF_A股纪律仪表盘模板.xlsx"
TEMP_TEMPLATE = ROOT / "V2.8.4_ETF_A股纪律仪表盘模板.tmp.xlsx"


def buy_filter_formulas(row: int) -> list[str]:
    r = row
    index_cond = f'OR(RIGHT(B{r},3)=".SH",RIGHT(B{r},3)=".SZ",RIGHT(B{r},4)=".CSI")'
    notes_lookup = f'IFERROR(VLOOKUP(B{r},Watchlist!$A$2:$L$51,12,FALSE),"")'
    legacy_cond = f'ISNUMBER(SEARCH("遗留仓",D{r}))'
    hold_cond = f'ISNUMBER(SEARCH("持有不加",{notes_lookup}))'
    missing_weight_cond = f'AND(F{r}<>"",E{r}="")'
    position_full_cond = f'AND(F{r}<>"",E{r}<>"",E{r}>=F{r})'
    price_veto_cond = (
        f'AND(Market_Data!E{r}<>"",Market_Data!H{r}<>"",K{r}<>"",'
        f'Market_Data!E{r}>Market_Data!H{r},K{r}<=0.2)'
    )
    hard_cond = (
        f'OR({index_cond},{legacy_cond},{hold_cond},{missing_weight_cond},'
        f'{position_full_cond},{price_veto_cond})'
    )

    return [
        f'=IF(B{r}="","",IF({hard_cond},"红",IF(S{r}>=3,"绿",IF(S{r}>=2,"黄","灰"))))',
        f'=IF(Market_Data!A{r}="","",Market_Data!A{r})',
        f'=IF(B{r}="","",Market_Data!B{r})',
        f'=IF(B{r}="","",IFERROR(VLOOKUP(B{r},Watchlist!$A$2:$L$51,3,FALSE),""))',
        f'=IF(B{r}="","",IF({index_cond},"",IFERROR(VLOOKUP(B{r},Positions!$A$5:$L$104,9,FALSE),0)))',
        f'=IF(B{r}="","",IFERROR(VLOOKUP(B{r},Watchlist!$A$2:$L$51,5,FALSE),""))',
        f'=IF(OR(B{r}="",E{r}="",F{r}=""),"",MAX(F{r}-E{r},0))',
        (
            f'=IF(B{r}="","",IF({index_cond},"指数锚点（不可交易）",'
            f'IF(F{r}="","未设置目标仓位",IF(E{r}="","当前仓位缺失，需人工确认",'
            f'IF(E{r}>F{r},"超目标 "&TEXT(E{r}-F{r},"0.00%"),'
            f'IF(E{r}=F{r},"已达目标仓位","低于目标，剩余 "&TEXT(G{r},"0.00%")))))))'
        ),
        f'=IF(OR(B{r}="",Market_Data!C{r}=""),"",Market_Data!C{r})',
        f'=IF(OR(B{r}="",Market_Data!D{r}=""),"",Market_Data!D{r})',
        (
            f'=IF(OR(B{r}="",I{r}="",Market_Data!F{r}="",Market_Data!G{r}="",'
            f'Market_Data!F{r}=Market_Data!G{r}),"",(I{r}-Market_Data!G{r})/(Market_Data!F{r}-Market_Data!G{r}))'
        ),
        f'=IF(OR(B{r}="",Market_Data!I{r}="",Market_Data!J{r}="",Market_Data!J{r}=0),"",Market_Data!I{r}/Market_Data!J{r})',
        (
            f'=IF(B{r}="","",IF(K{r}="","缺少日内数据",'
            f'IF(K{r}>=0.85,"接近日内高位",IF(K{r}<=0.25,"接近日内低位","日内中位震荡"))))'
        ),
        (
            f'=IF(B{r}="","",IF(J{r}="","缺少涨跌幅",IF(J{r}>0,'
            f'IF(OR(L{r}="",L{r}<=1.8),"温和上涨","放量上涨，防追高"),'
            f'IF(AND(J{r}<0,L{r}>=1.2),"放量下跌","弱势震荡"))))'
        ),
        f'=IF(B{r}="","",IF(Market_Data!K{r}="","量能替代：缺少份额数据",Market_Data!K{r}))',
        f'=IF(B{r}="","",IF(Market_Data!L{r}="","折溢价待填",Market_Data!L{r}))',
        f'=IF(B{r}="","",IF(Market_Data!M{r}="","待填",Market_Data!M{r}))',
        f'=IF(B{r}="","","待次日确认")',
        (
            f'=IF(B{r}="","",IF(OR(I{r}="",J{r}="",K{r}="",L{r}=""),0,'
            f'--AND(K{r}>=0.35,K{r}<=0.75)+--AND(J{r}>=-1%,J{r}<=3%)+'
            f'--AND(L{r}>=0.7,L{r}<=1.8)+--AND(J{r}>=0,K{r}<0.85)))'
        ),
        (
            f'=IF(B{r}="","",IF(OR(I{r}="",J{r}="",K{r}="",L{r}=""),"待补行情",'
            f'IF(AND(K{r}>=0.35,K{r}<=0.75),"日内位置适中；","")&'
            f'IF(AND(J{r}>=-1%,J{r}<=3%),"涨跌幅适中；","")&'
            f'IF(AND(L{r}>=0.7,L{r}<=1.8),"量能正常；","")&'
            f'IF(AND(J{r}>=0,K{r}<0.85),"未贴近日内高点","")))'
        ),
        (
            f'=IF(B{r}="","",IF(OR(I{r}="",J{r}="",K{r}="",L{r}=""),"缺少行情字段",'
            f'IF(AND(K{r}>=0.35,K{r}<=0.75),"","日内位置偏高/偏低；")&'
            f'IF(AND(J{r}>=-1%,J{r}<=3%),"","涨跌幅不适合追买；")&'
            f'IF(AND(L{r}>=0.7,L{r}<=1.8),"","量能不足或过热；")&'
            f'IF(AND(J{r}>=0,K{r}<0.85),"","可能追高或弱势")))'
        ),
        f'=IF(B{r}="","",IF({hard_cond},"是","否"))',
        (
            f'=IF(B{r}="","",IF({index_cond},"指数仅作市场锚点，不执行买入",'
            f'IF({legacy_cond},"遗留仓禁止新买，按减仓/退出纪律处理",'
            f'IF({hold_cond},"持有不加，禁止新增仓位",'
            f'IF({missing_weight_cond},"当前仓位缺失，禁止自动给出买入建议",'
            f'IF({position_full_cond},IF(E{r}>F{r},"当前仓位超目标 "&TEXT(E{r}-F{r},"0.00%"),"当前仓位已达目标"),'
            f'IF({price_veto_cond},"高开后接近日内低位","无")))))))'
        ),
        (
            f'=IF(B{r}="","",IF(V{r}="是",IF({index_cond},"指数锚点，仅观察，不执行买入",'
            f'IF({legacy_cond},"遗留仓禁止新买；"&H{r},IF({hold_cond},"持有不加，禁止新增；"&H{r},'
            f'IF({missing_weight_cond},"仓位数据缺失，先核对持仓，禁止买入",'
            f'IF({position_full_cond},H{r}&"，禁止加仓","不买/观察"))))),'
            f'IF(S{r}>=3,IF(G{r}<>"","可进入买点复核；累计加仓不超过剩余额度 "&TEXT(G{r},"0.00%"),'
            f'"可进入买点复核；需人工确认仓位上限"),IF(S{r}>=2,'
            f'IF(G{r}<>"","可观察（未达完整过滤器；剩余额度 "&TEXT(G{r},"0.00%")&"）",'
            f'"可观察（未达完整过滤器）"),"不买/观察（买点不合格；仓位额度不等于买入信号）"))))'
        ),
    ]


def copy_row_style(ws, source_row: int, target_row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        if source.has_style:
            target._style = copy(source._style)
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.number_format = source.number_format
        target.protection = copy(source.protection)


def sync_buy_filter(wb) -> None:
    ws = wb["Buy_Filter"]
    headers = main.BUY_FILTER_COLUMNS

    if "BuyFilterTable" in ws.tables:
        del ws.tables["BuyFilterTable"]

    for row in ws.iter_rows(min_row=1, max_row=51, min_col=1, max_col=26):
        for cell in row:
            cell.value = None

    dark_blue = PatternFill("solid", fgColor="1F4E78")
    white_bold = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    green = PatternFill("solid", fgColor="C6EFCE")
    yellow = PatternFill("solid", fgColor="FFEB9C")
    red = PatternFill("solid", fgColor="FFC7CE")
    grey = PatternFill("solid", fgColor="E7E6E6")

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(1, col, header)
        cell.fill = dark_blue
        cell.font = white_bold
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in range(2, 52):
        formulas = buy_filter_formulas(row)
        if len(formulas) != len(headers):
            raise RuntimeError(f"Formula/header mismatch at row {row}")
        for col, formula in enumerate(formulas, start=1):
            cell = ws.cell(row, col, formula)
            cell.font = Font(color="000000")
            cell.alignment = Alignment(vertical="center", wrap_text=col in {8, 13, 14, 15, 16, 17, 18, 20, 21, 23, 24})
            cell.border = border
        ws.row_dimensions[row].height = 38

    widths = {
        "买点灯号": 10, "Code": 12, "Name": 20, "Role": 14, "当前仓位": 11, "目标仓位": 11,
        "剩余额度": 11, "仓位状态": 24, "Latest": 11, "PctChg": 10, "日内位置": 11, "量能倍数": 11,
        "分时结构": 16, "量价关系": 18, "份额变动": 20, "折溢价": 18, "龙头成分": 18, "次日验证": 16,
        "通过项": 9, "通过明细": 36, "未通过项": 36, "一票否决": 11, "否决原因": 34, "建议": 44,
    }
    for col, header in enumerate(headers, start=1):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = widths[header]

    for row in range(2, 52):
        for col in (5, 6, 7):
            ws.cell(row, col).number_format = "0.00%"
        ws.cell(row, 9).number_format = "0.000"
        ws.cell(row, 10).number_format = "0.00%"
        for col in (11, 12):
            ws.cell(row, col).number_format = "0.00"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:X51"
    ws.sheet_view.showGridLines = False
    ws.conditional_formatting._cf_rules.clear()
    for text, fill in (("绿", green), ("黄", yellow), ("红", red), ("灰", grey)):
        ws.conditional_formatting.add(
            "A2:A51",
            FormulaRule(formula=[f'A2="{text}"'], fill=fill),
        )
    ws.conditional_formatting.add("V2:V51", FormulaRule(formula=['V2="是"'], fill=red))
    ws.conditional_formatting.add("V2:V51", FormulaRule(formula=['V2="否"'], fill=green))

    table = Table(displayName="BuyFilterTable", ref="A1:X51")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def sync_dependents(wb) -> None:
    watchlist = wb["Watchlist"]
    for row in range(2, 52):
        watchlist.cell(row, 9).value = f'=IFERROR(INDEX(Buy_Filter!$A$2:$A$51,MATCH(A{row},Buy_Filter!$B$2:$B$51,0)),"")'
        watchlist.cell(row, 10).value = f'=IFERROR(INDEX(Buy_Filter!$V$2:$V$51,MATCH(A{row},Buy_Filter!$B$2:$B$51,0)),"")'
        watchlist.cell(row, 11).value = f'=IFERROR(INDEX(Buy_Filter!$X$2:$X$51,MATCH(A{row},Buy_Filter!$B$2:$B$51,0)),"")'

    dashboard = wb["Dashboard"]
    dashboard["C12"] = "只考虑仓位未满、买点绿灯且无一票否决的标的；额度不等于买入信号。"
    dashboard["A16"] = "买点过滤器摘要（自动读取 Buy_Filter 前15行，含仓位约束）"
    summary_headers = ["Code", "Name", "Role", "当前仓位", "目标仓位", "仓位状态", "买点灯号", "通过项", "一票否决", "建议"]
    for col, value in enumerate(summary_headers, start=1):
        dashboard.cell(17, col).value = value
        if col > 7:
            dashboard.cell(17, col)._style = copy(dashboard.cell(17, 7)._style)
    source_cols = ["B", "C", "D", "E", "F", "H", "A", "S", "V", "X"]
    for dashboard_row, source_row in enumerate(range(2, 17), start=18):
        for col, source_col in enumerate(source_cols, start=1):
            if source_col in {"E", "F"}:
                formula = f'=IFERROR(IF(Buy_Filter!{source_col}{source_row}="","",Buy_Filter!{source_col}{source_row}),"")'
            else:
                formula = f'=IFERROR(Buy_Filter!{source_col}{source_row},"")'
            dashboard.cell(dashboard_row, col).value = formula
            if col > 7:
                dashboard.cell(dashboard_row, col)._style = copy(dashboard.cell(dashboard_row, 7)._style)
        dashboard.cell(dashboard_row, 4).number_format = "0.00%"
        dashboard.cell(dashboard_row, 5).number_format = "0.00%"
        dashboard.cell(dashboard_row, 10).alignment = Alignment(vertical="center", wrap_text=True)
        dashboard.row_dimensions[dashboard_row].height = 32
    for col, width in {"D": 12, "E": 12, "F": 24, "G": 11, "H": 9, "I": 11, "J": 44}.items():
        dashboard.column_dimensions[col].width = width

    rules = wb["Rules_Lists"]
    rules["B4"] = "高开低走接近日低、放量下跌、指数锚点、已达/超目标仓位、遗留仓、持有不加等均一票否决。"
    additions = [
        ("仓位硬约束", "已达或超过目标仓位时，买点条件再好也禁止加仓。"),
        ("指数锚点", "指数只用于判断市场环境，不作为交易标的。"),
        ("遗留仓/持有不加", "遗留仓和明确标注持有不加的标的禁止新增仓位。"),
        ("额度不等于信号", "剩余额度仅是上限；必须同时满足买点绿灯且无一票否决。"),
    ]
    for row, (rule, description) in enumerate(additions, start=9):
        copy_row_style(rules, 8, row, 9)
        rules.cell(row, 1).value = rule
        rules.cell(row, 2).value = description
        rules.cell(row, 2).alignment = Alignment(vertical="center", wrap_text=True)
        rules.row_dimensions[row].height = 30


def validate(path: Path) -> None:
    wb = load_workbook(path, data_only=False, keep_links=True)
    buy_filter = wb["Buy_Filter"]
    headers = [buy_filter.cell(1, col).value for col in range(1, 25)]
    if headers != main.BUY_FILTER_COLUMNS:
        raise RuntimeError("Template Buy_Filter headers do not match main.py")
    if buy_filter.tables["BuyFilterTable"].ref != "A1:X51":
        raise RuntimeError("BuyFilterTable range was not updated")
    stale_refs: list[str] = []
    formula_errors: list[str] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if not isinstance(value, str) or not value.startswith("="):
                    continue
                if "Buy_Filter!$A$2:$Q$51" in value:
                    stale_refs.append(f"{ws.title}!{cell.coordinate}")
                if value.count("(") != value.count(")"):
                    formula_errors.append(f"{ws.title}!{cell.coordinate}")
    if stale_refs:
        raise RuntimeError(f"Stale Buy_Filter references: {stale_refs[:10]}")
    if formula_errors:
        raise RuntimeError(f"Unbalanced formulas: {formula_errors[:10]}")


def main_sync() -> None:
    wb = load_workbook(TEMPLATE, data_only=False, keep_links=True)
    sync_buy_filter(wb)
    sync_dependents(wb)
    wb.calculation = CalcProperties(calcMode="auto", fullCalcOnLoad=True, forceFullCalc=True)
    wb.save(TEMP_TEMPLATE)
    validate(TEMP_TEMPLATE)
    TEMP_TEMPLATE.replace(TEMPLATE)
    print(TEMPLATE)


if __name__ == "__main__":
    main_sync()
