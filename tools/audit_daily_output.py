#!/usr/bin/env python3
from __future__ import annotations

import csv
import json
import re
import sys
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
TARGET_SHEETS = [
    "01_今日决策",
    "02_今日动作",
    "03_买入候选",
    "04_阶段推进",
    "买点计划",
    "05_持仓风险",
    "06_年度配置",
    "07_组合总览",
    "Double_Anchor",
    "Emotion",
    "Quality_Score",
    "Exposure",
    "Market_Data",
    "Buy_Filter",
    "Positions",
    "Broker_Snapshot",
    "Watchlist",
    "长期跟踪个股",
    "Investment_Profile",
    "Data_Sources",
    "Buy_Point_Plan_Source",
    "Framework_Rules",
    "Checks",
    "使用说明",
]
FRONT_SHEETS = TARGET_SHEETS[:8]
MISSING_TEXT_VALUES = {"", "-", "--", "—", "nan", "none", "null", "n/a", "na"}
MARKET_QUOTE_FIELDS = [
    ("Latest", "最新价", True),
    ("PctChg", "涨跌幅", False),
    ("Open", "今开", True),
    ("High", "最高", True),
    ("Low", "最低", True),
    ("PrevClose", "昨收", True),
    ("Amount", "成交额", False),
]


def normalize_code(value: Any) -> str:
    text = str(value or "").strip()
    if text.endswith(".0"):
        text = text[:-2]
    upper = text.upper()
    if upper.endswith((".SH", ".SZ", ".CSI")):
        return upper
    return text.zfill(6) if text.isdigit() else text


def csv_codes(path: Path) -> list[str]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        lines = [line for line in handle if not line.startswith("#")]
    return [normalize_code(row["Code"]) for row in csv.DictReader(lines)]


def csv_row_count(path: Path) -> int:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        return sum(1 for _ in csv.DictReader(handle))


def long_term_tracking_count(path: Path) -> int:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        rows = csv.DictReader(handle)
        count = 0
        for row in rows:
            text = " ".join(str(row.get(column, "") or "") for column in ["Role", "Action", "Notes", "Thesis"])
            if str(row.get("Asset Type", "")).strip().lower() == "stock" and ("长期跟踪" in text or "个股观察" in text):
                count += 1
        return count


def headers(ws) -> dict[str, int]:
    return {
        str(cell.value).strip(): index
        for index, cell in enumerate(ws[1], start=1)
        if cell.value is not None
    }


def blank_rows(ws, required: list[str]) -> list[dict[str, Any]]:
    mapping = headers(ws)
    issues: list[dict[str, Any]] = []
    for row in range(2, ws.max_row + 1):
        code = ws.cell(row, mapping.get("Code", 1)).value
        missing = [
            name
            for name in required
            if name not in mapping or ws.cell(row, mapping[name]).value in (None, "")
        ]
        if missing:
            issues.append({"row": row, "code": normalize_code(code), "missing": missing})
    return issues


def value_is_missing(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and value.strip().lower() in MISSING_TEXT_VALUES:
        return True
    return False


def numeric_value(value: Any) -> float | None:
    if value_is_missing(value):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def missing_market_quote_fields(row: dict[str, Any]) -> list[str]:
    missing: list[str] = []
    for field, label, must_be_positive in MARKET_QUOTE_FIELDS:
        numeric = numeric_value(row.get(field))
        if numeric is None or (must_be_positive and numeric <= 0):
            missing.append(label)
    return missing


def incomplete_market_row_has_diagnostics(row: dict[str, Any]) -> bool:
    status = str(row.get("接口状态") or "").strip()
    failure_reason = str(row.get("失败原因") or "").strip()
    retry_result = str(row.get("重试结果") or "").strip()
    status_is_explicit = bool(re.search(r"未刷新|失败|缺失|停牌|暂停|无成交", status))
    failure_is_explicit = failure_reason not in {"", "无", "nan", "None"}
    retry_is_explicit = retry_result not in {"", "实时行情与日K补齐可用", "nan", "None"}
    return status_is_explicit and failure_is_explicit and retry_is_explicit


def main(path: Path) -> int:
    result: dict[str, Any] = {"path": str(path.resolve()), "checks": {}, "issues": []}
    expected_suffix = re.compile(r"_\d{8}_\d{6}\.xlsx$")
    result["checks"]["timestamped_filename"] = bool(expected_suffix.search(path.name))

    with zipfile.ZipFile(path) as archive:
        bad_member = archive.testzip()
        result["checks"]["zip_integrity"] = bad_member is None
        result["checks"]["bad_zip_member"] = bad_member

    wb = load_workbook(path, data_only=False, read_only=False, keep_links=True)
    result["sheetnames"] = wb.sheetnames
    result["checks"]["target_sheets_present"] = all(name in wb.sheetnames for name in TARGET_SHEETS)
    result["checks"]["front_sheet_order_ok"] = wb.sheetnames[: len(FRONT_SHEETS)] == FRONT_SHEETS
    result["checks"]["external_links"] = len(getattr(wb, "_external_links", []))

    watch_codes = csv_codes(ROOT / "watchlist.csv")
    position_codes = csv_codes(ROOT / "positions.csv")
    expected_market = list(dict.fromkeys(watch_codes + position_codes))
    expected_counts = {
        "06_年度配置": csv_row_count(ROOT / "first_year_allocation.csv"),
        "Market_Data": len(expected_market),
        "Double_Anchor": 8,
        "03_买入候选": len(watch_codes),
        "04_阶段推进": len(watch_codes),
        "Buy_Filter": len(watch_codes),
        "Positions": len(position_codes),
        "05_持仓风险": len(position_codes),
        "长期跟踪个股": long_term_tracking_count(ROOT / "watchlist.csv"),
    }

    required = {
        "02_今日动作": [
            "优先级", "动作类型", "Code", "标的", "状态", "动作预算", "依据",
            "趋势结构", "单次仓位上限", "资产角色", "量化验证状态",
            "量化买点规则", "正期望检查", "量化风控规则",
            "阶段状态", "阶段推进结论", "阶段验证", "阶段占用率", "下一档单次上限", "阶段阻断原因",
            "仓位决策口径", "全资产当前仓位", "全资产目标仓位", "全资产剩余额度",
        ],
        "04_阶段推进": [
            "Code", "Name", "资产角色", "年度配置项", "买点灯号", "质量评分", "情绪温度",
            "阶段状态", "阶段目标下限", "阶段目标上限", "阶段占用率", "阶段验证",
            "全资产当前仓位", "全资产目标仓位", "全资产剩余额度",
            "阶段推进结论", "下一档单次上限", "阶段阻断原因", "阶段动作说明", "市场权限",
        ],
        "Market_Data": [
            "Code", "Name", "Latest", "PctChg", "Open", "High", "Low", "PrevClose",
            "Amount", "Avg20Amount", "Avg20AmountSource", "ETFShareChg", "Premium",
            "LeaderStatus", "PrevDayLow", "PrevDayLowSource", "QuoteTime", "DataSource",
            "接口状态", "失败原因", "重试结果",
        ],
        "03_买入候选": [
            "买点灯号", "Code", "Name", "建议买入区间", "建议", "阻断原因",
            "趋势结构", "单次仓位上限", "资产角色", "量化验证状态", "量化买点规则", "正期望检查", "量化风控规则",
            "阶段状态", "阶段推进结论", "阶段验证", "阶段占用率", "下一档单次上限", "阶段阻断原因",
            "机会成本检查", "认知防错",
            "接口状态", "失败原因", "重试结果", "质量状态", "评分输入状态",
            "评分缺失项", "通过项", "市场权限", "仓位决策口径", "数据状态",
        ],
        "Buy_Filter": [
            "Code", "Name", "建议买入区间", "趋势结构", "关键支撑纪律", "单次仓位上限", "爆量/量质吸纪律",
            "机会成本检查", "认知防错", "决策树情景", "赔率/最大损失", "反向失败路径", "能力圈边界",
            "资产角色", "量化验证分", "量化验证状态", "量化买点规则", "正期望检查", "量化风控规则",
            "接口状态", "失败原因", "重试结果", "评分输入状态", "评分缺失项", "通过项", "建议",
        ],
        "Quality_Score": [
            "Code", "Name", "类型", "数据完整度", "质量状态",
            "评分来源", "评分输入状态", "评分缺失项", "评分明细", "数据边界",
        ],
        "06_年度配置": [
            "配置项", "资产层级", "映射代码", "年度目标占比", "按当前全资产目标金额",
            "最新持仓金额", "年度资金缺口", "年度完成率", "目标收益下限", "目标收益上限",
            "进度状态", "执行约束", "来源",
        ],
        "Positions": [
            "Code", "Name", "Shares", "Cost", "Latest", "Latest Source", "Market Value",
            "P/L", "P/L%", "Weight", "Full Asset Weight", "Role", "Target Weight", "V2.8.5 Action",
        ],
        "05_持仓风险": [
            "风险级别", "Code", "Name", "动作建议", "触发原因",
            "全资产当前仓位", "证券账户仓位",
        ],
        "长期跟踪个股": [
            "Code", "Name", "跟踪分组", "Theme", "Latest", "PctChg", "Open",
            "High", "Low", "PrevClose", "Amount", "QuoteTime", "DataSource",
            "接口状态", "失败原因", "重试结果", "质量状态", "评分输入状态", "评分缺失项", "买点灯号",
            "阻断原因", "建议买入区间", "趋势结构", "单次仓位上限",
            "资产角色", "量化验证状态", "量化买点规则", "正期望检查", "量化风控规则",
            "机会成本检查", "反向失败路径", "仓位决策口径", "下一步",
        ],
        "Investment_Profile": [
            "Section", "Key", "Value", "Notes", "当前生效值", "维护文件",
        ],
        "Data_Sources": [
            "Module", "DataType", "PrimarySource", "FallbackSource", "FreshnessRule", "FailureBehavior", "ConfigFile",
        ],
    }

    formulas = 0
    formula_errors: list[dict[str, str]] = []
    hyperlinks: list[dict[str, str]] = []
    flagged_text: list[dict[str, str]] = []
    flags = ("未取到", "接口失败", "字段不匹配", "链接失效", "陈旧", "暂无接口", "待填")

    for sheet_name in TARGET_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        sheet_result: dict[str, Any] = {
            "data_rows": max(ws.max_row - 1, 0),
            "columns": ws.max_column,
            "freeze_panes": str(ws.freeze_panes or ""),
            "auto_filter": str(ws.auto_filter.ref or ""),
        }
        if sheet_name in expected_counts:
            sheet_result["expected_rows"] = expected_counts[sheet_name]
            sheet_result["row_count_ok"] = sheet_result["data_rows"] == expected_counts[sheet_name]
        if sheet_name in required:
            sheet_result["blank_required"] = blank_rows(ws, required[sheet_name])
        result[sheet_name] = sheet_result

        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.startswith("="):
                    formulas += 1
                    if any(token in value for token in ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A")):
                        formula_errors.append({"sheet": sheet_name, "cell": cell.coordinate, "value": value})
                if cell.hyperlink:
                    hyperlinks.append({"sheet": sheet_name, "cell": cell.coordinate, "target": str(cell.hyperlink.target)})
                if (
                    cell.row > 1
                    and isinstance(value, str)
                    and value != "未取到行情数量"
                    and any(flag in value for flag in flags)
                ):
                    flagged_text.append({"sheet": sheet_name, "cell": cell.coordinate, "value": value})

    market_ws = wb["Market_Data"]
    market_headers = headers(market_ws)
    market_codes = [normalize_code(market_ws.cell(row, market_headers["Code"]).value) for row in range(2, market_ws.max_row + 1)]
    result["checks"]["market_code_set_ok"] = set(market_codes) == set(expected_market)
    result["checks"]["market_missing_codes"] = sorted(set(expected_market) - set(market_codes))
    result["checks"]["market_extra_codes"] = sorted(set(market_codes) - set(expected_market))
    result["checks"]["market_duplicate_codes"] = sorted({code for code in market_codes if market_codes.count(code) > 1})

    quote_times: list[datetime] = []
    incomplete_market_rows: list[dict[str, Any]] = []
    incomplete_without_diagnostics: list[dict[str, Any]] = []
    for row in range(2, market_ws.max_row + 1):
        row_values = {
            name: market_ws.cell(row, column).value
            for name, column in market_headers.items()
        }
        missing_quote_fields = missing_market_quote_fields(row_values)
        if missing_quote_fields:
            item = {
                "row": row,
                "code": normalize_code(row_values.get("Code")),
                "name": row_values.get("Name"),
                "missing": missing_quote_fields,
                "接口状态": row_values.get("接口状态"),
                "失败原因": row_values.get("失败原因"),
                "重试结果": row_values.get("重试结果"),
            }
            incomplete_market_rows.append(item)
            if not incomplete_market_row_has_diagnostics(row_values):
                incomplete_without_diagnostics.append(item)
        raw = market_ws.cell(row, market_headers["QuoteTime"]).value
        if raw:
            try:
                quote_times.append(datetime.strptime(str(raw), "%Y-%m-%d %H:%M:%S"))
            except ValueError:
                result["issues"].append(f"Invalid QuoteTime at Market_Data row {row}: {raw}")
    result["quote_time_min"] = min(quote_times).isoformat(sep=" ") if quote_times else None
    result["quote_time_max"] = max(quote_times).isoformat(sep=" ") if quote_times else None
    result["checks"]["all_quotes_today"] = bool(quote_times) and all(item.date() == datetime.now().date() for item in quote_times)
    result["checks"]["market_quote_fields_complete"] = not incomplete_market_rows
    result["checks"]["market_incomplete_quote_rows"] = incomplete_market_rows
    result["checks"]["market_incomplete_quote_diagnostics_complete"] = not incomplete_without_diagnostics
    result["checks"]["market_incomplete_without_diagnostics"] = incomplete_without_diagnostics
    if incomplete_without_diagnostics:
        result["issues"].append(f"Incomplete Market_Data rows lack diagnostics: {incomplete_without_diagnostics}")

    candidate_ws = wb["03_买入候选"]
    candidate_headers = headers(candidate_ws)
    dual_basis_columns = [
        "仓位决策口径", "全资产当前仓位", "全资产目标仓位", "全资产剩余额度",
        "证券账户仓位",
    ]
    stage_columns = [
        "阶段状态", "阶段推进结论", "阶段验证", "阶段占用率", "下一档单次上限", "阶段阻断原因",
    ]
    result["checks"]["candidate_dual_basis_columns_present"] = all(
        column in candidate_headers for column in dual_basis_columns
    )
    result["checks"]["candidate_stage_columns_present"] = all(
        column in candidate_headers for column in stage_columns
    )
    missing_dual_basis_columns = [column for column in dual_basis_columns if column not in candidate_headers]
    missing_candidate_stage_columns = [column for column in stage_columns if column not in candidate_headers]
    result["checks"]["candidate_dual_basis_missing_columns"] = missing_dual_basis_columns
    result["checks"]["candidate_stage_missing_columns"] = missing_candidate_stage_columns
    if missing_dual_basis_columns:
        result["issues"].append(f"Missing dual-basis columns in 03_买入候选: {missing_dual_basis_columns}")
    if missing_candidate_stage_columns:
        result["issues"].append(f"Missing stage columns in 03_买入候选: {missing_candidate_stage_columns}")
    risk_ws = wb["05_持仓风险"]
    risk_headers = headers(risk_ws)
    risk_basis_columns = [
        "全资产当前仓位", "全资产目标仓位", "全资产剩余额度",
        "证券账户仓位",
    ]
    result["checks"]["risk_dual_basis_columns_present"] = all(
        column in risk_headers for column in risk_basis_columns
    )
    missing_risk_basis_columns = [column for column in risk_basis_columns if column not in risk_headers]
    result["checks"]["risk_dual_basis_missing_columns"] = missing_risk_basis_columns
    if missing_risk_basis_columns:
        result["issues"].append(f"Missing dual-basis columns in 05_持仓风险: {missing_risk_basis_columns}")
    missing_red_blockers = []
    blank_buy_ranges = []
    stale_numeric_ranges = []
    unsafe_blocked_numeric_ranges = []
    numeric_range_pattern = re.compile(r"\d+(?:\.\d+)?–\d+(?:\.\d+)?")
    for row in range(2, candidate_ws.max_row + 1):
        range_text = str(candidate_ws.cell(row, candidate_headers["建议买入区间"]).value or "").strip()
        if not range_text:
            blank_buy_ranges.append(row)
        has_numeric_range = bool(numeric_range_pattern.search(range_text))
        data_status = str(candidate_ws.cell(row, candidate_headers.get("数据状态", 1)).value or "")
        signal = str(candidate_ws.cell(row, candidate_headers["买点灯号"]).value or "")
        position_status = str(candidate_ws.cell(row, candidate_headers.get("仓位状态", 1)).value or "")
        blocker = str(candidate_ws.cell(row, candidate_headers["阻断原因"]).value or "")
        market_permission = str(candidate_ws.cell(row, candidate_headers.get("市场权限", 1)).value or "")
        full_position_exception = (
            range_text.startswith("暂不建议买入；")
            and (
                (("已达" in position_status and "目标" in position_status) or ("超" in position_status and "目标" in position_status))
                or ("仓位" in blocker and ("目标" in blocker or "超过" in blocker or "已达" in blocker))
            )
            and "仓位" in blocker
        )
        if has_numeric_range and data_status == "行情未刷新":
            stale_numeric_ranges.append(row)
        if (
            has_numeric_range
            and not full_position_exception
            and (
                signal not in {"绿", "黄"}
                or range_text.startswith("暂不建议买入")
                or (market_permission and market_permission != "开放买点复核")
            )
        ):
            unsafe_blocked_numeric_ranges.append(row)
        if str(candidate_ws.cell(row, candidate_headers["买点灯号"]).value) != "红":
            continue
        if blocker in (None, ""):
            missing_red_blockers.append(row)
    result["checks"]["red_candidate_blockers_complete"] = not missing_red_blockers
    result["checks"]["red_candidate_missing_blocker_rows"] = missing_red_blockers
    result["checks"]["buy_range_text_complete"] = not blank_buy_ranges
    result["checks"]["buy_range_blank_rows"] = blank_buy_ranges
    result["checks"]["stale_rows_have_no_numeric_range"] = not stale_numeric_ranges
    result["checks"]["stale_numeric_range_rows"] = stale_numeric_ranges
    result["checks"]["blocked_numeric_ranges_safe"] = not unsafe_blocked_numeric_ranges
    result["checks"]["unsafe_blocked_numeric_range_rows"] = unsafe_blocked_numeric_ranges
    if blank_buy_ranges:
        result["issues"].append(f"Blank suggested buy range text at rows: {blank_buy_ranges}")
    if stale_numeric_ranges:
        result["issues"].append(f"Stale rows contain numeric buy ranges: {stale_numeric_ranges}")
    if unsafe_blocked_numeric_ranges:
        result["issues"].append(f"Blocked rows contain unsafe numeric buy ranges: {unsafe_blocked_numeric_ranges}")

    dashboard_ws = wb["07_组合总览"]
    dashboard = {
        str(dashboard_ws.cell(row, 1).value): dashboard_ws.cell(row, 2).value
        for row in range(2, dashboard_ws.max_row + 1)
    }
    result["dashboard"] = dashboard
    result["checks"]["dashboard_missing_count"] = dashboard.get("未取到行情数量")
    result["checks"]["dashboard_missing_zero"] = dashboard.get("未取到行情数量") == 0
    result["checks"]["formula_cells"] = formulas
    result["checks"]["formula_errors"] = formula_errors
    result["checks"]["hyperlinks"] = hyperlinks
    result["checks"]["flagged_text"] = flagged_text

    print(json.dumps(result, ensure_ascii=False, indent=2, default=str))
    return 0


if __name__ == "__main__":
    if len(sys.argv) != 2:
        raise SystemExit("Usage: audit_daily_output.py <workbook.xlsx>")
    raise SystemExit(main(Path(sys.argv[1])))
