from __future__ import annotations

import json
import math
import sys
import time
import urllib.request
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_INPUT = ROOT / "output" / "V2.8.4_每日行情输出_20260617_103244_filled.xlsx"

ANCHORS = ["000001.SH", "000300.SH", "000510.CSI", "000688.SH", "399006.SZ"]
WATCH_CODES = [
    "159516",
    "562500",
    "159381",
    "159819",
    "516630",
    "588700",
    "159338",
    "512890",
    "511180",
    "511010",
    "002594",
    "688083",
    "562550",
    "159566",
    "510210",
]


def normalize_code(value: Any) -> str:
    text = str(value or "").strip()
    if text.endswith(".0"):
        text = text[:-2]
    upper = text.upper()
    if upper.endswith((".SH", ".SZ", ".CSI")):
        return upper
    if text.isdigit():
        return text.zfill(6)
    return text


def secid(code: str) -> str:
    code = normalize_code(code)
    if code == "000001.SH":
        return "1.000001"
    if code == "000300.SH":
        return "1.000300"
    if code == "000510.CSI":
        return "1.000510"
    if code == "000688.SH":
        return "1.000688"
    if code == "399006.SZ":
        return "0.399006"
    if code.startswith(("0", "1", "2", "3")):
        return f"0.{code}"
    return f"1.{code}"


def request_json(url: str) -> dict[str, Any]:
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=25) as resp:
        return json.loads(resp.read().decode("utf-8"))


def fetch_quotes(codes: list[str]) -> dict[str, dict[str, Any]]:
    fields = "f12,f14,f2,f3,f6,f15,f16,f17,f18,f152"
    url = (
        "https://push2.eastmoney.com/api/qt/ulist.np/get"
        f"?fltt=2&invt=2&fields={fields}&secids="
        + ",".join(secid(code) for code in codes)
    )
    data = request_json(url).get("data", {}).get("diff", [])
    quotes: dict[str, dict[str, Any]] = {}
    for item in data:
        raw = normalize_code(item.get("f12"))
        mapped = raw
        if raw == "000001":
            mapped = "000001.SH"
        elif raw == "000300":
            mapped = "000300.SH"
        elif raw == "000510":
            mapped = "000510.CSI"
        elif raw == "000688":
            mapped = "000688.SH"
        elif raw == "399006":
            mapped = "399006.SZ"
        quotes[mapped] = {
            "Name": item.get("f14"),
            "Latest": item.get("f2"),
            "PctChg": item.get("f3"),
            "Open": item.get("f17"),
            "High": item.get("f15"),
            "Low": item.get("f16"),
            "PrevClose": item.get("f18"),
            "Amount": item.get("f6"),
            "Premium": item.get("f152"),
        }
    return quotes


def fetch_kline_stats(code: str) -> dict[str, Any]:
    url = (
        "https://push2his.eastmoney.com/api/qt/stock/kline/get"
        f"?secid={secid(code)}&fields1=f1,f2,f3,f4,f5,f6"
        "&fields2=f51,f52,f53,f54,f55,f56,f57&klt=101&fqt=1&end=20500101&lmt=21"
    )
    try:
        klines = request_json(url).get("data", {}).get("klines") or []
        parsed = [line.split(",") for line in klines]
        amounts = [float(row[6]) for row in parsed[-20:] if len(row) > 6]
        prev_day_low = float(parsed[-2][4]) if len(parsed) >= 2 and len(parsed[-2]) > 4 else None
        return {
            "Avg20Amount": sum(amounts) / len(amounts) if amounts else None,
            "PrevDayLow": prev_day_low,
        }
    except Exception:
        return {"Avg20Amount": None, "PrevDayLow": None}


def header_map(ws) -> dict[str, int]:
    return {str(cell.value).strip(): idx for idx, cell in enumerate(ws[1], start=1) if cell.value is not None}


def set_value(ws, row: int, headers: dict[str, int], key: str, value: Any) -> None:
    if key in headers:
        ws.cell(row, headers[key]).value = value


def as_float(value: Any) -> float | None:
    try:
        if value is None or (isinstance(value, float) and math.isnan(value)):
            return None
        return float(value)
    except Exception:
        return None


def intraday_position(latest: Any, low: Any, high: Any) -> float | None:
    latest_f = as_float(latest)
    low_f = as_float(low)
    high_f = as_float(high)
    if latest_f is None or low_f is None or high_f is None or high_f <= low_f:
        return None
    return round((latest_f - low_f) / (high_f - low_f), 4)


def volume_ratio(amount: Any, avg20: Any) -> float | None:
    amount_f = as_float(amount)
    avg20_f = as_float(avg20)
    if amount_f is None or avg20_f is None or avg20_f <= 0:
        return None
    return round(amount_f / avg20_f, 2)


def buy_filter_row(q: dict[str, Any]) -> dict[str, Any]:
    latest = q.get("Latest")
    pct = as_float(q.get("PctChg"))
    open_price = as_float(q.get("Open"))
    prev_close = as_float(q.get("PrevClose"))
    high = q.get("High")
    low = q.get("Low")
    pos = intraday_position(latest, low, high)
    vr = volume_ratio(q.get("Amount"), q.get("Avg20Amount"))
    veto = "否"
    veto_reason = ""
    if open_price is not None and prev_close is not None and pos is not None:
        if open_price > prev_close and pos <= 0.2:
            veto = "是"
            veto_reason = "高开后接近日内低位"
    pass_count = 0
    if pos is not None and 0.35 <= pos <= 0.75:
        pass_count += 1
    if pct is not None and -1.0 <= pct <= 3.0:
        pass_count += 1
    if vr is not None and 0.7 <= vr <= 1.8:
        pass_count += 1
    if pct is not None and pct >= 0 and pos is not None and pos < 0.85:
        pass_count += 1
    if veto == "是":
        suggestion = "不买/观察"
    elif pass_count >= 3:
        suggestion = "可进入买点复核"
    elif pass_count >= 2:
        suggestion = "可观察（未达完整过滤器）"
    else:
        suggestion = "不买/观察"
    if pos is None:
        structure = "缺少日内数据"
    elif pos >= 0.85:
        structure = "接近日内高位"
    elif pos <= 0.25:
        structure = "接近日内低位"
    else:
        structure = "日内中位震荡"
    if vr is None:
        volume_text = "缺少20日均额"
    elif vr >= 1.8:
        volume_text = "显著放量"
    elif vr >= 0.8:
        volume_text = "量能正常"
    else:
        volume_text = "缩量"
    if pct is None:
        price_volume = "缺少涨跌幅"
    elif pct > 0 and (vr or 0) <= 1.8:
        price_volume = "温和上涨"
    elif pct > 0:
        price_volume = "放量上涨，防追高"
    elif pct < 0 and (vr or 0) >= 1.2:
        price_volume = "放量下跌"
    else:
        price_volume = "弱势震荡"
    return {
        "日内位置": pos,
        "量能倍数": vr,
        "分时结构": structure,
        "量价关系": price_volume,
        "份额变动": "暂无接口",
        "龙头成分": "待人工确认",
        "次日验证": "待次日确认",
        "通过项": pass_count,
        "一票否决": veto,
        "否决原因": veto_reason,
        "建议": suggestion,
        "量能说明": volume_text,
    }


def anchor_light(q: dict[str, Any]) -> tuple[str, str]:
    latest = as_float(q.get("Latest"))
    pct = as_float(q.get("PctChg"))
    low = as_float(q.get("Low"))
    prev_low = as_float(q.get("PrevDayLow"))
    if latest is None or pct is None:
        return "待确认", "缺少最新价或涨跌幅"
    if prev_low is not None and latest < prev_low:
        return "红灯", "最新价跌破昨日低点"
    if pct <= -1.5:
        return "黄灯偏弱", "指数跌幅超过1.5%"
    if low is not None and latest > low and pct > 0:
        return "黄绿", "当日未创新低，且涨跌幅为正"
    return "观察", "未触发黄绿、黄灯偏弱或红灯"


def summarize_group(lights: list[str]) -> str:
    if any(x == "红灯" for x in lights):
        return "红灯"
    if any(x == "黄灯偏弱" for x in lights):
        return "黄灯偏弱"
    if any(x == "黄绿" for x in lights):
        return "黄绿"
    return "观察"


def complete_workbook(path: Path) -> Path:
    codes = list(dict.fromkeys(ANCHORS + WATCH_CODES))
    quotes = fetch_quotes(codes)
    for code in codes:
        quotes.setdefault(code, {})
        quotes[code].update(fetch_kline_stats(code))
        time.sleep(0.03)

    wb = load_workbook(path)

    if "Market_Data" in wb.sheetnames:
        ws = wb["Market_Data"]
        headers = header_map(ws)
        for row in range(2, ws.max_row + 1):
            code = normalize_code(ws.cell(row, headers["Code"]).value)
            q = quotes.get(code)
            if not q:
                continue
            for key in ["Latest", "PctChg", "Open", "High", "Low", "PrevClose", "Amount", "Avg20Amount", "Premium", "PrevDayLow"]:
                set_value(ws, row, headers, key, q.get(key))
            set_value(ws, row, headers, "DataSource", "东方财富补齐")
            set_value(ws, row, headers, "ETFShareChg", "暂无接口")
            set_value(ws, row, headers, "LeaderStatus", "待人工确认")
            note = ws.cell(row, headers["Notes"]).value if "Notes" in headers else ""
            note_text = str(note or "").replace("；未取到行情，使用券商截图价格", "").replace("；未取到行情", "")
            set_value(ws, row, headers, "Notes", (note_text + "；行情已补齐").strip("；"))

    buy_sheet = "03_买入候选" if "03_买入候选" in wb.sheetnames else "Buy_Filter"
    if buy_sheet in wb.sheetnames:
        ws = wb[buy_sheet]
        headers = header_map(ws)
        for row in range(2, ws.max_row + 1):
            code = normalize_code(ws.cell(row, headers["Code"]).value)
            q = quotes.get(code)
            if not q:
                continue
            for key in ["Latest", "PctChg"]:
                set_value(ws, row, headers, key, q.get(key))
            result = buy_filter_row(q)
            for key in ["日内位置", "量能倍数", "分时结构", "量价关系", "份额变动", "龙头成分", "次日验证", "通过项", "一票否决", "否决原因", "建议"]:
                set_value(ws, row, headers, key, result.get(key))

    risk_sheet = "04_持仓风险" if "04_持仓风险" in wb.sheetnames else "Positions_Action"
    if risk_sheet in wb.sheetnames:
        ws = wb[risk_sheet]
        headers = header_map(ws)
        for row in range(2, ws.max_row + 1):
            code = normalize_code(ws.cell(row, headers["Code"]).value)
            q = quotes.get(code)
            if not q:
                continue
            for key in ["Latest", "PctChg", "High", "Low", "Amount", "Avg20Amount"]:
                set_value(ws, row, headers, key, q.get(key))

    if "Double_Anchor" in wb.sheetnames:
        ws = wb["Double_Anchor"]
        headers = header_map(ws)
        total_lights: list[str] = []
        growth_lights: list[str] = []
        for row in range(2, ws.max_row + 1):
            code_cell = ws.cell(row, headers.get("Code", 2)).value
            code = normalize_code(code_cell)
            if code not in quotes:
                continue
            q = quotes[code]
            light, note = anchor_light(q)
            set_value(ws, row, headers, "最新价", q.get("Latest"))
            set_value(ws, row, headers, "涨跌幅", q.get("PctChg"))
            set_value(ws, row, headers, "当日最低", q.get("Low"))
            set_value(ws, row, headers, "昨日低点", q.get("PrevDayLow"))
            set_value(ws, row, headers, "灯号", light)
            set_value(ws, row, headers, "说明", note)
            anchor_type = str(ws.cell(row, headers.get("锚点", 1)).value or "")
            if "总量" in anchor_type:
                total_lights.append(light)
            if "成长" in anchor_type:
                growth_lights.append(light)
        total = summarize_group(total_lights)
        growth = summarize_group(growth_lights)
        combined = "双绿" if total == "黄绿" and growth == "黄绿" else ("成长红" if growth == "红灯" else "观察")
        for row in range(2, ws.max_row + 1):
            anchor_type = str(ws.cell(row, headers.get("锚点", 1)).value or "")
            if anchor_type == "总量锚综合":
                set_value(ws, row, headers, "灯号", total)
            elif anchor_type == "成长锚综合":
                set_value(ws, row, headers, "灯号", growth)
            elif anchor_type == "综合灯号":
                set_value(ws, row, headers, "灯号", combined)
                set_value(ws, row, headers, "说明", "允许研究买入，但仍需买点过滤器" if combined == "双绿" else "继续观察")

    for sheet_name in ["Market_Data", buy_sheet, risk_sheet, "Double_Anchor"]:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for col in ws.columns:
                header = str(col[0].value or "")
                if header in {"Amount", "Avg20Amount"}:
                    for cell in col[1:]:
                        cell.number_format = "#,##0"
                elif header in {"PctChg", "涨跌幅", "量能倍数"}:
                    for cell in col[1:]:
                        cell.number_format = "0.00"
                elif header in {"Latest", "Open", "High", "Low", "PrevClose", "最新价", "当日最低", "昨日低点"}:
                    for cell in col[1:]:
                        cell.number_format = "0.000"
                ws.column_dimensions[col[0].column_letter].width = max(ws.column_dimensions[col[0].column_letter].width or 10, min(18, max(10, len(header) + 4)))

    output = path.with_name(path.stem.replace("_filled", "_complete") + path.suffix)
    wb.save(output)
    return output


def main() -> int:
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_INPUT
    output = complete_workbook(path)
    print(output)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
